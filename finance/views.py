from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Q
from django.utils import timezone
from django.http import JsonResponse
import json
import datetime
from .models import Transaction, Category, Source
from .forms import TransactionForm, CategoryForm
from branches.models import LocationSite
from accounts.views import get_admin_id
from accounts.date_utils import today_ist
from accounts.cycle_utils import get_salary_cycle


# Marker prefix that attendance/signals.py stamps on every auto-generated
# salary expense. Grep target for the "salary vs non-salary" split.
AUTO_SAL_PREFIX = '[AUTO-SAL:'


def _cycle_label(start, end):
    """Human label like '26 Jul → 25 Aug 2026'. Same year → year printed once."""
    if start.year == end.year:
        return f"{start.day:02d} {start.strftime('%b')} → {end.day:02d} {end.strftime('%b %Y')}"
    return (
        f"{start.day:02d} {start.strftime('%b %Y')} → "
        f"{end.day:02d} {end.strftime('%b %Y')}"
    )


def _available_cycles(n=8):
    """Return the last N salary cycles (most-recent first, including current).

    Each entry: {'start': date, 'end': date, 'label': str, 'key': 'YYYY-MM-DD_YYYY-MM-DD'}
    Reused by the top filter dropdown and every per-card cycle picker.
    """
    from dateutil.relativedelta import relativedelta
    out = []
    current = get_salary_cycle(today_ist())
    ref = current['start']
    for _ in range(n):
        c = get_salary_cycle(ref)
        out.append({
            'start': c['start'],
            'end':   c['end'],
            'label': _cycle_label(c['start'], c['end']),
            'key':   f"{c['start'].isoformat()}_{c['end'].isoformat()}",
        })
        ref = c['start'] - datetime.timedelta(days=1)
    return out


def _resolve_active_cycle(request):
    """Pick the active cycle from ?cycle=<start_iso>_<end_iso> or fall back
    to the current cycle (today IST). Returns (start, end, label)."""
    raw = (request.GET.get('cycle') or '').strip()
    if raw and '_' in raw:
        try:
            s, e = raw.split('_', 1)
            start = datetime.date.fromisoformat(s)
            end   = datetime.date.fromisoformat(e)
            return start, end, _cycle_label(start, end)
        except (ValueError, TypeError):
            pass
    current = get_salary_cycle(today_ist())
    return current['start'], current['end'], _cycle_label(current['start'], current['end'])


def _location_sites_json(user):
    admin_id = get_admin_id(user)
    names = list(LocationSite.objects.filter(admin_id=admin_id).values_list('name', flat=True))
    return json.dumps(names)


def _shared_sources_json(user):
    admin_id = get_admin_id(user)
    names = list(Source.objects.filter(admin_id=admin_id, type='income').values_list('name', flat=True))
    return json.dumps(names)


def _shared_accounts_json(user):
    admin_id = get_admin_id(user)
    names = list(Source.objects.filter(admin_id=admin_id, type='account').values_list('name', flat=True))
    return json.dumps(names)


def _balance_by_combo_expense(user):
    """Return balance per (Income Source, Account) for the expense page.
    Keys are normalised to lower-case so matching is case-insensitive; the display
    values ('source', 'account') keep the original casing from the income record.

    Defensive: each cross-module query is wrapped + uses `.only(...)` so a
    transient migration mismatch on either side doesn't 500 the page.
    """
    admin_id = get_admin_id(user)
    combo_map = {}

    try:
        from income.models import Income
        income_qs = Income.objects.filter(admin_id=admin_id).only(
            'amount', 'payment_by', 'payment_mode',
        )
        for i in income_qs:
            source  = (i.payment_by or '').strip()
            account = (i.payment_mode or '').strip()
            if source and account:
                k = f"{source.lower()}|||{account.lower()}"
                if k not in combo_map:
                    combo_map[k] = {'source': source, 'account': account, 'income': 0, 'expense': 0}
                combo_map[k]['income'] += float(i.amount)
    except Exception:
        pass

    try:
        expense_qs = Transaction.objects.filter(admin_id=admin_id, type='expense').only(
            'amount', 'income_source', 'payment_mode',
        )
        for e in expense_qs:
            source  = (e.income_source or '').strip()
            account = (e.payment_mode or '').strip()
            if source and account:
                k = f"{source.lower()}|||{account.lower()}"
                if k not in combo_map:
                    combo_map[k] = {'source': source, 'account': account, 'income': 0, 'expense': 0}
                combo_map[k]['expense'] += float(e.amount)
    except Exception:
        pass

    result = []
    for v in combo_map.values():
        v['balance'] = v['income'] - v['expense']
        result.append(v)
    return sorted(result, key=lambda x: (-abs(x['balance']), x['source']))


def _is_ajax(request):
    return (
        request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        or 'application/json' in request.headers.get('Accept', '')
    )


# Fixed 8-slot accent palette — assigned per site by hashing the site name so
# the same site always renders with the same border accent in both the Expense
# Manager and Income module cards. Kept readable on the light theme.
SITE_ACCENT_PALETTE = [
    '#2563eb', '#10b981', '#f59e0b', '#ef4444',
    '#8b5cf6', '#06b6d4', '#ec4899', '#84cc16',
]


def _site_color_index(name):
    """Stable 0..7 slot from a site name for the card border accent."""
    n = (name or '').strip().lower()
    if not n:
        return 0
    total = 0
    for ch in n:
        total += ord(ch)
    return total % len(SITE_ACCENT_PALETTE)


def _group_expenses_by_site(expenses, user, seed_names=None):
    """
    Site-based card view for the Expense page (Income/Expense restructure).

    Groups expense Transactions by `location_site` (case-insensitive). For
    each site returns credit (sum of Income amounts for the SAME site —
    pulled live from income.Income, never duplicated), debit (sum of
    Transaction amounts), and balance = credit - debit. Sort order: most
    recent expense activity first.

    Sites that have income but zero expense are still surfaced so the
    admin can see the credit balance even before any expense is logged.

    Defensive: the cross-module Income query is wrapped in try/except so a
    transient Income-side failure (e.g. mid-deploy when one migration has
    run and the other hasn't) degrades to an empty credit map instead of
    blowing up the entire Expense page.
    """
    import datetime as _dt
    from decimal import Decimal as _D
    from accounts.views import get_admin_id
    admin_id = get_admin_id(user)

    # Normalise the (raw -> key, display) for a site value so the "(No Site)"
    # bucket collapses every shape an empty/no-site record can take: NULL,
    # empty string, whitespace, or the literal string "(No Site)" that
    # earlier site-detail Add-Row submissions wrote back into location_site.
    # This is what fixes the "two (No Site) cards" symptom — the grouping
    # key is now identical regardless of how the row was created.
    def _site_key_display(raw):
        s = (raw or '').strip()
        if not s or s.lower() == '(no site)':
            return '', '(No Site)'
        return s.lower(), s

    def _cycle_key_for(d):
        """Cycle bucket key for a date. Uses SPIM 26→25 cycle month_key
        ('YYYY-MM' of the cycle's END month) so per-card breakdown
        aggregates by payroll cycle instead of calendar month."""
        try:
            return get_salary_cycle(d)['month_key']
        except Exception:
            return d.strftime('%Y-%m')

    # Credit map: site → total income amount for that tenant. Built in a
    # single sweep so site_card render stays O(N) over expenses + incomes.
    # credit_by_cycle[site_key][cycle_key] = Decimal — feeds the card-level
    # cycle dropdown so each card can show per-cycle Credit/Debit/Balance.
    credit_map  = {}
    site_disp   = {}
    income_dates = {}
    credit_by_cycle = {}
    try:
        from income.models import Income
        income_qs = Income.objects.filter(admin_id=admin_id).only(
            'amount', 'location_site', 'date',
        )
        for i in income_qs:
            key, disp = _site_key_display(i.location_site)
            credit_map[key] = credit_map.get(key, _D('0')) + (i.amount or _D('0'))
            if key not in site_disp or not site_disp[key]:
                site_disp[key] = disp
            if i.date:
                ckey = _cycle_key_for(i.date)
                cb = credit_by_cycle.setdefault(key, {})
                cb[ckey] = cb.get(ckey, _D('0')) + (i.amount or _D('0'))
                prev = income_dates.get(key)
                if prev is None or i.date > prev:
                    income_dates[key] = i.date
    except Exception:
        # Income table unavailable (e.g. migration not yet applied). Render
        # the Expense page with zero credit rather than 500-ing.
        credit_map  = {}
        site_disp   = {}
        income_dates = {}
        credit_by_cycle = {}

    debit_by_cycle = {}
    groups = {}

    # Seed with the canonical site list from Projects/Attendance/WorkLog so
    # sites with zero transactions still render as editable zero-row cards.
    # Projects.Site casing wins for display — matches the source-of-truth rule.
    seeded_keys = set()
    if seed_names:
        for n in seed_names:
            key = (n or '').strip().lower()
            if not key or key in groups:
                continue
            seeded_keys.add(key)
            site_disp[key] = n
            groups[key] = {
                'site':         n,
                'site_key':     key,
                'credit':       credit_map.get(key, _D('0')),
                'debit':        _D('0'),
                'balance':      _D('0'),
                'count':        0,
                'latest_date':  income_dates.get(key),
                'transactions': [],
            }

    for e in expenses:
        key, disp = _site_key_display(getattr(e, 'location_site', None))
        g = groups.get(key)
        if g is None:
            g = {
                'site':         site_disp.get(key) or disp,
                'site_key':     key,
                'credit':       credit_map.get(key, _D('0')),
                'debit':        _D('0'),
                'balance':      _D('0'),
                'count':        0,
                'latest_date':  e.date,
                'transactions': [],
            }
            groups[key] = g
        g['debit'] += (e.amount or _D('0'))
        g['count'] += 1
        g['transactions'].append(e)
        if e.date:
            ckey = _cycle_key_for(e.date)
            db = debit_by_cycle.setdefault(key, {})
            db[ckey] = db.get(ckey, _D('0')) + (e.amount or _D('0'))
        if e.date and (g['latest_date'] is None or e.date > g['latest_date']):
            g['latest_date'] = e.date
            # Don't overwrite the canonical display with a blank — keep the
            # first non-empty real-site spelling, otherwise leave "(No Site)".
            # Also never overwrite when the group came from a seed (Projects.Site
            # casing wins over whatever ad-hoc string the expense row carries).
            if key and key not in seeded_keys:
                g['site'] = disp

    # Surface income-only sites (credit > 0, no expense yet) so the card
    # list reflects every site the tenant is tracking — admin still needs
    # to see the positive balance.
    for key, credit in credit_map.items():
        if key in groups:
            continue
        groups[key] = {
            'site':         site_disp.get(key) or '(No Site)',
            'site_key':     key,
            'credit':       credit,
            'debit':        _D('0'),
            'balance':      credit,
            'count':        0,
            'latest_date':  income_dates.get(key),
            'transactions': [],
        }

    today = today_ist()
    current_cycle = get_salary_cycle(today)
    current_cycle_key = current_cycle['month_key']

    # Preload a lookup of cycle_key → (start, end, label) covering the last
    # 24 cycles so option labels don't recompute per-card. 24 keeps two
    # full years addressable in the picker.
    cycle_meta = {}
    for c in _available_cycles(n=24):
        cs = c['start']
        ce = c['end']
        cycle_meta[ce.strftime('%Y-%m')] = {
            'start': cs, 'end': ce, 'label': c['label'],
        }

    for g in groups.values():
        g['balance'] = g['credit'] - g['debit']
        g['transactions'].sort(
            key=lambda x: (x.date or _dt.date.min),
            reverse=True,
        )
        # Per-site cycle breakdown — fed to the card via data-monthly so
        # the card's cycle dropdown can repaint Credit/Debit/Balance tiles
        # without a round-trip. Key kept as 'YYYY-MM' (the cycle END month)
        # so the existing repaint JS works unchanged.
        ck = g['site_key']
        cm = credit_by_cycle.get(ck, {})
        dm = debit_by_cycle.get(ck, {})
        all_cycles = set(cm.keys()) | set(dm.keys())
        all_cycles.add(current_cycle_key)
        by_cycle = {}
        for m in all_cycles:
            c = cm.get(m, _D('0'))
            d = dm.get(m, _D('0'))
            by_cycle[m] = {
                'credit':  float(c),
                'debit':   float(d),
                'balance': float(c - d),
            }
        # `monthly_json` name preserved so existing card-repaint JS keeps
        # working; the values inside are now per-cycle aggregates.
        g['monthly_json'] = json.dumps(by_cycle)
        g['totals_json']  = json.dumps({
            'credit':  float(g['credit']),
            'debit':   float(g['debit']),
            'balance': float(g['balance']),
        })
        # Pretty cycle options for the dropdown — sorted most-recent first.
        opts = []
        for m in sorted(all_cycles, reverse=True):
            meta = cycle_meta.get(m)
            if meta:
                label = meta['label']
            else:
                # Fall back to computing on-demand for cycles older than
                # the 24-cycle preload window.
                try:
                    y, mo = m.split('-')
                    end = datetime.date(int(y), int(mo), 25)
                    start = get_salary_cycle(end)['start']
                    label = _cycle_label(start, end)
                except Exception:
                    label = m
            opts.append({'key': m, 'label': label})
        g['month_options'] = opts
        # Per-site accent for the card border (Issue 4 — color palette).
        g['color_index'] = _site_color_index(g['site'])

    return sorted(
        groups.values(),
        key=lambda g: (g['latest_date'] or _dt.date.min),
        reverse=True,
    )


def _dynamic_site_options(admin_id, month_yyyy_mm=None, exact_date=None):
    """Site NAMES to surface in the Expense-page site filter dropdown.

    Scope is EITHER a single date (`exact_date`, higher priority) OR a
    month string 'YYYY-MM'. Union of names from:
      * attendance.AttendanceRecord.site (CharField, mobile back-compat)
      * attendance.AttendanceRecord.site_ref.name (FK on projects.Site)
      * projects.WorkLog.site.name (FK on projects.Site)
      * distinct non-empty Transaction.location_site strings (legacy)

    Returns [{'name': <display>}] deduped on lower(name), sorted by name.
    Returns [] when scope is empty (no date, no month).
    """
    if not exact_date and not month_yyyy_mm:
        return []

    def _scope_filter(qs, date_field):
        if exact_date is not None:
            return qs.filter(**{date_field: exact_date})
        return qs.filter(**{f"{date_field}__startswith": month_yyyy_mm})

    names = set()

    try:
        from attendance.models import AttendanceRecord
        att_qs = _scope_filter(
            AttendanceRecord.objects.filter(admin_id=admin_id),
            'date',
        )
        for n in att_qs.exclude(site__isnull=True).exclude(site='') \
                       .values_list('site', flat=True).distinct():
            if n:
                names.add(n.strip())
        for n in att_qs.filter(site_ref__isnull=False) \
                       .values_list('site_ref__name', flat=True).distinct():
            if n:
                names.add(n.strip())
    except Exception:  # noqa: BLE001
        pass

    try:
        from projects.models import WorkLog
        for n in _scope_filter(
            WorkLog.objects.filter(admin_id=admin_id, site__isnull=False),
            'date',
        ).values_list('site__name', flat=True).distinct():
            if n:
                names.add(n.strip())
    except Exception:  # noqa: BLE001
        pass

    for n in _scope_filter(
        Transaction.objects.filter(admin_id=admin_id), 'date',
    ).exclude(location_site__isnull=True).exclude(location_site='') \
     .values_list('location_site', flat=True).distinct():
        if n:
            names.add(n.strip())

    dedup = {}
    for n in names:
        key = n.lower()
        if key and key not in dedup:
            dedup[key] = n
    return sorted(
        ({'name': v} for v in dedup.values()),
        key=lambda o: o['name'].lower(),
    )


def _group_expenses_by_account(qs):
    """
    Group an already-filtered Transaction queryset by `payment_mode` (the
    column that holds the Account name, e.g. "MURUGAN NALLAKANNU").

    Returns a list of dicts, sorted by latest_date descending:
        {
            'account':      <display name, casing of the most recent row>,
            'account_key':  <lower-cased account, the dedup key>,
            'total':        Decimal — sum of all amounts in this group,
            'count':        int — number of transactions,
            'latest_date':  date — most recent transaction date,
            'transactions': [Transaction, ...] — sorted by date desc,
        }

    Records with a blank account fall under the "(No Account)" bucket so
    they are still visible to the admin (instead of silently disappearing).
    Lower-cased keys mirror the case-insensitive matching used by
    `_balance_by_combo_expense` above — two rows with "Cash" / "CASH"
    collapse into one card.
    """
    import datetime as _dt
    from decimal import Decimal as _D
    groups = {}
    for t in qs:
        acc = (getattr(t, 'payment_mode', None) or '').strip()
        key = acc.lower()
        g = groups.get(key)
        if g is None:
            g = {
                'account':      acc or '(No Account)',
                'account_key':  key,
                'total':        _D('0'),
                'count':        0,
                'latest_date':  t.date,
                'transactions': [],
            }
            groups[key] = g
        g['total'] += t.amount
        g['count'] += 1
        g['transactions'].append(t)
        if t.date and (g['latest_date'] is None or t.date > g['latest_date']):
            g['latest_date'] = t.date
            # Refresh display casing to the most recent row's spelling.
            g['account'] = acc or '(No Account)'
    for g in groups.values():
        g['transactions'].sort(
            key=lambda x: (x.date or _dt.date.min),
            reverse=True,
        )
    return sorted(
        groups.values(),
        key=lambda g: (g['latest_date'] or _dt.date.min),
        reverse=True,
    )


def _expense_combo_data(t, user):
    """Return the combo balance dict for this transaction's (income_source, account) pair.
    Comparison is case-insensitive to match the normalised _balance_by_combo_expense keys."""
    src = (t.income_source or '').strip().lower()
    acc = (t.payment_mode or '').strip().lower()
    if not (src and acc):
        return None
    for c in _balance_by_combo_expense(user):
        if c['source'].lower() == src and c['account'].lower() == acc:
            return c
    return None


def _expense_to_dict(t):
    return {
        'id': t.pk,
        'date': t.date.isoformat() if t.date else '',
        'date_display': t.date.strftime('%d %b, %Y') if t.date else '',
        'amount': str(t.amount),
        'amount_display': '{:,.2f}'.format(float(t.amount)),
        'category': t.category.name if t.category else 'General',
        'category_id': t.category_id,
        # Fixed-choice category for the site-detail panel (Income/Expense
        # restructure). Stored on the row as a code ('food', 'fuel', …);
        # the display label is resolved via Transaction.get_expense_category_display().
        'expense_category':         getattr(t, 'expense_category', '') or '',
        'expense_category_display': t.get_expense_category_display() if getattr(t, 'expense_category', '') else '',
        'expense_type': t.purpose or '',
        'location': t.location_site or '',
        'payment_by': t.payment_by or '',
        'payment_to': t.vendor or '',
        'payment_mode': t.payment_mode or '',
        'income_source': t.income_source or '',
        'description': t.description or '',
        'edit_url': f'/expenses/{t.pk}/edit/',
        'delete_url': f'/expenses/{t.pk}/delete/',
    }

FILTER_PARAM_KEYS = (
    'date', 'month', 'from_date', 'to_date',
    'site', 'income_source', 'category', 'search', 'account',
)


def _clean_multi(request, key):
    """Read a multi-value querystring param. Accepts ?key=X&key=Y (native
    <select multiple>) and also ?key=X,Y (comma-separated). Empty entries
    stripped. Returns a list of non-empty strings."""
    raw = request.GET.getlist(key)
    out = []
    for item in raw:
        if item is None:
            continue
        for part in str(item).split(','):
            p = part.strip()
            if p:
                out.append(p)
    # Dedup preserving order (case-insensitive for string-name filters).
    seen = set()
    dedup = []
    for x in out:
        k = x.lower()
        if k in seen:
            continue
        seen.add(k)
        dedup.append(x)
    return dedup


def _parse_iso_date(raw):
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    try:
        return datetime.date.fromisoformat(s)
    except (ValueError, TypeError):
        return None


def _apply_expense_filters(request, admin_id):
    """
    Parse the Expense page's querystring filters and apply them to a
    Transaction queryset. Returns (qs, filters, has_explicit_filter,
    exact_date, month, today_default_active).

    Multi-value filters (site / category / account / income_source): read
    with getlist so ?site=A&site=B applies OR-within-field. AND across
    fields (chained .filter calls).

    Date precedence: from_date/to_date range takes precedence over the
    month picker. If either is set, month is ignored.

    Rule: if the user has set ANY filter param, honour it verbatim. If not
    (fresh landing on /expenses/), narrow to today's active sites — a soft
    default the moment any filter is touched.
    """
    qs = (
        Transaction.objects
        .filter(admin_id=admin_id, type='expense')
        .select_related('category', 'branch')
    )

    cat_ids        = _clean_multi(request, 'category')
    accounts       = _clean_multi(request, 'account')
    income_sources = _clean_multi(request, 'income_source')
    sites          = _clean_multi(request, 'site')

    month     = (request.GET.get('month') or '').strip()
    search    = (request.GET.get('search') or '').strip()
    from_date = _parse_iso_date(request.GET.get('from_date'))
    to_date   = _parse_iso_date(request.GET.get('to_date'))

    # from/to takes precedence over month.
    if from_date or to_date:
        month = ''

    # Invalid range (to < from) — ignore both and surface an error via filters.
    date_range_error = ''
    if from_date and to_date and to_date < from_date:
        date_range_error = 'To date is before From date — date range ignored.'
        from_date = None
        to_date = None

    has_explicit_filter = any((
        cat_ids, accounts, income_sources, sites,
        month, search, from_date, to_date,
        (request.GET.get('date') or '').strip(),
    ))

    exact_date = _parse_iso_date(request.GET.get('date'))

    # Cycle scope — the fresh default and the value the top-of-page
    # dropdown drives. Precedence order for the date window:
    #   1. exact ?date=YYYY-MM-DD  (already handled below)
    #   2. ?from_date / ?to_date   (manual override)
    #   3. ?month                  (legacy scalar override)
    #   4. ?cycle=START_END        (explicit cycle pick)
    #   5. current cycle (fallback default)
    cycle_start, cycle_end, cycle_label = _resolve_active_cycle(request)
    cycle_scope_active = False

    if cat_ids:
        qs = qs.filter(category_id__in=cat_ids)

    if exact_date:
        qs = qs.filter(date=exact_date)
    elif from_date or to_date:
        if from_date:
            qs = qs.filter(date__gte=from_date)
        if to_date:
            qs = qs.filter(date__lte=to_date)
    elif month:
        qs = qs.filter(date__startswith=month)
    else:
        # No manual date narrowing → apply the salary cycle window.
        qs = qs.filter(date__gte=cycle_start, date__lte=cycle_end)
        cycle_scope_active = True

    # Legacy alias — the template gates the banner on `today_default_active`.
    # It now means "cycle scope is in effect" (fresh landing or explicit
    # cycle pick). Renamed in-context below without breaking the template.
    today_default_active = cycle_scope_active

    if search:
        qs = qs.filter(
            Q(description__icontains=search) | Q(vendor__icontains=search) | Q(reference__icontains=search)
        )

    if accounts:
        aq = Q()
        for a in accounts:
            aq |= Q(payment_mode__iexact=a)
        qs = qs.filter(aq)

    if income_sources:
        sq = Q()
        for s in income_sources:
            sq |= Q(income_source__iexact=s)
        qs = qs.filter(sq)

    if sites:
        sq = Q()
        for s in sites:
            if s == 'UNASSIGNED':
                sq |= Q(location_site__isnull=True) | Q(location_site='')
            else:
                sq |= Q(location_site__iexact=s)
        qs = qs.filter(sq)

    filters = {
        # Multi-value lists — template iterates these for `selected` state.
        'categories':    cat_ids,
        'accounts':      accounts,
        'income_sources': income_sources,
        'sites':         sites,
        # Scalars.
        'month':         month,
        'search':        search,
        'from_date':     from_date.isoformat() if from_date else '',
        'to_date':       to_date.isoformat() if to_date else '',
        'date':          exact_date.isoformat() if exact_date else '',
        'date_range_error': date_range_error,
        'cycle':         f"{cycle_start.isoformat()}_{cycle_end.isoformat()}",
        # Back-compat scalar aliases — legacy template bits still read these
        # (e.g. the Grouped-by-Site "no expenses for {{ filters.date }}"
        # empty-state copy). First selected value if present, else ''.
        'category':      cat_ids[0] if cat_ids else '',
        'account':       accounts[0] if accounts else '',
        'income_source': income_sources[0] if income_sources else '',
        'site':          sites[0] if sites else '',
    }
    return (qs, filters, has_explicit_filter, exact_date, month,
            today_default_active, cycle_start, cycle_end, cycle_label)


def _all_site_options(admin_id):
    """Canonical site names for the Expense filter dropdown.

    Drawn from projects.Site ∪ Attendance-used sites via
    projects.utils.sites_for_admin — legacy Transaction.location_site
    orphans are deliberately excluded. Returned as [{'name': <display>}]
    sorted A→Z (already deduped case-insensitively by the helper).
    """
    from projects.utils import sites_for_admin
    return [{'name': n} for n in sites_for_admin(admin_id)]


@login_required
def transaction_list(request):
    admin_id = get_admin_id(request.user)
    (qs, filters, has_explicit_filter, exact_date, month,
     today_default_active, cycle_start, cycle_end, cycle_label) = \
        _apply_expense_filters(request, admin_id)

    # Total Expenses / Balance Summary — cycle-scoped to the SAME window
    # (get_salary_cycle(today_ist())) as the This-Cycle split below and
    # as the Income Manager KPI tiles. Keeps every currency KPI across
    # the suite on the same 26→25 payroll cycle so numbers agree.
    from accounts.cycle_utils import get_salary_cycle as _kpi_get_cycle
    from accounts.date_utils import today_ist as _kpi_today
    _kpi_cycle = _kpi_get_cycle(_kpi_today())
    _kpi_start, _kpi_end = _kpi_cycle['start'], _kpi_cycle['end']
    total_expense = (
        Transaction.objects
        .filter(admin_id=admin_id, type='expense',
                date__gte=_kpi_start, date__lte=_kpi_end)
        .aggregate(t=Sum('amount'))['t'] or 0
    )

    # Total Income (same cycle window) — feeds the Balance Summary box so
    # its numbers stay consistent with the Total Expenses card above.
    # Defensive: a transient Income-side schema mismatch falls back to 0.
    try:
        from income.models import Income as _Income
        total_income = (
            _Income.objects
            .filter(admin_id=admin_id, date__gte=_kpi_start, date__lte=_kpi_end)
            .aggregate(t=Sum('amount'))['t'] or 0
        )
    except Exception:
        total_income = 0
    net_balance = (total_income or 0) - (total_expense or 0)

    # This-cycle split — salary vs non-salary. The cycle window is
    # already applied to `qs`; split by the AUTO-SAL marker prefix.
    salary_qs     = qs.filter(reference__startswith=AUTO_SAL_PREFIX)
    non_salary_qs = qs.exclude(reference__startswith=AUTO_SAL_PREFIX)
    total_salary_expense     = salary_qs.aggregate(t=Sum('amount'))['t'] or 0
    total_non_salary_expense = non_salary_qs.aggregate(t=Sum('amount'))['t'] or 0

    balance_combos = _balance_by_combo_expense(request.user)
    # Keyed by lower-case (source, account) so lookup is case-insensitive.
    combo_map = {(c['source'].lower(), c['account'].lower()): c for c in balance_combos}

    # Defensive fetch: if migration 0008 (expense_category) hasn't been
    # applied on this DB yet, the default `list(qs[:200])` would 500 because
    # the SELECT includes a column that doesn't exist. Falling back to a
    # deferred fetch + a stubbed value in __dict__ prevents the template
    # from triggering a lazy load that would also fail.
    # The site cards + inline list explicitly EXCLUDE salary rows — those
    # get a dedicated panel (`salary_panel`). Auto-salary transactions are
    # a per-day payroll shadow of Attendance, not something the admin edits
    # inline like a food / diesel receipt.
    non_salary_display_qs = non_salary_qs
    try:
        transactions_list = list(non_salary_display_qs[:200])
    except Exception:
        transactions_list = list(non_salary_display_qs.defer('expense_category')[:200])
        for _t in transactions_list:
            _t.__dict__['expense_category'] = ''

    for t in transactions_list:
        src = (t.income_source or '').strip().lower()
        acc = (t.payment_mode or '').strip().lower()
        t.combo_data = combo_map.get((src, acc)) if (src and acc) else None

    # ── Serialize to JSON for client-side JS rendering ──
    _mode_display = dict(Transaction.PAYMENT_MODE)
    transactions_json = json.dumps([{
        'id':          t.pk,
        'date':        t.date.isoformat() if t.date else '',
        'type':        t.type,
        'desc':        t.description or '',
        'amount':      float(t.amount),
        'catId':       t.category_id,
        'catName':     t.category.name if t.category else 'General',
        'catColor':    t.category.color if t.category else '#94a3b8',
        'branchId':    t.branch_id,
        'branchName':  t.branch.name if t.branch else '',
        'mode':        _mode_display.get(t.payment_mode, t.payment_mode or ''),
        'account':     t.payment_mode or '',
        'vendor':      t.vendor or '',
        'ref':         t.reference or '',
        'incomeSource': t.income_source or '',
    } for t in transactions_list])

    # Compute the site-name seed list for the site cards. Rules:
    #   * today_default_active (fresh landing, no filters) → today's active
    #     Attendance/WorkLog sites only, plus Projects.Site master + legacy
    #     Transaction distincts (the helper always includes those).
    #   * Explicit site filter → just that one site.
    #   * Any other filter → full union across Projects/Attendance/Transaction.
    from finance.services.site_cards import get_expense_card_seed, has_unassigned_transactions
    selected_sites = [s for s in filters.get('sites') or [] if s != 'UNASSIGNED']
    seed_site_names = get_expense_card_seed(
        admin_id,
        selected_sites=selected_sites,
        restrict_today=today_default_active,
        today=timezone.localdate() if today_default_active else None,
        cycle_start=cycle_start,
        cycle_end=cycle_end,
    )

    # Group the filtered transactions by Account (legacy) and Site (new
    # site-based card view, Income/Expense restructure). The site card uses
    # live Income totals as Credit — never duplicated. Salary rows were
    # already stripped from transactions_list; the seed above may include
    # salary-only sites so they still render as empty cards.
    accounts_grouped = _group_expenses_by_account(transactions_list)
    sites_grouped    = _group_expenses_by_site(transactions_list, request.user, seed_names=seed_site_names)

    # Only surface the "(No Site)" / Unassigned bucket when a legacy row
    # without a location_site actually exists — otherwise the card is empty
    # clutter. Drop it from sites_grouped when the check comes back False.
    has_unassigned = has_unassigned_transactions(admin_id, txn_type='expense')
    if not has_unassigned:
        sites_grouped = [g for g in sites_grouped if g['site_key']]

    # Site dropdown — full universe of sites the tenant has ever used, so
    # the filter never looks broken.
    site_options = _all_site_options(admin_id)

    # Unique Credit From / Credit To values across this tenant's expense
    # rows — feeds the inline-edit dropdowns (datalists) in the site detail
    # panel. No model fields added; we just surface what's already stored.
    payment_by_options = list(
        Transaction.objects.filter(admin_id=admin_id, type='expense')
        .exclude(payment_by__isnull=True).exclude(payment_by='')
        .values_list('payment_by', flat=True).distinct().order_by('payment_by')
    )
    vendor_options = list(
        Transaction.objects.filter(admin_id=admin_id, type='expense')
        .exclude(vendor__isnull=True).exclude(vendor='')
        .values_list('vendor', flat=True).distinct().order_by('vendor')
    )

    # Structured option lists for the checkbox-dropdown widget (Account,
    # Source, Category, Site). Each entry is {value, display}. Site adds an
    # optional "Unassigned" bucket when legacy null-location rows exist.
    _cat_qs = Category.objects.filter(admin_id=admin_id, type='expense')
    _acct_names = list(
        Source.objects.filter(admin_id=admin_id, type='account').values_list('name', flat=True)
    )
    _src_names = list(
        Source.objects.filter(admin_id=admin_id, type='income').values_list('name', flat=True)
    )
    account_options  = [{'value': n, 'display': n} for n in _acct_names if n]
    source_options   = [{'value': n, 'display': n} for n in _src_names if n]
    category_options = [{'value': str(c.pk), 'display': c.name} for c in _cat_qs]
    site_ms_options  = [{'value': s['name'], 'display': s['name']} for s in site_options]
    if has_unassigned:
        site_ms_options.append({'value': 'UNASSIGNED', 'display': '— Unassigned —'})

    # Salary panel data — grouped per site, one totals row + drill-down of
    # per-date entries. Scoped to the active cycle so the panel matches
    # the top summary card.
    from finance.services.salary_panel import build_salary_panel
    salary_panel = build_salary_panel(admin_id, cycle_start, cycle_end)

    # Available cycles for the top filter dropdown AND every per-card
    # cycle picker (they share the same list).
    available_cycles = _available_cycles(n=8)
    active_cycle_key = f"{cycle_start.isoformat()}_{cycle_end.isoformat()}"
    current_cycle_month_key = get_salary_cycle(today_ist())['month_key']

    return render(request, 'finance/list.html', {
        'transactions':        transactions_list,
        'transactions_json':   transactions_json,
        'accounts_grouped':    accounts_grouped,
        'sites_grouped':       sites_grouped,
        'total_expense':       total_expense,
        'total_income':        total_income,
        'net_balance':         net_balance,
        # Cycle-scoped totals (top "This Cycle" card).
        'total_salary_expense':     total_salary_expense,
        'total_non_salary_expense': total_non_salary_expense,
        'cycle_start':          cycle_start,
        'cycle_end':            cycle_end,
        'cycle_label':          cycle_label,
        'active_cycle_key':     active_cycle_key,
        'available_cycles':     available_cycles,
        'salary_panel':         salary_panel,
        'categories':          Category.objects.filter(admin_id=admin_id, type='expense'),
        'filters':               filters,
        'site_options':          site_options,
        'has_unassigned':        has_unassigned,
        'today_default_active':  today_default_active,
        'is_date_scoped':        bool(exact_date),
        'today_ist_iso':         timezone.localdate().isoformat(),
        'location_sites_json':  _location_sites_json(request.user),
        'income_sources_json':  _shared_sources_json(request.user),
        'accounts_json':        _shared_accounts_json(request.user),
        # JSON-encoded selected values (legacy — no longer read by the
        # checkbox-dropdown widget, kept in case any embedded JS still needs them).
        'selected_accounts_json':       json.dumps(filters.get('accounts') or []),
        'selected_income_sources_json': json.dumps(filters.get('income_sources') or []),
        # Widget option lists + selected values (for _multiselect.html).
        'account_options':         account_options,
        'source_options':          source_options,
        'category_options':        category_options,
        'site_ms_options':         site_ms_options,
        'selected_accounts':       filters.get('accounts') or [],
        'selected_income_sources': filters.get('income_sources') or [],
        'selected_categories':     filters.get('categories') or [],
        'selected_sites':          filters.get('sites') or [],
        'balance_data':         balance_combos,
        # Per-card default selection — matches the page-level active cycle
        # so cards land showing the same window as the top summary. Value
        # is the cycle's END month key ('YYYY-MM'), stored in the option
        # `value` and matched against `data-monthly` bucket keys by JS.
        # Legacy variable name kept for template compatibility.
        'current_month':        current_cycle_month_key,
        # Unique values for the inline-edit Credit From / Credit To dropdowns.
        'payment_by_options':   payment_by_options,
        'vendor_options':       vendor_options,
    })

@login_required
def add_transaction(request):
    # Default any GET hit on /expenses/add/ back to the list page (modal-only flow).
    if request.method != 'POST':
        return redirect('expenses:list')

    # Force the type to expense regardless of what the client sent.
    post_data = request.POST.copy()
    post_data['type'] = 'expense'

    form = TransactionForm(request.user, post_data)
    ajax = _is_ajax(request)

    if form.is_valid():
        t = form.save(commit=False)
        t.user     = request.user
        t.type     = 'expense'
        t.admin_id = get_admin_id(request.user)
        t.save()
        if t.location_site:
            admin_id = get_admin_id(request.user)
            name = t.location_site.strip()
            if name and not LocationSite.objects.filter(admin_id=admin_id, name__iexact=name).exists():
                LocationSite.objects.create(admin_id=admin_id, name=name, created_by=request.user)
        if ajax:
            d = _expense_to_dict(t)
            d['combo'] = _expense_combo_data(t, request.user)
            return JsonResponse({'success': True, 'expense': d})
        messages.success(request, "Expense recorded.")
        return redirect('expenses:list')

    if ajax:
        errors = {f: [str(e) for e in errs] for f, errs in form.errors.items()}
        first_error = next(iter(errors.values()), ['Invalid data.'])[0]
        return JsonResponse({'success': False, 'errors': errors, 'message': first_error}, status=400)

    # Non-AJAX fallback: still keep user on the list page; surface the error in flash.
    messages.error(request, "Could not save expense. Please try again.")
    return redirect('expenses:list')

@login_required
def edit_transaction(request, pk):
    t = get_object_or_404(Transaction, pk=pk, admin_id=get_admin_id(request.user))

    # Modal-only flow: any GET hit on /expenses/<pk>/edit/ bounces to the list.
    if request.method != 'POST':
        return redirect('expenses:list')

    post_data = request.POST.copy()
    post_data['type'] = 'expense'

    form = TransactionForm(request.user, post_data, instance=t)
    ajax = _is_ajax(request)

    if form.is_valid():
        obj = form.save(commit=False)
        obj.user     = request.user
        obj.type     = 'expense'
        obj.admin_id = get_admin_id(request.user)
        obj.save()
        if obj.location_site:
            admin_id = get_admin_id(request.user)
            name = obj.location_site.strip()
            if name and not LocationSite.objects.filter(admin_id=admin_id, name__iexact=name).exists():
                LocationSite.objects.create(admin_id=admin_id, name=name, created_by=request.user)
        if ajax:
            d = _expense_to_dict(obj)
            d['combo'] = _expense_combo_data(obj, request.user)
            return JsonResponse({'success': True, 'expense': d})
        messages.success(request, "Expense updated.")
        return redirect('expenses:list')

    if ajax:
        errors = {f: [str(e) for e in errs] for f, errs in form.errors.items()}
        first_error = next(iter(errors.values()), ['Invalid data.'])[0]
        return JsonResponse({'success': False, 'errors': errors, 'message': first_error}, status=400)

    messages.error(request, "Could not update expense. Please try again.")
    return redirect('expenses:list')

@login_required
def delete_transaction(request, pk):
    t = get_object_or_404(Transaction, pk=pk, admin_id=get_admin_id(request.user))
    if request.method == 'POST':
        t.delete()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        messages.success(request, "Deleted.")
    return redirect('expenses:list')


@login_required
def delete_expenses_by_sites(request):
    """Bulk-delete every expense Transaction whose location_site matches
    one of the site names in `sites[]` (case-insensitive). Scope is
    strictly the current admin tenant — never touches Projects.Site or
    AttendanceRecord rows. POST only, atomic, returns per-site counts."""
    from django.views.decorators.http import require_POST as _require_POST  # local import
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)

    admin_id = get_admin_id(request.user)
    if not admin_id:
        return JsonResponse({'success': False, 'error': 'Forbidden'}, status=403)

    site_names = request.POST.getlist('sites') or request.POST.getlist('sites[]')
    site_names = [s.strip() for s in site_names if (s or '').strip()]
    if not site_names:
        return JsonResponse({'success': False, 'error': 'No sites selected'}, status=400)

    from django.db import transaction as _db_tx
    import logging as _logging
    _log = _logging.getLogger(__name__)

    # TODO remove debug logging after delete flow verified
    _log.info(
        "DEBUG delete_expenses_by_sites: admin_id=%r raw_body=%r POST=%r parsed_sites=%r",
        admin_id, request.body[:500], dict(request.POST.lists()), site_names,
    )

    # Case-insensitive OR across site names, scoped to this admin + expenses.
    # FIX 2 (belt-and-suspenders): also match against the internal-whitespace
    # collapsed form so "Foo  Bar" from a card matches a "Foo Bar" row.
    import re as _re
    site_q = Q()
    for name in site_names:
        stripped = name.strip()
        collapsed = _re.sub(r'\s+', ' ', stripped)
        site_q |= Q(location_site__iexact=stripped)
        if collapsed != stripped:
            site_q |= Q(location_site__iexact=collapsed)

    base_qs = Transaction.objects.filter(admin_id=admin_id, type='expense').filter(site_q)

    from finance.models import ModuleHiddenSite

    deleted_by_site = {}
    hidden_site_names = []
    with _db_tx.atomic():
        # Count per-site FIRST (for the response), then delete in one shot.
        for name in site_names:
            deleted_by_site[name] = base_qs.filter(location_site__iexact=name).count()
        pre_delete_total = base_qs.count()  # TODO remove debug logging after delete flow verified
        deleted_count, _ = base_qs.delete()

        # Hide the site from Expense Manager view for this admin, whether or
        # not any rows actually got deleted. Projects/Attendance untouched.
        for name in site_names:
            stripped = name.strip()
            if not stripped:
                continue
            _, created = ModuleHiddenSite.objects.get_or_create(
                admin_id=admin_id,
                module='expense',
                site_name=stripped,
                defaults={'hidden_by': request.user if request.user.is_authenticated else None},
            )
            hidden_site_names.append(stripped)

    _log.info(
        "Bulk expense delete: admin_id=%s sites=%s pre_count=%s deleted=%s per_site=%s hidden=%s",
        admin_id, site_names, pre_delete_total, deleted_count, deleted_by_site, hidden_site_names,
    )

    return JsonResponse({
        'success': True,
        'deleted_row_count':  deleted_count,
        'hidden_site_count':  len(hidden_site_names),
        'hidden_site_names':  hidden_site_names,
        'deleted_by_site':    deleted_by_site,
        # Back-compat alias for older JS callers still on this key.
        'deleted_count': deleted_count,
    })


@login_required
def debug_site_match(request):
    """Diagnostic-only endpoint. For each card-list site name reports how
    many Transaction rows this admin actually owns under that name (exact,
    iexact, icontains) plus up to 3 similar samples from the raw distinct
    location_site set. Gated by ?debug=1 AND admin-scope so it can be
    reached in production without polluting the URL surface.

    TODO remove after delete-flow verified.
    """
    if request.GET.get('debug') != '1':
        return JsonResponse({'error': 'not enabled'}, status=404)

    admin_id = get_admin_id(request.user)
    if not admin_id:
        return JsonResponse({'error': 'no admin scope'}, status=403)

    from finance.services.site_cards import get_active_site_names
    card_names = get_active_site_names(admin_id, restrict_today=False)

    raw_qs = Transaction.objects.filter(admin_id=admin_id, type='expense')
    distinct_stored = sorted(
        set(
            (v or '') for v in raw_qs.values_list('location_site', flat=True).distinct()
        ),
        key=lambda s: s.lower(),
    )
    stored_lower_prefix = [(s, s.strip().lower()[:5]) for s in distinct_stored]

    per_card = []
    for name in card_names:
        stripped = (name or '').strip()
        prefix = stripped.lower()[:5]
        exact_ct    = raw_qs.filter(location_site=name).count()
        iexact_ct   = raw_qs.filter(location_site__iexact=stripped).count()
        icontain_ct = raw_qs.filter(location_site__icontains=stripped).count() if stripped else 0
        similar = [s for (s, p) in stored_lower_prefix if p and prefix and p == prefix][:3]
        per_card.append({
            'card_data_site_name': name,
            'stripped_for_match':  stripped,
            'exact_count':         exact_ct,
            'iexact_count':        iexact_ct,
            'icontains_count':     icontain_ct,
            'similar_stored_samples': similar,
        })

    # Sanity — first 3 raw rows: are admin_id + type as expected?
    top_rows = list(raw_qs.order_by('-created_at').values(
        'id', 'admin_id', 'type', 'location_site', 'amount', 'date',
    )[:3])

    return JsonResponse({
        'admin_id_in_scope':       admin_id,
        'total_transaction_rows':  Transaction.objects.filter(admin_id=admin_id).count(),
        'total_expense_rows':      raw_qs.count(),
        'distinct_location_sites_stored': distinct_stored,
        'card_names_from_union':   card_names,
        'per_card_match_counts':   per_card,
        'top_3_expense_rows':      top_rows,
    }, json_dumps_params={'indent': 2})

# ═══════════════════════════════════════════════════════════════════════
# Export — PDF + Excel. Both reuse `_apply_expense_filters` so the report
# reflects the exact filter state the user has on screen (querystring is
# forwarded from the Download links).
# ═══════════════════════════════════════════════════════════════════════
EXPORT_CATEGORIES = ['food', 'room', 'diesel', 'travel', 'ticket', 'other']
EXPORT_CATEGORY_LABELS = {
    'food':   'Food',
    'room':   'Room',
    'diesel': 'Diesel',
    'travel': 'Travel',
    'ticket': 'Ticket',
    'other':  'Misc/Others',
}


def _filter_summary(filters):
    def _join(v):
        return ', '.join(v) if isinstance(v, (list, tuple)) else str(v)
    parts = []
    if filters.get('month'):         parts.append(f"Month={filters['month']}")
    if filters.get('date'):          parts.append(f"Date={filters['date']}")
    if filters.get('from_date') or filters.get('to_date'):
        parts.append(f"Range={filters.get('from_date') or '…'}→{filters.get('to_date') or '…'}")
    if filters.get('sites'):         parts.append(f"Site={_join(filters['sites'])}")
    if filters.get('income_sources'): parts.append(f"Source={_join(filters['income_sources'])}")
    if filters.get('accounts'):      parts.append(f"Account={_join(filters['accounts'])}")
    if filters.get('categories'):    parts.append(f"Category={_join(filters['categories'])}")
    if filters.get('search'):        parts.append(f"Search={filters['search']}")
    return ' | '.join(parts) if parts else 'None'


def _row_category_amounts(t):
    """Return dict {food, room, diesel, travel, ticket, other} — amount in
    the row's expense_category slot, 0 elsewhere. Diesel absorbs 'fuel';
    an unset category falls into 'other' so nothing goes missing."""
    cat = (getattr(t, 'expense_category', '') or '').strip().lower()
    if cat == 'fuel':
        cat = 'diesel'
    if cat not in EXPORT_CATEGORIES:
        cat = 'other'
    amt = float(t.amount or 0)
    return {c: (amt if c == cat else 0.0) for c in EXPORT_CATEGORIES}


@login_required
def export_expenses_pdf(request):
    from django.http import HttpResponse
    try:
        return _render_expenses_pdf(request)
    except Exception as exc:
        # Surface the traceback in the browser instead of a raw 500 —
        # Railway prod is where this gets tested, and grabbing the
        # traceback out of the platform logs is friction. plain/text
        # response keeps the browser from trying to render HTML.
        import traceback
        tb = traceback.format_exc()
        body = (
            'Expense PDF export failed.\n\n'
            f'Exception: {type(exc).__name__}: {exc}\n\n'
            'Traceback (most recent call last):\n'
            f'{tb}\n'
        )
        resp = HttpResponse(body, content_type='text/plain; charset=utf-8', status=500)
        resp['X-Export-Error'] = '1'
        return resp


def _render_expenses_pdf(request):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak,
    )
    from django.http import HttpResponse
    from django.utils.html import escape as _esc
    from io import BytesIO
    from decimal import Decimal as _D

    admin_id = get_admin_id(request.user)
    qs, filters, _has, _ed, _mo, _td = _apply_expense_filters(request, admin_id)
    try:
        rows = list(qs.order_by('location_site', 'date'))
    except Exception:
        rows = list(qs.defer('expense_category').order_by('location_site', 'date'))
        for _t in rows:
            _t.__dict__.setdefault('expense_category', '')

    sites_grouped = _group_expenses_by_site(rows, request.user)

    try:
        from income.models import Income as _Income
        total_income = _Income.objects.filter(admin_id=admin_id).aggregate(t=Sum('amount'))['t'] or 0
    except Exception:
        total_income = 0
    total_expense = sum((float(t.amount or 0) for t in rows), 0.0)
    net_balance = float(total_income) - total_expense

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=10 * mm, bottomMargin=10 * mm,
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle('H1', parent=styles['Heading1'], fontSize=14, spaceAfter=4)
    small = ParagraphStyle('small', parent=styles['Normal'], fontSize=8, textColor=colors.grey)
    label = ParagraphStyle('label', parent=styles['Normal'], fontSize=9)

    story = []
    ts = timezone.localtime().strftime('%d %b %Y, %H:%M IST')
    admin_name = (
        request.user.get_full_name() or request.user.username or request.user.email or 'Admin'
    )
    story.append(Paragraph('SPIM Suite — Expense Report', h1))
    story.append(Paragraph(_esc(f'Generated {ts} · {admin_name}'), small))
    story.append(Paragraph(f'<b>Filters:</b> {_esc(_filter_summary(filters))}', label))
    story.append(Spacer(1, 6))

    header_cells = ['Date', 'Credit', 'From', 'To'] + \
        [EXPORT_CATEGORY_LABELS[c] for c in EXPORT_CATEGORIES] + ['Remarks']

    for g in sites_grouped:
        story.append(Spacer(1, 6))
        story.append(Paragraph(
            f"<b>{_esc(g['site'])}</b> &nbsp;·&nbsp; "
            f"Credit ₹{float(g['credit']):,.2f} &nbsp;·&nbsp; "
            f"Debit ₹{float(g['debit']):,.2f} &nbsp;·&nbsp; "
            f"Balance ₹{float(g['balance']):,.2f}",
            label,
        ))
        data = [header_cells]
        for t in g['transactions']:
            cats = _row_category_amounts(t)
            data.append([
                t.date.strftime('%d %b %Y') if t.date else '',
                f"{float(g['credit']):,.2f}" if g['credit'] else '',
                (t.payment_by or '')[:24],
                (t.vendor or '')[:24],
                *(f"{cats[c]:,.2f}" if cats[c] else '' for c in EXPORT_CATEGORIES),
                (t.description or '')[:60],
            ])
        tbl = Table(data, repeatRows=1, colWidths=[
            22 * mm, 20 * mm, 26 * mm, 26 * mm,
            18 * mm, 18 * mm, 18 * mm, 18 * mm, 18 * mm, 22 * mm,
            50 * mm,
        ])
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f1f5f9')),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.HexColor('#334155')),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 0), (-1, -1), 8),
            ('GRID',       (0, 0), (-1, -1), 0.25, colors.HexColor('#cbd5e1')),
            ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN',      (4, 1), (9, -1), 'RIGHT'),
            ('ALIGN',      (1, 1), (1, -1), 'RIGHT'),
        ]))
        story.append(tbl)

    story.append(Spacer(1, 10))
    story.append(Paragraph(
        f"<b>Grand Total</b> &nbsp;·&nbsp; Income ₹{float(total_income):,.2f} &nbsp;·&nbsp; "
        f"Expense ₹{total_expense:,.2f} &nbsp;·&nbsp; Net ₹{net_balance:,.2f}",
        label,
    ))
    if not sites_grouped:
        story.append(Paragraph('<i>No transactions match the current filter.</i>', label))

    doc.build(story)
    fname = 'spim_expenses_' + timezone.localtime().strftime('%Y-%m-%d_%H%M%S') + '.pdf'
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


@login_required
def export_expenses_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from io import BytesIO

    admin_id = get_admin_id(request.user)
    qs, filters, _has, _ed, _mo, _td = _apply_expense_filters(request, admin_id)
    try:
        rows = list(qs.order_by('location_site', 'date'))
    except Exception:
        rows = list(qs.defer('expense_category').order_by('location_site', 'date'))
        for _t in rows:
            _t.__dict__.setdefault('expense_category', '')

    try:
        from income.models import Income as _Income
        total_income = float(_Income.objects.filter(admin_id=admin_id).aggregate(t=Sum('amount'))['t'] or 0)
    except Exception:
        total_income = 0.0
    total_expense = sum((float(t.amount or 0) for t in rows), 0.0)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Expenses'

    ts = timezone.localtime().strftime('%d %b %Y, %H:%M IST')
    ws.cell(row=1, column=1, value=f'SPIM Suite — Expense Report  (Generated {ts})').font = Font(bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    ws.cell(row=2, column=1, value=f'Filters: {_filter_summary(filters)}').font = Font(italic=True, color='64748B')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)

    headers = ['Site', 'Date', 'Credit', 'From', 'To'] + \
        [EXPORT_CATEGORY_LABELS[c] for c in EXPORT_CATEGORIES] + ['Remarks']
    header_fill = PatternFill('solid', fgColor='F1F5F9')
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(bold=True, color='334155')
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center')

    r = 5
    rows_by_site = {}
    for t in rows:
        rows_by_site.setdefault((t.location_site or '(No Site)'), []).append(t)
    for site in sorted(rows_by_site.keys(), key=lambda s: s.lower()):
        for t in sorted(rows_by_site[site], key=lambda x: (x.date or datetime.date.min)):
            cats = _row_category_amounts(t)
            ws.cell(row=r, column=1,  value=site)
            ws.cell(row=r, column=2,  value=t.date.strftime('%Y-%m-%d') if t.date else '')
            ws.cell(row=r, column=3,  value='')
            ws.cell(row=r, column=4,  value=t.payment_by or '')
            ws.cell(row=r, column=5,  value=t.vendor or '')
            for i, k in enumerate(EXPORT_CATEGORIES):
                cell = ws.cell(row=r, column=6 + i, value=(cats[k] or None))
                if cats[k]:
                    cell.number_format = '#,##0.00'
            ws.cell(row=r, column=12, value=t.description or '')
            r += 1

    ws.freeze_panes = 'A5'
    for i in range(1, 13):
        col = get_column_letter(i)
        max_len = 10
        for row_cells in ws[col]:
            v = row_cells.value
            if v is None:
                continue
            max_len = max(max_len, min(40, len(str(v)) + 2))
        ws.column_dimensions[col].width = max_len

    summary_row = r + 1
    ws.cell(row=summary_row,     column=1, value='Total Income').font  = Font(bold=True)
    ws.cell(row=summary_row,     column=2, value=total_income).number_format = '#,##0.00'
    ws.cell(row=summary_row + 1, column=1, value='Total Expense').font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=2, value=total_expense).number_format = '#,##0.00'
    ws.cell(row=summary_row + 2, column=1, value='Net Balance').font   = Font(bold=True)
    ws.cell(row=summary_row + 2, column=2, value=total_income - total_expense).number_format = '#,##0.00'

    buf = BytesIO()
    wb.save(buf)
    fname = 'spim_expenses_' + timezone.localtime().strftime('%Y-%m-%d_%H%M%S') + '.xlsx'
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


@login_required
def category_list(request):
    return redirect('categories:expense_list')

@login_required
def delete_category(request, pk):
    c = get_object_or_404(Category, pk=pk, admin_id=get_admin_id(request.user))
    if request.method == 'POST':
        c.delete()
        messages.success(request, "Category removed.")
    return redirect('expenses:categories')
