from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.db.models import Q, Sum
from django.db import transaction
from django.utils import timezone
import calendar
import datetime
import json
import random
import secrets
import string
from decimal import Decimal
from django.contrib.auth.hashers import make_password
from .models import Employee, BankDetail, PFDetail, SalaryUpdate, SalaryStructure, JobRole, EmployeeLevel
from .forms import EmployeeForm, BankDetailForm, PFDetailForm, SalaryForm
from branches.models import LocationSite
from accounts.views import get_admin_id


def _sync_employee_location(employee, user):
    """Register an employee's location/site in the centralized LocationSite table."""
    admin_id = get_admin_id(user)
    loc  = (employee.location or '').strip()
    site = (employee.site or '').strip()
    if not loc and not site:
        return
    combined = f"{loc} / {site}" if loc and site else (loc or site)
    if not LocationSite.objects.filter(admin_id=admin_id, name__iexact=combined).exists():
        LocationSite.objects.create(admin_id=admin_id, name=combined, created_by=user)


def _generate_employee_id(admin_id):
    """Return the next available SPIMXXX employee ID for the given tenant."""
    existing_ids = Employee.objects.filter(admin_id=admin_id).values_list('employee_id', flat=True)
    max_num = 0
    for eid in existing_ids:
        if eid and eid.upper().startswith('SPIM'):
            try:
                num = int(eid[4:])
                if num > max_num:
                    max_num = num
            except (ValueError, IndexError):
                pass
    return f"SPIM{(max_num + 1):03d}"


def _apply_role_level_salary(employee, admin_id, force=False):
    """
    Auto-fill an Employee's monetary fields from the matching SalaryStructure
    (job_role + level). Lookup keys: (admin_id, job_role, level). When `force`
    is True (Role/Level changed on edit), overwrites existing base_salary and
    fixed_allowance so the centralized config remains the source of truth.
    Otherwise preserves any explicit override the admin typed into the form.

    Custom-override guard: when Employee.salary_is_custom_override is True,
    base_salary is never touched (Case 3 / Case 4 of the manual-override
    spec). fixed_allowance is allowed to refresh because it represents the
    food-allowance line, which is configured per Role+Level, not per
    employee — admins manage food allowance via the Salary Config.
    """
    if getattr(employee, 'salary_is_custom_override', False):
        # Honor the per-employee manual override — do not refresh base_salary
        # from the centralized config, even when force=True.
        force_base = False
    else:
        force_base = force
    level_val = (employee.level or '').strip()
    if not level_val:
        return
    structure = None
    if employee.job_role_id:
        structure = SalaryStructure.objects.filter(
            admin_id=admin_id, job_role_id=employee.job_role_id, level=level_val,
        ).first()
    # Fallback: match by JobRole.name == Employee.designation so that the
    # Salary Config table works even when the employee has no explicit
    # JobRole FK (the Add modal stores role as free-text designation).
    if not structure:
        designation = (employee.designation or '').strip()
        if designation:
            structure = SalaryStructure.objects.filter(
                admin_id=admin_id,
                job_role__admin_id=admin_id,
                job_role__name__iexact=designation,
                level=level_val,
            ).first()
    if not structure:
        return
    if force_base or not employee.base_salary:
        employee.base_salary = structure.base_salary
    # `fixed_allowance` carries the food allowance in the existing schema.
    if force or not employee.fixed_allowance:
        employee.fixed_allowance = structure.food_allowance


def _generate_app_password():
    """
    Return a cryptographically random mobile-app password.

    Requirements (SPIM Lite):
      * Minimum 8 characters
      * Includes uppercase, lowercase, digits, special characters
      * Generated via `secrets` for cryptographic strength
    """
    uppers   = string.ascii_uppercase
    lowers   = string.ascii_lowercase
    digits   = string.digits
    specials = '!@#$%^&*'
    # Guarantee at least one of each required class
    required = [
        secrets.choice(uppers),
        secrets.choice(lowers),
        secrets.choice(digits),
        secrets.choice(specials),
    ]
    pool      = uppers + lowers + digits + specials
    remaining = [secrets.choice(pool) for _ in range(6)]  # 4 + 6 = 10 chars
    chars     = required + remaining
    # Shuffle without bias
    secrets.SystemRandom().shuffle(chars)
    return ''.join(chars)


def _location_sites_json(user):
    """Return a JSON string of all centralized location names for this tenant."""
    admin_id = get_admin_id(user)
    names = list(LocationSite.objects.filter(admin_id=admin_id).values_list('name', flat=True))
    return json.dumps(names)


def timezone_month_map(month_name):
    months = {
        "January": 1, "February": 2, "March": 3, "April": 4, "May": 5, "June": 6,
        "July": 7, "August": 8, "September": 9, "October": 10, "November": 11, "December": 12
    }
    return months.get(month_name, 1)


@login_required
def employee_list(request):
    search_q = request.GET.get('q', '')
    location_f = request.GET.get('location', '')
    site_f = request.GET.get('site', '')
    admin_id = get_admin_id(request.user)
    qs = Employee.objects.filter(admin_id=admin_id)
    if search_q:
        qs = qs.filter(
            Q(name__icontains=search_q) |
            Q(employee_id__icontains=search_q) |
            Q(designation__icontains=search_q)
        )
    if location_f:
        qs = qs.filter(location=location_f)
    if site_f:
        qs = qs.filter(site=site_f)

    locations = Employee.objects.filter(admin_id=admin_id).values_list('location', flat=True).distinct().order_by('location')
    sites = Employee.objects.filter(admin_id=admin_id).values_list('site', flat=True).distinct().order_by('site')

    # ── Global master dropdowns (Task 4 / 2026-05-24) ──────────────────
    # Single source of truth per master. Roles/levels/locations/sites are
    # collected from EVERY table that already holds them so a value typed
    # anywhere in the system shows up in every dependent dropdown.
    def _dedup_sorted(values):
        seen = set()
        out  = []
        for v in values:
            v = (v or '').strip()
            if v and v.lower() not in seen:
                seen.add(v.lower())
                out.append(v)
        return sorted(out, key=lambda x: x.lower())

    # Local re-import guards against a stale autoreload session where the
    # module-level import didn't get refreshed.
    from .models import (
        JobRole as _JobRole,
        SalaryStructure as _SalaryStructure,
        EmployeeLevel as _EmployeeLevel,
    )
    role_pool = list(
        Employee.objects.filter(admin_id=admin_id).values_list('designation', flat=True)
    ) + list(
        _JobRole.objects.filter(admin_id=admin_id).values_list('name', flat=True)
    )
    level_pool = list(
        Employee.objects.filter(admin_id=admin_id).values_list('level', flat=True)
    ) + list(
        _SalaryStructure.objects.filter(admin_id=admin_id).values_list('level', flat=True)
    ) + list(
        _EmployeeLevel.objects.filter(admin_id=admin_id).values_list('name', flat=True)
    )
    # Location pool — Employee.location PLUS the part before " / " in
    # LocationSite.name (centralized branches registry).
    loc_pool = list(Employee.objects.filter(admin_id=admin_id).values_list('location', flat=True))
    site_pool = list(Employee.objects.filter(admin_id=admin_id).values_list('site', flat=True))
    for combined in LocationSite.objects.filter(admin_id=admin_id).values_list('name', flat=True):
        if ' / ' in combined:
            l, s = combined.split(' / ', 1)
            loc_pool.append(l); site_pool.append(s)
        else:
            loc_pool.append(combined)

    # Split into people vs vehicles for the two-section Master layout. Both
    # inherit Employee.Meta.ordering = ['name']; the split is purely visual.
    qs = qs.select_related('bank_details', 'pf_details')
    return render(request, 'employees/list.html', {
        'employees':           qs.filter(is_vehicle=False),
        'vehicles':            qs.filter(is_vehicle=True),
        'all_employees':       Employee.objects.filter(admin_id=admin_id).order_by('name'),
        'locations':           locations,
        'sites':               sites,
        'search_q':            search_q,
        'location_f':          location_f,
        'site_f':              site_f,
        'location_sites_json': _location_sites_json(request.user),
        'next_employee_id':    _generate_employee_id(admin_id),
        'preview_password':    _generate_app_password(),
        # Global masters — consumed by the four <datalist>s in the template.
        'master_roles':        _dedup_sorted(role_pool),
        'master_levels':       _dedup_sorted(level_pool),
        'master_locations':    _dedup_sorted(loc_pool),
        'master_sites':        _dedup_sorted(site_pool),
    })


@login_required
def add_employee(request):
    if request.method == 'POST':
        form = EmployeeForm(request.POST)
        if form.is_valid():
            employee = form.save(commit=False)
            employee.created_by = request.user
            admin_id = get_admin_id(request.user)
            employee.admin_id = admin_id

            # Auto-generate Employee ID if blank, ensure uniqueness
            emp_id = (employee.employee_id or '').strip()
            if not emp_id:
                emp_id = _generate_employee_id(admin_id)
            while Employee.objects.filter(admin_id=admin_id, employee_id=emp_id).exists():
                emp_id = _generate_employee_id(admin_id)
            employee.employee_id = emp_id

            # Auto-generate Mobile App Password if blank.
            # The plaintext is shown ONCE in the admin UI (existing behavior)
            # but the source of truth for authentication is mobile_password_hash.
            if not (employee.mobile_app_password or '').strip():
                employee.mobile_app_password = _generate_app_password()
            employee.mobile_password_hash = make_password(employee.mobile_app_password)

            # SPIM Lite login id mirrors the Employee ID.
            employee.employee_login_id = employee.employee_id
            employee.mobile_account_active = True

            # Role + Level → SalaryStructure auto-fill (Task 5)
            _apply_role_level_salary(employee, admin_id)

            employee.save()
            _sync_employee_location(employee, request.user)
            messages.success(request, f"Employee {employee.name} added successfully.")

            bank_name = request.POST.get('bank_name')
            if bank_name:
                BankDetail.objects.create(
                    employee=employee,
                    bank_name=bank_name,
                    account_holder=request.POST.get('account_holder', ''),
                    account_number=request.POST.get('account_number', ''),
                    ifsc_code=request.POST.get('ifsc_code', ''),
                    branch=request.POST.get('branch', ''),
                )

            pf_number = request.POST.get('pf_number')
            if pf_number:
                PFDetail.objects.create(
                    employee=employee,
                    pf_number=pf_number,
                    uan_number=request.POST.get('uan_number', ''),
                    esic_number=request.POST.get('esic_number', ''),
                    status='added'
                )

            return redirect('employees:list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.capitalize()}: {error}")
            return redirect('employees:list')
    else:
        return redirect('employees:list')


@login_required
def edit_employee(request, pk):
    employee = get_object_or_404(Employee, pk=pk, admin_id=get_admin_id(request.user))
    original_emp_id     = employee.employee_id
    original_role       = (employee.designation or '').strip().lower()
    original_level      = (employee.level or '').strip().lower()
    original_job_role   = employee.job_role_id
    if request.method == 'POST':
        form = EmployeeForm(request.POST, instance=employee)
        if form.is_valid():
            new_emp_id = (form.cleaned_data.get('employee_id') or '').strip()
            if new_emp_id and new_emp_id != original_emp_id:
                employee.employee_login_id = new_emp_id
            updated = form.save()
            # Re-apply Salary Config when Role / Level / JobRole changed so
            # base_salary stays in sync with the centralized SalaryStructure.
            role_changed  = (updated.designation or '').strip().lower() != original_role
            level_changed = (updated.level or '').strip().lower() != original_level
            jr_changed    = updated.job_role_id != original_job_role
            if role_changed or level_changed or jr_changed:
                admin_id = get_admin_id(request.user)
                _apply_role_level_salary(updated, admin_id, force=True)
                updated.save(update_fields=['base_salary', 'fixed_allowance'])
            _sync_employee_location(updated, request.user)
            messages.success(request, f"Employee {employee.name} updated.")
            return redirect('employees:list')
    else:
        form = EmployeeForm(instance=employee)
    return render(request, 'employees/edit_form.html', {
        'form': form,
        'employee': employee,
        'location_sites_json': _location_sites_json(request.user),
    })


@login_required
def delete_employee(request, pk):
    employee = get_object_or_404(Employee, pk=pk, admin_id=get_admin_id(request.user))
    if request.method == 'POST':
        name = employee.name
        employee.delete()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'name': name})
        messages.success(request, f"Employee {name} removed.")
    return redirect('employees:list')


@login_required
def employee_edit_ajax(request, pk):
    """AJAX-only endpoint: update core employee fields without a full-page reload."""
    if request.headers.get('X-Requested-With') != 'XMLHttpRequest':
        return redirect('employees:list')
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'})
    employee = get_object_or_404(Employee, pk=pk, admin_id=get_admin_id(request.user))
    name = request.POST.get('name', '').strip()
    if not name:
        return JsonResponse({'success': False, 'error': 'Name is required'})
    original_role  = (employee.designation or '').strip().lower()
    original_level = (employee.level or '').strip().lower()
    employee.name                = name
    posted_emp_id = request.POST.get('employee_id', employee.employee_id).strip()
    if posted_emp_id != employee.employee_id:
        employee.employee_login_id = posted_emp_id
    employee.employee_id         = posted_emp_id
    employee.designation         = request.POST.get('designation', employee.designation).strip()
    # Persist level / mobile / branch — previously silently dropped because
    # this handler only updated a hard-coded subset of fields. Without these
    # assignments, the in-page Edit modal would save (HTTP 200) but the new
    # level/mobile/branch would be lost on the server, breaking the
    # SPIM Lite reflection chain.
    employee.level               = request.POST.get('level', employee.level).strip()
    employee.mobile              = request.POST.get('mobile', employee.mobile).strip()
    employee.branch              = request.POST.get('branch', employee.branch).strip()
    employee.location            = request.POST.get('location', employee.location).strip()
    employee.site                = request.POST.get('site', employee.site).strip()
    posted_pw = request.POST.get('mobile_app_password', employee.mobile_app_password).strip()
    if posted_pw and posted_pw != employee.mobile_app_password:
        employee.mobile_app_password = posted_pw
        employee.mobile_password_hash = make_password(posted_pw)
    salary_str = request.POST.get('base_salary', '').strip()
    if salary_str:
        try:
            employee.base_salary = float(salary_str)
        except ValueError:
            pass
    # Auto-assign base salary from centralized Salary Config when role/level
    # changes (Tasks 3 / 4). When admin explicitly typed a base_salary in the
    # same submission we honor it (kept above); otherwise the structure value
    # wins. force=True ensures stale base salaries follow the new mapping.
    admin_id = get_admin_id(request.user)
    role_changed  = (employee.designation or '').strip().lower() != original_role
    level_changed = (employee.level or '').strip().lower() != original_level
    if (role_changed or level_changed) and not salary_str:
        _apply_role_level_salary(employee, admin_id, force=True)
    employee.save()
    _sync_employee_location(employee, request.user)
    return JsonResponse({'success': True})


@login_required
def export_json_employees(request):
    """Download all employees for this tenant as a JSON file."""
    from django.http import HttpResponse
    admin_id = get_admin_id(request.user)
    qs = Employee.objects.filter(admin_id=admin_id).values(
        'name', 'employee_id', 'designation', 'department',
        'location', 'site', 'base_salary', 'status', 'joining_date'
    )
    payload = json.dumps({'employees': list(qs)}, indent=2, default=str)
    resp = HttpResponse(payload, content_type='application/json')
    resp['Content-Disposition'] = 'attachment; filename="employees_export.json"'
    return resp


@login_required
def export_excel_employees(request):
    """Download the Employee Master list for this tenant as an XLSX file.

    Mirrors the Employee Master (`employee_list`) queryset so the export
    always reflects what the admin sees on screen: same tenant scope, same
    optional q/location/site filters, same model-default ordering by name.
    """
    from io import BytesIO
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    admin_id = get_admin_id(request.user)
    search_q   = request.GET.get('q', '')
    location_f = request.GET.get('location', '')
    site_f     = request.GET.get('site', '')

    employees = (
        Employee.objects.filter(admin_id=admin_id)
        .select_related('job_role', 'bank_details', 'pf_details')
    )
    if search_q:
        employees = employees.filter(
            Q(name__icontains=search_q) |
            Q(employee_id__icontains=search_q) |
            Q(designation__icontains=search_q)
        )
    if location_f:
        employees = employees.filter(location=location_f)
    if site_f:
        employees = employees.filter(site=site_f)
    # Employee.Meta.ordering = ['name'] — match Master page exactly.
    employees = employees.order_by('name')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Employees'

    headers = [
        'Employee ID', 'Password', 'Employee Name', 'Mobile', 'Role', 'Level',
        'Department', 'Designation', 'Branch', 'Location', 'Site',
        'Status', 'Joining Date', 'Salary Type', 'Base Salary',
        'Fixed Allowance', 'Bank Details Status', 'PF Status',
    ]
    ws.append(headers)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1E293B')
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for emp in employees:
        role_name = emp.job_role.name if emp.job_role_id else ''
        bank_status = ''
        if hasattr(emp, 'bank_details'):
            bank_status = emp.bank_details.get_status_display() or 'Added'
        else:
            bank_status = 'Not Added'
        pf_status = ''
        if hasattr(emp, 'pf_details'):
            pf_status = emp.pf_details.get_status_display() or 'Added'
        else:
            pf_status = 'Not Added'

        ws.append([
            emp.employee_id or '',
            emp.mobile_app_password or '',
            emp.name or '',
            emp.mobile or '',
            role_name,
            emp.level or '',
            emp.department or '',
            emp.designation or '',
            emp.branch or '',
            emp.location or '',
            emp.site or '',
            emp.get_status_display() if emp.status else '',
            emp.joining_date,
            emp.get_salary_type_display() if emp.salary_type else '',
            float(emp.base_salary) if emp.base_salary is not None else 0,
            float(emp.fixed_allowance) if emp.fixed_allowance is not None else 0,
            bank_status,
            pf_status,
        ])

    # Format Joining Date column as text-date and salaries as numeric
    for row in ws.iter_rows(min_row=2, min_col=13, max_col=13):
        for cell in row:
            if cell.value:
                cell.number_format = 'yyyy-mm-dd'
    for row in ws.iter_rows(min_row=2, min_col=15, max_col=16):
        for cell in row:
            cell.number_format = '#,##0.00'

    # Auto-fit column widths based on the longest cell value per column
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            val = cell.value
            if val is None:
                continue
            if hasattr(val, 'strftime'):
                length = len(val.strftime('%Y-%m-%d'))
            else:
                length = len(str(val))
            if length > max_len:
                max_len = length
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 40)

    ws.freeze_panes = 'A2'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename = f"Employee_Master_{timezone.localdate().strftime('%Y-%m-%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@require_POST
def import_excel_employees(request):
    """Stub: accept Excel file upload (full implementation coming soon)."""
    messages.info(request, "Excel import is coming soon. Please use JSON import for now.")
    return redirect('employees:list')


@login_required
@require_POST
def import_json_employees(request):
    """Import employees from a previously exported JSON file.
    Supports both SPIM Suite export keys and SPIM Lite JSON keys.
    Uses update_or_create so re-importing the same file updates rather
    than duplicating records. Bank details are upserted when present."""
    json_file = request.FILES.get('json_file')
    if not json_file:
        messages.error(request, "No file selected.")
        return redirect('employees:list')
    try:
        data = json.loads(json_file.read())
        emp_list = data.get('employees', data) if isinstance(data, dict) else data
        admin_id = get_admin_id(request.user)
        count = 0
        for emp in emp_list:
            if not isinstance(emp, dict) or not emp.get('name'):
                continue

            # ── Field aliases: accept both export and SPIM Lite key names ──
            employee_id = emp.get('employee_id') or str(emp.get('id', ''))
            designation = emp.get('designation') or emp.get('role', '')
            # base_salary: prefer explicit key; fall back to 'salary'
            raw_salary  = emp.get('base_salary')
            if raw_salary is None:
                raw_salary = emp.get('salary', 0)

            employee_obj, _ = Employee.objects.update_or_create(
                admin_id    = admin_id,
                employee_id = employee_id,
                defaults={
                    'name':        emp.get('name', ''),
                    'designation': designation,
                    'department':  emp.get('department', ''),
                    'location':    emp.get('location', ''),
                    'site':        emp.get('site', ''),
                    'base_salary': raw_salary or 0,
                    'status':      emp.get('status', 'active'),
                    'created_by':  request.user,
                    # JSON contains per-employee individual salaries — mark as
                    # custom override so sc_save never silently overwrites them.
                    'salary_is_custom_override': True,
                }
            )
            count += 1

            # ── Bank details: upsert if any bank field is provided ──────────
            bank_name      = emp.get('bank', '')
            account_holder = emp.get('holder', '')
            account_number = emp.get('account', '')
            ifsc_code      = emp.get('ifsc', '')
            if any([bank_name, account_holder, account_number, ifsc_code]):
                BankDetail.objects.update_or_create(
                    employee=employee_obj,
                    defaults={
                        'bank_name':      bank_name,
                        'account_holder': account_holder,
                        'account_number': account_number,
                        'ifsc_code':      ifsc_code,
                    }
                )

        messages.success(request, f"Imported {count} employee(s) successfully.")
    except Exception as e:
        messages.error(request, f"Import failed: {str(e)}")
    return redirect('employees:list')


@login_required
def employee_list_json(request):
    """Return the employee list as JSON for the attendance page live-fetch.
    Bypasses browser HTML caching so salary changes are always reflected
    immediately without requiring a hard reload."""
    admin_id = get_admin_id(request.user)
    employees = Employee.objects.filter(admin_id=admin_id)
    emp_list = []
    for emp in employees:
        # EMP ID must come from the employee_id column only — never the PK.
        # `pk` is an internal-only field that lets the attendance JS resolve
        # the right Employee row when `employee_id` is blank. It is never
        # displayed as the EMP ID.
        emp_list.append({
            'id':           emp.employee_id or '',
            'pk':           emp.pk,
            'name':         emp.name,
            'dept':         emp.department or '',
            'role':         emp.designation or '',
            'mainLocation': emp.location or '',
            'site':         emp.site or '',
            'leave':        '0',
            'baseSalary':   float(emp.base_salary) if emp.base_salary else 0,
            'salaryType':   getattr(emp, 'salary_type', 'base_salary') or 'base_salary',
            # Additive flag used by the attendance bulk table to group vehicle
            # rows under a "Vehicles" heading. Marking/saving is unaffected.
            'isVehicle':    bool(getattr(emp, 'is_vehicle', False)),
        })
    return JsonResponse({'employees': emp_list})


@login_required
def bank_details(request, pk):
    employee = get_object_or_404(Employee, pk=pk, admin_id=get_admin_id(request.user))
    bank_detail, _ = BankDetail.objects.get_or_create(
        employee=employee,
        defaults={'bank_name': '', 'account_holder': '', 'account_number': '', 'ifsc_code': ''},
    )
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if request.method == 'POST':
        if is_ajax:
            # Modal only sends bank_name/account_holder/account_number/ifsc_code.
            # Fill modal-excluded required fields from the current record to prevent
            # form validation failures on fields the modal never submits.
            post_data = request.POST.copy()
            if 'status' not in post_data:
                post_data['status'] = bank_detail.status or 'pending'
            if 'branch' not in post_data:
                post_data['branch'] = bank_detail.branch or ''
            form = BankDetailForm(post_data, instance=bank_detail)
        else:
            form = BankDetailForm(request.POST, instance=bank_detail)
        if form.is_valid():
            form.save()
            if is_ajax:
                return JsonResponse({'success': True})
            messages.success(request, "Bank details updated.")
            return redirect('employees:list')
        if is_ajax:
            return JsonResponse({'success': False, 'errors': form.errors.as_json()}, status=400)
    else:
        form = BankDetailForm(instance=bank_detail)
    return render(request, 'employees/bank_form.html', {'form': form, 'employee': employee})


@login_required
def pf_details(request, pk):
    employee = get_object_or_404(Employee, pk=pk, admin_id=get_admin_id(request.user))
    pf_detail, _ = PFDetail.objects.get_or_create(employee=employee)
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if request.method == 'POST':
        if is_ajax:
            # Modal only sends pf_number and esic_number.
            # Fill modal-excluded required fields from the current record to prevent
            # form validation failures on fields the modal never submits.
            post_data = request.POST.copy()
            if 'status' not in post_data:
                post_data['status'] = pf_detail.status or 'pending'
            if 'uan_number' not in post_data:
                post_data['uan_number'] = pf_detail.uan_number or ''
            if 'employee_contribution' not in post_data:
                post_data['employee_contribution'] = str(pf_detail.employee_contribution or '0')
            if 'employer_contribution' not in post_data:
                post_data['employer_contribution'] = str(pf_detail.employer_contribution or '0')
            form = PFDetailForm(post_data, instance=pf_detail)
        else:
            form = PFDetailForm(request.POST, instance=pf_detail)
        if form.is_valid():
            form.save()
            if is_ajax:
                return JsonResponse({'success': True})
            messages.success(request, "PF details updated.")
            return redirect('employees:list')
        if is_ajax:
            return JsonResponse({'success': False, 'errors': form.errors.as_json()}, status=400)
    else:
        form = PFDetailForm(instance=pf_detail)
    return render(request, 'employees/pf_form.html', {'form': form, 'employee': employee})


@login_required
def salary_management(request, pk):
    employee = get_object_or_404(Employee, pk=pk, admin_id=get_admin_id(request.user))
    salaries = employee.salary_history.all()
    return render(request, 'employees/salary_list.html', {
        'employee': employee,
        'salaries': salaries,
    })


def compute_cycle_salary_total(admin_id, cycle_start, cycle_end):
    """Actual salary accrual for a cycle, matching the Salary Report PDF.

    Iterates every Employee for the tenant, computes per-employee
    `attendance_earnings + ot - advance - ded` (identical formula to
    salary_report_download / _render_salary_pdf) and returns the Decimal
    sum. This is the SINGLE SOURCE OF TRUTH for cycle salary totals
    across the suite — Salary Management's Total Payout tile AND Expense
    Manager's "This Cycle → Salary" tile both consume this so their
    numbers always agree with the PDF report.

    Reads AttendanceRecord directly via _compute_attendance_breakdown
    (NOT [AUTO-SAL:*] Transaction rows), which is why it stays correct
    even for attendance predating the AUTO-SAL live-sync cutover.
    """
    from decimal import Decimal
    target_date = cycle_end.replace(day=1)
    queryset = Employee.objects.filter(admin_id=admin_id)
    total = Decimal('0')
    for e in queryset:
        salary_record = e.salary_history.filter(
            month__year=target_date.year, month__month=target_date.month,
        ).first()
        gross_base = Decimal(str(salary_record.basic_salary if salary_record else (e.base_salary or 0)))
        ot         = Decimal(str(salary_record.ot_allowance    if salary_record else 0))
        advance    = Decimal(str(salary_record.advance_pay     if salary_record else 0))
        ded        = Decimal(str(salary_record.total_deduction if salary_record else 0))
        try:
            earn_dec, _ = _compute_attendance_breakdown(e, target_date, gross_base)
        except Exception:
            earn_dec = gross_base
        total += (earn_dec + ot - advance - ded)
    return total


@login_required
def salary_dashboard(request):
    """
    Main Salary Management Dashboard (The premium multi-filter UI).
    """
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.GET.get('format') == 'json':
        # AJAX JSON Request: Return stats and salary list
        month_name = request.GET.get('month', '')
        year = request.GET.get('year', '')
        
        # Filters from UI
        f_location = request.GET.get('location', '')
        f_site = request.GET.get('site', '')
        f_bank = request.GET.get('bank', '')
        f_search = request.GET.get('search', '').lower()

        admin_id = get_admin_id(request.user)
        queryset = Employee.objects.filter(admin_id=admin_id).order_by('name')
        
        # Apply filters if provided
        if f_location:
            queryset = queryset.filter(location=f_location)
        if f_site:
            queryset = queryset.filter(site=f_site)
        if f_bank:
            queryset = queryset.filter(bank_details__bank_name=f_bank)
        if f_search:
            queryset = queryset.filter(Q(name__icontains=f_search) | Q(employee_id__icontains=f_search))

        emp_list = []
        total_net = 0
        total_ded = 0
        total_ot = 0
        total_advance = 0
        processed_count = 0

        # For selecting specific month salary
        target_date = None
        if month_name and year:
            try:
                m_idx = timezone_month_map(month_name)
                target_date = datetime.date(int(year), m_idx, 1)
            except: pass

        # Case 2 — back-fill base_salary for employees that have no salary on
        # record yet (and no manual override) from the centralized Salary
        # Config. Idempotent: only touches rows where base_salary == 0.
        for e in queryset:
            if (not e.base_salary) and (not e.salary_is_custom_override) and (e.level or '').strip():
                _apply_role_level_salary(e, admin_id, force=False)
                if e.base_salary:
                    e.save(update_fields=['base_salary', 'fixed_allowance'])

        for e in queryset:
            # Find salary record for specific month or last one
            if target_date:
                salary_record = e.salary_history.filter(month__year=target_date.year, month__month=target_date.month).first()
            else:
                salary_record = e.salary_history.order_by('-month').first()
            
            bank = getattr(e, 'bank_details', None)
            pf = getattr(e, 'pf_details', None)
            
            gross_base = float(salary_record.basic_salary if salary_record else e.base_salary)
            ot = float(salary_record.ot_allowance if salary_record else 0)
            advance = float(salary_record.advance_pay if salary_record else 0)
            ded = float(salary_record.total_deduction if salary_record else 0)
            pf_amount = float(salary_record.pf_employee_snapshot if salary_record else 0)
            # Always compute the attendance-prorated payable base so the Edit
            # Salary modal can render the same value as the dashboard list
            # (single source of truth — same helper the Save handler uses).
            # The compute is per (employee, month, base), so it's also valid
            # when a salary_record already exists.
            ae_month = (
                target_date
                or (salary_record.month if salary_record else timezone.now().date().replace(day=1))
            )
            try:
                _earn_dec, _paid_dec = _compute_attendance_breakdown(
                    e, ae_month, gross_base,
                )
                attendance_earnings = float(_earn_dec)
                attendance_paid_days = float(_paid_dec)
            except Exception as exc:
                # Defensive: never let a per-row compute kill the whole
                # dashboard. Surface in logs and fall back to gross_base.
                print(f'[salary_dashboard] attendance compute failed for emp {e.pk}: {exc}')
                attendance_earnings = gross_base
                attendance_paid_days = 0.0

            # Always derive net from the LIVE attendance_earnings so the
            # dashboard reflects current attendance data. salary_record.net_pay
            # is a snapshot stamped at Save time for payslip rendering — using
            # it here would lock the row at a stale value (e.g. ₹0 from an
            # earlier save when attendance had not yet been marked).
            net = attendance_earnings + ot - advance - ded

            if salary_record:
                processed_count += 1

            # KPI accumulators for OT / Advance / Deductions stay scoped to
            # employees with an actual SalaryUpdate record in the cycle.
            # Total Payout (total_net) is NOT accumulated here anymore —
            # it's computed once below from posted [AUTO-SAL:*] Transactions
            # so it matches Expense Manager's "This Cycle → Salary" tile
            # exactly (same query, same data source).
            if salary_record:
                total_ded += ded
                total_ot += ot
                total_advance += advance

            food_allowance = float(salary_record.food_allowance if salary_record else 0)
            food_usage     = float(salary_record.food_usage if salary_record else 0)

            emp_list.append({
                'id': e.id,
                'employee_name': e.name,
                'employee_id': e.employee_id,
                'role': e.designation or '',
                'level': e.level or '',
                'location': e.location,
                'site': e.site,
                'is_payslip_generated': bool(salary_record and salary_record.is_payslip_generated),
                # is_download_active drives the per-row Download Payslip button
                # colour. Green until the 15th of the month AFTER generation
                # (or later, if HR ticked "Reactivate expired downloads"),
                # grey after that. See accounts.cycle_utils.
                'is_download_active': bool(salary_record and salary_record.is_download_active),
                'payslip_id': salary_record.id if salary_record else None,
                'month': month_name or (salary_record.month.strftime('%B') if salary_record else 'N/A'),
                'year': year or (salary_record.month.year if salary_record else 'N/A'),
                'base_salary': float(e.base_salary),
                # Salary Type surfaced so the Edit modal can render the
                # dropdown with the employee's current selection and swap
                # the "Base Salary" label to "Daily Salary" for daily-basis
                # rows. Defaults to 'base_salary' for rows created before
                # the field existed (migration 0012).
                'salary_type': getattr(e, 'salary_type', 'base_salary') or 'base_salary',
                # Additive flag consumed by renderTable() to group vehicle rows
                # under a "Vehicles" section header. Does not change any amount.
                'is_vehicle': bool(getattr(e, 'is_vehicle', False)),
                'salary_is_custom_override': bool(e.salary_is_custom_override),
                'gross_salary': gross_base + ot, # simplified gross
                'advance_pay': advance,
                'deduction': ded,
                'pf_amount': pf_amount,
                'net_payable': net,
                # Server-computed payable base from AttendanceRecord rows for
                # the row's month — Edit Salary modal reads this instead of
                # its localStorage fallback so it always matches the page.
                'attendance_earnings': attendance_earnings,
                # Paid-days breakdown (Present + WeekOff + Holiday + 0.5·HalfDay)
                # surfaced so the Edit Salary modal can display it read-only.
                'paid_days': float(attendance_paid_days),
                'overtime': ot,
                'food_allowance': food_allowance,
                'food_usage': food_usage,
                'bank_name': bank.bank_name if bank else '',
                'account_holder': bank.account_holder if bank else '',
                'account_number': bank.account_number if bank else '',
                'ifsc_code': bank.ifsc_code if bank else '',
                'pf_number': pf.pf_number if pf else '',
                'esi_number': pf.esic_number if pf else '',
            })

        # Total Payout = actual cycle salary accrual — recomputed from
        # AttendanceRecord (same source as the Salary Report PDF) so the
        # KPI never drifts from what the PDF prints. The [AUTO-SAL:*]
        # Transaction pool used previously was incomplete for attendance
        # entered before the live-sync cutover. Cycle window comes from
        # get_salary_cycle(today_ist()) — the canonical 26→25 helper.
        from accounts.cycle_utils import get_salary_cycle as _kpi_get_cycle
        from accounts.date_utils import today_ist as _kpi_today
        _kc = _kpi_get_cycle(_kpi_today())
        total_net = float(compute_cycle_salary_total(admin_id, _kc['start'], _kc['end']))

        return JsonResponse({
            'salaries': emp_list,
            'stats': {
                'total_employees': queryset.count(),
                'processed_count': processed_count,
                'total_net_payable': total_net,
                'total_deductions': total_ded,
                'total_ot': total_ot,
                'total_advance': total_advance,
            },
            'filters': {
                'locations': list(Employee.objects.filter(admin_id=admin_id).values_list('location', flat=True).distinct()),
                'sites': list(Employee.objects.filter(admin_id=admin_id).values_list('site', flat=True).distinct()),
                'banks': list(BankDetail.objects.filter(employee__admin_id=admin_id).values_list('bank_name', flat=True).distinct()),
            }
        })

    # Provide the same master pools the Employee Add/Edit dropdowns use so
    # Salary Config's Role + Level <select>s render from a single source.
    from .models import JobRole as _JR, EmployeeLevel as _EL, SalaryStructure as _SS
    from branches.models import LocationSite as _LS
    def _dedup(values):
        seen = set(); out = []
        for v in values:
            v = (v or '').strip()
            if v and v.lower() not in seen:
                seen.add(v.lower()); out.append(v)
        return sorted(out, key=lambda x: x.lower())
    admin_id = get_admin_id(request.user)
    master_roles = _dedup(
        list(Employee.objects.filter(admin_id=admin_id).values_list('designation', flat=True)) +
        list(_JR.objects.filter(admin_id=admin_id).values_list('name', flat=True))
    )
    master_levels = _dedup(
        list(Employee.objects.filter(admin_id=admin_id).values_list('level', flat=True)) +
        list(_EL.objects.filter(admin_id=admin_id).values_list('name', flat=True)) +
        list(_SS.objects.filter(admin_id=admin_id).values_list('level', flat=True))
    )
    return render(request, 'employees/salary_manager.html', {
        'master_roles':  master_roles,
        'master_levels': master_levels,
    })


# ─── Server-side Salary Report Download ─────────────────────────────────────
# The previous Download Report button relied on jsPDF + SheetJS loaded from
# cdnjs.cloudflare.com. On networks that block cdnjs (corporate proxies,
# aggressive ad-blockers, some Railway egress paths) the libraries never
# parse and the click does nothing. This server-side path uses reportlab
# (PDF) and openpyxl (XLSX), both already pinned in requirements.txt
# (lines 32, 54), so no CDN dependency remains.
#
# The view mirrors salary_dashboard's JSON filter pipeline + per-employee
# salary derivation, then streams a binary file with Content-Disposition:
# attachment. The frontend simply navigates the browser to the URL — no
# client-side JS library involvement at all.
# ────────────────────────────────────────────────────────────────────────────
@login_required
def salary_report_download(request):
    """
    GET /employees/salary/report/download/?format=pdf|xlsx
        &month=June&year=2026
        &location=...&site=...&bank=...&search=...

    Returns a PDF (default) or XLSX file containing the same salary rows
    the Salary Management page is currently showing, with these 8 columns:
        EMP ID | EMP NAME | SITE | SALARY (₹) | BANK NAME | ACCOUNT HOLDER
        | ACCOUNT NUMBER | IFSC CODE
    Filename: Salary_Report_<Month>_<Year>.{pdf,xlsx}

    Auth note: @login_required protects this view against unauthenticated
    callers. If reached without a session it will 302 to LOGIN_URL (see
    config/settings.py). In practice this can only happen if a user's
    session expired between the page render and the download click — the
    salary_manager page itself is @login_required too, so an unauthenticated
    user can't see the Download button in the first place. The frontend
    uses an <a download> click pattern that does NOT replace the current
    URL, so if Django redirects to login the user stays on the salary
    page and can re-authenticate without losing context.
    """
    fmt = (request.GET.get('format') or 'pdf').strip().lower()
    if fmt not in ('pdf', 'xlsx'):
        fmt = 'pdf'

    month_name = (request.GET.get('month') or '').strip()
    year       = (request.GET.get('year') or '').strip()
    f_location = request.GET.get('location', '')
    f_site     = request.GET.get('site', '')
    f_bank     = request.GET.get('bank', '')
    f_search   = (request.GET.get('search', '') or '').lower()

    admin_id = get_admin_id(request.user)
    queryset = Employee.objects.filter(admin_id=admin_id).order_by('name')
    if f_location:
        queryset = queryset.filter(location=f_location)
    if f_site:
        queryset = queryset.filter(site=f_site)
    if f_bank:
        queryset = queryset.filter(bank_details__bank_name=f_bank)
    if f_search:
        queryset = queryset.filter(
            Q(name__icontains=f_search) | Q(employee_id__icontains=f_search)
        )

    target_date = None
    if month_name and year:
        try:
            m_idx = timezone_month_map(month_name)
            target_date = datetime.date(int(year), m_idx, 1)
        except Exception:
            target_date = None

    # Build rows — same derivation as salary_dashboard JSON branch.
    rows = []
    total_net = 0.0
    for e in queryset:
        # Mirror dashboard's idempotent base-salary back-fill so the download
        # never reports ₹0 for an employee just because their base hasn't
        # been viewed in the dashboard yet.
        if (not e.base_salary) and (not e.salary_is_custom_override) and (e.level or '').strip():
            _apply_role_level_salary(e, admin_id, force=False)
            if e.base_salary:
                e.save(update_fields=['base_salary', 'fixed_allowance'])

        if target_date:
            salary_record = e.salary_history.filter(
                month__year=target_date.year, month__month=target_date.month
            ).first()
        else:
            salary_record = e.salary_history.order_by('-month').first()

        gross_base    = float(salary_record.basic_salary if salary_record else e.base_salary)
        ot            = float(salary_record.ot_allowance if salary_record else 0)
        extra_allow   = float(salary_record.extra_allowance if salary_record else 0)
        advance       = float(salary_record.advance_pay if salary_record else 0)
        ded           = float(salary_record.total_deduction if salary_record else 0)
        pf_amount     = float(salary_record.pf_employee_snapshot if salary_record else 0)
        food_allow    = float(salary_record.food_allowance if salary_record else 0)
        food_usage    = float(salary_record.food_usage if salary_record else 0)
        ae_month      = (
            target_date
            or (salary_record.month if salary_record else timezone.now().date().replace(day=1))
        )
        try:
            earn_dec, paid_dec = _compute_attendance_breakdown(e, ae_month, gross_base)
            attendance_earnings = float(earn_dec)
            paid_days           = float(paid_dec)
        except Exception as exc:
            print(f'[salary_report_download] attendance compute failed for emp {e.pk}: {exc}')
            attendance_earnings = gross_base
            paid_days           = 0.0
        net = attendance_earnings + ot - advance - ded

        # Report column 19: Total Deduction per spec — advance + PF + any
        # food-usage overshoot. Distinct from the SalaryUpdate.total_deduction
        # snapshot (which is HR-entered generic deductions).
        food_overshoot = max(0.0, food_usage - food_allow)
        report_total_deduction = advance + pf_amount + food_overshoot

        salary_type = getattr(e, 'salary_type', 'base_salary') or 'base_salary'
        is_daily    = salary_type == 'daily_basis'
        emp_base    = float(e.base_salary or 0)

        bank = getattr(e, 'bank_details', None)
        rows.append({
            'emp_id':         e.employee_id or '',
            'emp_name':       e.name or '',
            'site':           e.site or '',
            'salary':         net,
            'bank_name':      bank.bank_name      if bank else '',
            'account_holder': bank.account_holder if bank else '',
            'account_number': bank.account_number if bank else '',
            'ifsc_code':      bank.ifsc_code      if bank else '',
            'is_vehicle':     bool(getattr(e, 'is_vehicle', False)),
            # Extended report-only columns (9-20). Not surfaced in the on-
            # screen table — the download output is the ONLY consumer.
            'salary_type_label':    'Daily Pay' if is_daily else 'Base Pay',
            'base_salary_col':      '' if is_daily else emp_base,
            'daily_salary_col':     emp_base if is_daily else '',
            'paid_days':            paid_days,
            'attendance_earnings':  attendance_earnings,
            'ot_extra':             ot + extra_allow,
            'advance_pay':          advance,
            'pf_amount':            pf_amount,
            'food_allowance':       food_allow,
            'food_usage':           food_usage,
            'total_deduction':      report_total_deduction,
            'net_pay':              net,
        })
        total_net += net

    # Group people first, then vehicles, preserving name order within each
    # group (queryset is already name-ordered; a stable sort keeps that).
    # total_net stays a single combined total — no per-group subtotals.
    rows.sort(key=lambda r: r['is_vehicle'])

    safe_month = (month_name or 'AllMonths').replace(' ', '_')
    safe_year  = (str(year) if year else 'AllYears').replace(' ', '_')
    filename_base = f"Salary_Report_{safe_month}_{safe_year}"
    period_label  = f"{month_name} {year}".strip() or "All Periods"

    if fmt == 'xlsx':
        return _render_salary_xlsx(rows, total_net, filename_base, period_label)
    return _render_salary_pdf(rows, total_net, filename_base, period_label)


def _render_salary_xlsx(rows, total_net, filename_base, period_label):
    """Build the salary report as an XLSX HttpResponse using openpyxl."""
    from io import BytesIO
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = 'Salary Report'

    headers = [
        'EMP ID', 'EMP NAME', 'SITE',
        'BANK NAME', 'ACCOUNT HOLDER', 'ACCOUNT NUMBER', 'IFSC CODE',
        'SALARY (₹)',
        # Extended columns (report-only)
        'SALARY TYPE', 'BASE SALARY (₹)', 'DAILY SALARY (₹)', 'PAID DAYS',
        'ATTENDANCE EARNINGS (₹)', 'OT / EXTRA (₹)', 'ADVANCE PAY (₹)',
        'PF AMOUNT (₹)', 'FOOD ALLOWANCE (₹)', 'ACTUAL FOOD USAGE (₹)',
        'TOTAL DEDUCTION (₹)', 'NET PAY (₹)',
    ]
    ws.append(headers)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1E293B')
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')

    # Full-width group-header row (merged across all columns). Only emitted
    # for a non-empty group — no empty "Vehicles" heading is ever written.
    group_fill = PatternFill('solid', fgColor='E2E8F0')
    group_font = Font(bold=True, color='0F172A')

    def _append_section(label):
        ws.append([label])
        ridx = ws.max_row
        ws.merge_cells(start_row=ridx, start_column=1, end_row=ridx, end_column=len(headers))
        cell = ws.cell(row=ridx, column=1)
        cell.font = group_font
        cell.fill = group_fill
        cell.alignment = Alignment(horizontal='left')

    def _append_data(r):
        def _n(v):
            return '' if v == '' or v is None else float(v)
        ws.append([
            r['emp_id'],
            r['emp_name'],
            r['site'],
            r['bank_name'],
            r['account_holder'],
            r['account_number'],
            r['ifsc_code'],
            float(r['salary'] or 0),
            r.get('salary_type_label', ''),
            _n(r.get('base_salary_col', '')),
            _n(r.get('daily_salary_col', '')),
            float(r.get('paid_days') or 0),
            float(r.get('attendance_earnings') or 0),
            float(r.get('ot_extra') or 0),
            float(r.get('advance_pay') or 0),
            float(r.get('pf_amount') or 0),
            float(r.get('food_allowance') or 0),
            float(r.get('food_usage') or 0),
            float(r.get('total_deduction') or 0),
            float(r.get('net_pay') or 0),
        ])

    people   = [r for r in rows if not r['is_vehicle']]
    vehicles = [r for r in rows if r['is_vehicle']]
    # "Employees" header only when vehicles also exist (a single flat list
    # keeps its original headerless look otherwise).
    if people and vehicles:
        _append_section('Employees')
    for r in people:
        _append_data(r)
    if vehicles:
        _append_section('Vehicles')
        for r in vehicles:
            _append_data(r)

    # Blank separator + totals footer. Only column 8 (SALARY) and 20
    # (NET PAY) carry monetary totals; other extended columns stay blank.
    ws.append([])
    total_row = [''] * len(headers)
    total_row[0]  = f'Total Employees: {len(rows)}'
    total_row[4]  = 'Total Salary Paid'
    total_row[7]  = float(total_net or 0)
    total_row[19] = float(total_net or 0)
    ws.append(total_row)

    widths = [
        14, 24, 18, 22, 26, 22, 18, 14,
        14, 16, 16, 12, 20, 16, 16, 14, 18, 20, 18, 14,
    ]
    from openpyxl.utils import get_column_letter as _colletter
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[_colletter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
    return response


def _render_salary_pdf(rows, total_net, filename_base, period_label):
    """Build the salary report as a landscape A4 PDF HttpResponse using reportlab."""
    from io import BytesIO
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import landscape, A3
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    # Page size: A3 landscape (1190.55 x 841.89 pt).
    # A4 landscape only offers ~841 pt wide; even after narrowing margins to
    # 10pt and shrinking every column to its legibility floor, the 20-column
    # salary report sums to ~1140+ pt. A4 clipped Emp ID / Emp Name / Total
    # Deduction / Net Pay into the page margins. A3 gives 1190pt so the same
    # widths render fully with a small right-side buffer. XLSX renderer is
    # unchanged — this only affects the printable PDF snapshot.
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A3),
        leftMargin=10, rightMargin=10, topMargin=18, bottomMargin=18,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title14', parent=styles['Title'], fontSize=14, alignment=0)
    sub_style   = ParagraphStyle('Sub8',    parent=styles['Normal'], fontSize=8)

    story = [
        Paragraph(f"Salary Report — {period_label}", title_style),
        Paragraph(
            f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}",
            sub_style,
        ),
        Spacer(1, 6),
    ]

    table_data = [[
        'EMP ID', 'EMP NAME', 'SITE',
        'BANK NAME', 'ACCOUNT HOLDER', 'ACCOUNT NUMBER', 'IFSC CODE',
        'SALARY (Rs)',
        'SALARY TYPE', 'BASE (Rs)', 'DAILY (Rs)', 'PAID DAYS',
        'ATT EARN (Rs)', 'OT/EXTRA (Rs)', 'ADVANCE (Rs)',
        'PF (Rs)', 'FOOD ALLOW (Rs)', 'FOOD USED (Rs)',
        'TOT DED (Rs)', 'NET PAY (Rs)',
    ]]
    NUM_COLS = len(table_data[0])

    def _money_or_blank(v):
        if v == '' or v is None:
            return '-'
        return '{:,.2f}'.format(float(v))

    def _data_row(r):
        return [
            r['emp_id']         or '-',
            r['emp_name']       or '-',
            r['site']           or '-',
            r['bank_name']      or '-',
            r['account_holder'] or '-',
            r['account_number'] or '-',
            r['ifsc_code']      or '-',
            '{:,.2f}'.format(float(r['salary'] or 0)),
            r.get('salary_type_label', '-') or '-',
            _money_or_blank(r.get('base_salary_col', '')),
            _money_or_blank(r.get('daily_salary_col', '')),
            '{:g}'.format(float(r.get('paid_days') or 0)),
            '{:,.2f}'.format(float(r.get('attendance_earnings') or 0)),
            '{:,.2f}'.format(float(r.get('ot_extra') or 0)),
            '{:,.2f}'.format(float(r.get('advance_pay') or 0)),
            '{:,.2f}'.format(float(r.get('pf_amount') or 0)),
            '{:,.2f}'.format(float(r.get('food_allowance') or 0)),
            '{:,.2f}'.format(float(r.get('food_usage') or 0)),
            '{:,.2f}'.format(float(r.get('total_deduction') or 0)),
            '{:,.2f}'.format(float(r.get('net_pay') or 0)),
        ]

    people   = [r for r in rows if not r['is_vehicle']]
    vehicles = [r for r in rows if r['is_vehicle']]
    # Track which table_data row indices are full-width section headers so
    # their SPAN + background styles can be applied after the table is built.
    section_header_idx = []

    def _append_section(label):
        section_header_idx.append(len(table_data))
        table_data.append([label] + [''] * (NUM_COLS - 1))

    # "Employees" header only when vehicles also exist; "Vehicles" header
    # omitted entirely when there are no vehicles — no empty headings.
    if people and vehicles:
        _append_section('Employees')
    for r in people:
        table_data.append(_data_row(r))
    if vehicles:
        _append_section('Vehicles')
        for r in vehicles:
            table_data.append(_data_row(r))

    total_row_cells = [''] * NUM_COLS
    total_row_cells[0]  = f'Total Employees: {len(rows)}'
    total_row_cells[4]  = 'Total Salary Paid'
    total_row_cells[7]  = '{:,.2f}'.format(float(total_net or 0))
    total_row_cells[19] = '{:,.2f}'.format(float(total_net or 0))
    table_data.append(total_row_cells)

    # 20-column landscape A3 layout. Widths respect the design caps:
    #   text cols <= 75pt, amount cols <= 55pt, Emp Name >= 90pt.
    # Sum = 1165pt vs A3 landscape printable width 1170.55pt (10pt margins).
    col_widths = [
        45,  # 0  EMP ID
        90,  # 1  EMP NAME
        60,  # 2  SITE
        70,  # 3  BANK NAME
        75,  # 4  ACCOUNT HOLDER
        70,  # 5  ACCOUNT NUMBER
        65,  # 6  IFSC CODE
        55,  # 7  SALARY
        50,  # 8  SALARY TYPE
        55,  # 9  BASE
        55,  # 10 DAILY
        40,  # 11 PAID DAYS
        55,  # 12 ATT EARN
        55,  # 13 OT/EXTRA
        55,  # 14 ADVANCE
        50,  # 15 PF
        55,  # 16 FOOD ALLOW
        55,  # 17 FOOD USED
        55,  # 18 TOT DED
        55,  # 19 NET PAY
    ]
    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0),  colors.HexColor('#1E293B')),
        ('TEXTCOLOR',  (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',   (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 5.5),
        # Right-align numeric columns in data rows: salary(7) + 9..19
        ('ALIGN',      (7, 1),  (7, -1),  'RIGHT'),
        ('ALIGN',      (9, 1),  (-1, -1), 'RIGHT'),
        ('VALIGN',     (0, 0),  (-1, -1), 'MIDDLE'),
        ('GRID',       (0, 0),  (-1, -1), 0.25, colors.HexColor('#CBD5E1')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F0FDF4')),
        ('TEXTCOLOR',  (0, -1), (-1, -1), colors.HexColor('#059669')),
        ('FONTNAME',   (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]
    # Merge each section-header row across all 8 columns and shade it. The
    # right-align on column 7 does not apply to these rows because the SPAN
    # makes the first cell own the full width.
    for ridx in section_header_idx:
        style_cmds.append(('SPAN', (0, ridx), (-1, ridx)))
        style_cmds.append(('BACKGROUND', (0, ridx), (-1, ridx), colors.HexColor('#E2E8F0')))
        style_cmds.append(('TEXTCOLOR', (0, ridx), (-1, ridx), colors.HexColor('#0F172A')))
        style_cmds.append(('FONTNAME', (0, ridx), (-1, ridx), 'Helvetica-Bold'))
        style_cmds.append(('ALIGN', (0, ridx), (-1, ridx), 'LEFT'))
    tbl.setStyle(TableStyle(style_cmds))
    story.append(tbl)
    doc.build(story)

    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
    return response


def _attendance_day_weight(status, is_daily):
    """Per-day pay weight for a single AttendanceRecord.status. Shared with
    _compute_attendance_breakdown and the attendance→expense signal so
    monthly totals and per-day auto-expense rows can never diverge."""
    if is_daily:
        if status in ('present', 'week_off'):
            return Decimal('1')
        return Decimal('0')
    if status in ('present', 'week_off', 'holiday'):
        return Decimal('1')
    if status == 'half_day':
        return Decimal('0.5')
    return Decimal('0')


def compute_daily_salary_for_attendance(attendance):
    """Rupee amount payable for a single AttendanceRecord day. Uses the
    same weight table + 26→25 cycle divisor as the monthly Salary Manager
    helper so `sum(daily rows) == monthly total`."""
    from accounts.cycle_utils import get_cycle_ending_in_month

    emp = attendance.employee
    is_daily = getattr(emp, 'salary_type', 'base_salary') == 'daily_basis'
    base = Decimal(str(emp.base_salary or 0))
    weight = _attendance_day_weight(attendance.status, is_daily)
    if weight == 0 or base == 0:
        return Decimal('0.00')

    if is_daily:
        return (base * weight).quantize(Decimal('0.01'))

    cycle = get_cycle_ending_in_month(attendance.date)
    cycle_days = (cycle['end'] - cycle['start']).days + 1
    if cycle_days <= 0:
        return Decimal('0.00')
    daily_rate = base / Decimal(str(cycle_days))
    return (daily_rate * weight).quantize(Decimal('0.01'))


def _compute_attendance_breakdown(employee, month_date, basic_salary):
    """
    Compute both the attendance-prorated payable base and the paid-days
    count for the given (employee, payroll month). Returns (final_salary,
    paid_days) both as Decimal.

    Cycle window: 26th of the previous month → 25th of `month_date`'s
    month — the SPIM 26→25 attendance cycle, matching the frontend
    attendance calendar.

    Future dates are excluded entirely (no pay for days that haven't
    happened yet). Future Sundays auto-marked Holiday by the backend
    backfill are also excluded by the explicit Sunday rule, kept as a
    defensive guard.

    See _compute_attendance_earnings for the formula and fallback semantics.
    """
    from attendance.models import AttendanceRecord
    from accounts.cycle_utils import get_cycle_ending_in_month

    # Coerce to Decimal to prevent float × Decimal TypeError
    basic_salary = Decimal(str(basic_salary))

    # ---- 26→25 cycle anchored on month_date's END month ----
    # month_date names the payroll month (the month the cycle ENDS in).
    # get_cycle_ending_in_month is deterministic w.r.t. day-of-month, so
    # this returns the same window whether callers pass day=1, day=25, or
    # day=31. Do NOT swap this for get_salary_cycle — that one is day-
    # sensitive and would flip to "next cycle" whenever month_date.day>=26.
    cycle = get_cycle_ending_in_month(month_date)
    cycle_start = cycle['start']
    cycle_end   = cycle['end']

    # ---- CycleDays = actual calendar days in the cycle window ----
    cycle_days = (cycle_end - cycle_start).days + 1

    if cycle_days <= 0:
        # Defensive: never divide by zero.
        return basic_salary, Decimal('0')

    today = datetime.date.today()

    records = AttendanceRecord.objects.filter(
        employee=employee,
        date__gte=cycle_start,
        date__lte=cycle_end,
    ).filter(date__lte=today)

    # Salary Type branch. 'daily_basis' employees never receive the fixed
    # monthly-salary fallback and never divide by cycle_days — their
    # `base_salary` field holds a per-day rate and earnings are literally
    # rate × paid_days. 'base_salary' path below is byte-for-byte the
    # existing formula.
    is_daily = getattr(employee, 'salary_type', 'base_salary') == 'daily_basis'

    if not records.exists():
        if is_daily:
            # Daily-basis: no attendance → no earnings. Never surface the
            # per-day rate as a monthly "safe fallback".
            return Decimal('0'), Decimal('0')
        # No attendance marked — treat as fully paid (safe fallback).
        # Paid-days surfaced as CycleDays so the modal stays consistent
        # with the full-base earnings figure.
        return basic_salary, Decimal(str(cycle_days))

    # Defensive Sunday rule: even if a future-Sunday row leaked past the
    # date__lte=today filter, do not credit it.
    records_paid = records.exclude(
        date__week_day=1, date__gt=today,
    )

    present_days  = records_paid.filter(status='present').count()
    holiday_days  = records_paid.filter(status='holiday').count()
    half_days     = records_paid.filter(status='half_day').count()
    week_off_days = records_paid.filter(status='week_off').count()

    # Weights sourced from the shared _attendance_day_weight table so the
    # per-day auto-expense signal (attendance/signals.py) never diverges
    # from this monthly aggregate.
    w_present  = _attendance_day_weight('present',  is_daily)
    w_week_off = _attendance_day_weight('week_off', is_daily)
    w_holiday  = _attendance_day_weight('holiday',  is_daily)
    w_half_day = _attendance_day_weight('half_day', is_daily)

    paid_days = (
        Decimal(str(present_days))  * w_present
        + Decimal(str(week_off_days)) * w_week_off
        + Decimal(str(holiday_days))  * w_holiday
        + Decimal(str(half_days))     * w_half_day
    )

    if is_daily:
        # Daily basis: no cycle divisor — basic_salary is a per-day rate.
        final_salary = (basic_salary * paid_days).quantize(Decimal('0.01'))
        return final_salary, paid_days

    cycle_dec    = Decimal(str(cycle_days))
    daily_salary = basic_salary / cycle_dec
    final_salary = (daily_salary * paid_days).quantize(Decimal('0.01'))
    return final_salary, paid_days


def _compute_attendance_earnings(employee, month_date, basic_salary):
    """
    Attendance-prorated payable base from AttendanceRecord rows.

    Divisor:
      CycleDays = full calendar days in the month (28 / 29 / 30 / 31).

    Status pay weights (per business rule):
      - present     → 1.0  (full day pay)
      - week_off    → 1.0  (full day pay)
      - holiday     → 1.0  (full day pay; includes Sundays auto-marked Holiday)
      - half_day    → 0.5
      - leave       → 0.0
      - no_week_off → 0.0

    PaidDays    = Present + WeekOff + Holiday + 0.5·HalfDay
    DailySalary = MonthlySalary / CycleDays
    FinalSalary = DailySalary × PaidDays

    Safe fallback: if no attendance records exist for the month, return
    the full basic_salary so payroll is never silently zeroed.

    The import of AttendanceRecord is deferred to avoid a circular import:
    attendance.models imports Employee from employees.models.
    """
    final_salary, _ = _compute_attendance_breakdown(employee, month_date, basic_salary)
    return final_salary


@login_required
@require_POST
def manage_ajax(request):
    """
    Unified AJAX endpoint for Salary Manager (salary_manager.html).
    """
    import json
    try:
        data = json.loads(request.body)
        emp_id = data.get('id')
        
        # 1. DELETE ACTION
        if data.get('action') == 'delete':
            employee = get_object_or_404(Employee, pk=emp_id, admin_id=get_admin_id(request.user))
            employee.delete()
            return JsonResponse({'success': True})

        # 1z. SALARY CONFIG CRUD ACTIONS (Tasks 3 + 4)
        # Body: { "action": "sc_list" }
        #     -> { success, configs: [{id, role, level, base_salary}, ...] }
        # Body: { "action": "sc_save", "role": "...", "level": "...", "base_salary": 30000 }
        #     -> upserts SalaryStructure(admin_id, JobRole(name=role), level)
        # Body: { "action": "sc_delete", "id": <pk> }
        #     -> deletes the SalaryStructure row (scoped to admin_id)
        # Source of truth: employees.SalaryStructure. Auto-applied to Employee
        # at Add and on role/level change via _apply_role_level_salary().
        if data.get('action') == 'sc_list':
            admin_id = get_admin_id(request.user)
            rows = (SalaryStructure.objects
                    .filter(admin_id=admin_id)
                    .select_related('job_role')
                    .order_by('job_role__name', 'level'))
            return JsonResponse({
                'success': True,
                'configs': [{
                    'id':          s.id,
                    'role':        s.job_role.name if s.job_role_id else '',
                    'level':       s.level,
                    'base_salary': float(s.base_salary),
                } for s in rows],
            })

        if data.get('action') == 'sc_save':
            role  = (data.get('role')  or '').strip()
            level = (data.get('level') or '').strip()
            try:
                base  = Decimal(str(data.get('base_salary') or 0))
            except Exception:
                base = Decimal('0')
            if not role or not level:
                return JsonResponse({'success': False, 'error': 'Role and Level are required.'}, status=400)
            if base <= 0:
                return JsonResponse({'success': False, 'error': 'Base salary must be greater than 0.'}, status=400)
            admin_id = get_admin_id(request.user)
            # All three DB writes are atomic: if the Employee bulk-update fails
            # the SalaryStructure write rolls back too — no partial state.
            with transaction.atomic():
                # Reuse-or-create JobRole by name (centralized role master).
                job_role, _ = JobRole.objects.get_or_create(admin_id=admin_id, name=role)
                structure, created = SalaryStructure.objects.update_or_create(
                    admin_id=admin_id, job_role=job_role, level=level,
                    defaults={'base_salary': base},
                )
                # Push the new base to every employee that already matches this
                # mapping so SPIM Lite Salary / Payslip generation reflect right
                # away (Task 4 — "Prevent stale values"). Employees with a manual
                # salary override are deliberately skipped (Case 3 of the
                # override spec).
                Employee.objects.filter(
                    admin_id=admin_id, level=level, job_role=job_role,
                    salary_is_custom_override=False,
                ).update(base_salary=base)
                Employee.objects.filter(
                    admin_id=admin_id, level=level, designation__iexact=role,
                    salary_is_custom_override=False,
                ).update(base_salary=base)
            return JsonResponse({
                'success': True, 'created': created,
                'config': {
                    'id': structure.id, 'role': role,
                    'level': level, 'base_salary': float(base),
                },
            })

        if data.get('action') == 'sc_delete':
            sid = data.get('id')
            if not sid:
                return JsonResponse({'success': False, 'error': 'id is required.'}, status=400)
            admin_id = get_admin_id(request.user)
            SalaryStructure.objects.filter(admin_id=admin_id, pk=sid).delete()
            return JsonResponse({'success': True})

        # 1a. MASTER ADD ACTION
        # Body: { "action": "master_add", "kind": "role|level|location|site", "value": "..." }
        # Upserts the new master value into the existing model for the
        # current tenant. Returns the canonical value so the client can
        # append it to the dropdown without a page reload.
        if data.get('action') == 'master_add':
            kind  = (data.get('kind')  or '').strip().lower()
            value = (data.get('value') or '').strip()
            if not value:
                return JsonResponse({'success': False, 'error': 'Value cannot be empty.'}, status=400)
            admin_id = get_admin_id(request.user)
            if kind == 'role':
                JobRole.objects.get_or_create(admin_id=admin_id, name=value)
            elif kind == 'level':
                EmployeeLevel.objects.get_or_create(admin_id=admin_id, name=value)
            elif kind in ('location', 'site'):
                # LocationSite holds combined "<location> / <site>" entries.
                # When admin adds a bare location or bare site, store it as
                # a standalone name so it shows up in either dropdown's pool.
                from branches.models import LocationSite as _LS
                _LS.objects.get_or_create(admin_id=admin_id, name=value, defaults={'created_by': request.user})
            else:
                return JsonResponse({'success': False, 'error': f'Unknown kind: {kind}'}, status=400)
            return JsonResponse({'success': True, 'kind': kind, 'value': value})

        # 1c. BATCH GENERATE PAYSLIPS ACTION
        # Body: { "action": "generate_payslips_batch", "month": "January", "year": "2026" }
        # Single-click flow:
        #   1. Flip is_payslip_generated on every SalaryUpdate row for the
        #      tenant + cycle. Skips rows that are already generated so a
        #      repeat click is a no-op for that subset.
        #   2. Salary expense rows are NO LONGER written here — they are
        #      posted per-day by attendance/signals.py against each
        #      AttendanceRecord (marker: "[AUTO-SAL:{pk}]"). This endpoint
        #      only locks the payslip snapshot.
        #   3. If every row in the cycle was already generated, return an
        #      early "already generated" message without touching anything.
        if data.get('action') == 'generate_payslips_batch':
            from accounts.cycle_utils import (
                get_salary_cycle, get_previous_cycle, get_force_active_until,
            )
            import logging as _logging
            _log = _logging.getLogger(__name__)

            admin_id       = get_admin_id(request.user)
            month_name     = (data.get('month') or '').strip()
            year_str       = (data.get('year') or '').strip()
            cycle_month_key = (data.get('cycle_month_key') or '').strip()
            reactivate     = bool(data.get('reactivate_grey'))

            # Resolve the target cycle. cycle_month_key ('YYYY-MM') wins;
            # otherwise fall back to month+year; otherwise use the most
            # recent completed cycle from today (IST).
            try:
                if cycle_month_key:
                    y_s, m_s = cycle_month_key.split('-', 1)
                    cycle = get_salary_cycle(datetime.date(int(y_s), int(m_s), 25))
                elif month_name and year_str:
                    m_idx = timezone_month_map(month_name)
                    cycle = get_salary_cycle(datetime.date(int(year_str), m_idx, 25))
                else:
                    cycle = get_previous_cycle()
            except (ValueError, TypeError):
                return JsonResponse({'success': False, 'error': 'Invalid cycle month.'}, status=400)

            target_date = cycle['end'].replace(day=1)   # payroll month first-of-month
            month_name  = target_date.strftime('%B')
            year_str    = str(target_date.year)

            # Include every employee under this admin — same scope the
            # salary_manager dashboard renders. A `status='active'` filter
            # here previously left inactive / on_leave rows permanently grey
            # even after HR clicked Generate Payslip, because they still
            # showed up in the dashboard but were silently skipped by the
            # batch. Attendance-driven earnings + salary_record.net_pay
            # already handle "no attendance = ₹0", so it is safe to include
            # every dashboard row.
            active_emps = Employee.objects.filter(admin_id=admin_id)
            if not active_emps.exists():
                return JsonResponse({
                    'success': False,
                    'error': 'No employees under this admin.',
                }, status=400)

            gen_time    = timezone.now()
            force_until = get_force_active_until(gen_time.date()) if reactivate else None
            generated   = 0
            reactivated = 0
            recomputed_deltas = 0
            errors      = []

            for emp in active_emps:
                try:
                    with transaction.atomic():
                        salary_record, created = SalaryUpdate.objects.get_or_create(
                            employee=emp,
                            month=target_date,
                            defaults={
                                'admin_id':     admin_id,
                                'basic_salary': emp.base_salary or Decimal('0'),
                                'food_allowance': emp.fixed_allowance or Decimal('0'),
                                'created_by':   request.user,
                            },
                        )
                        # Ensure admin_id is set on rows created before that
                        # field carried an explicit default.
                        if not salary_record.admin_id or salary_record.admin_id == 'PENDING':
                            salary_record.admin_id = admin_id

                        # Bug 3B: recompute net_pay from CURRENT attendance
                        # before locking, so the payslip snapshot matches
                        # what HR saw at generation time. Warn on drift.
                        basic          = Decimal(str(salary_record.basic_salary or 0))
                        ot             = Decimal(str(salary_record.ot_allowance or 0))
                        advance        = Decimal(str(salary_record.advance_pay or 0))
                        deduction      = Decimal(str(salary_record.total_deduction or 0))
                        food_allowance = Decimal(str(salary_record.food_allowance or 0))
                        food_usage     = Decimal(str(salary_record.food_usage or 0))
                        food_adj       = food_allowance - food_usage

                        payable_base = _compute_attendance_earnings(emp, target_date, basic)
                        new_net_pay  = payable_base + ot - advance - deduction + food_adj

                        prev_net = Decimal(str(salary_record.net_pay or 0))
                        if abs(new_net_pay - prev_net) > Decimal('1'):
                            recomputed_deltas += 1
                            _log.warning(
                                "Payslip net_pay drift for emp=%s cycle=%s: "
                                "stored=%s recomputed=%s (attendance changed since Save)",
                                emp.pk, target_date.isoformat(), prev_net, new_net_pay,
                            )
                        salary_record.net_pay = new_net_pay

                        # Lock the payslip. Reactivate override picks a new
                        # force_active_until when HR asked for it.
                        was_locked = salary_record.is_payslip_generated
                        salary_record.is_payslip_generated = True
                        salary_record.payslip_generated_at = gen_time
                        salary_record.payslip_generated_by = request.user
                        if reactivate:
                            salary_record.payslip_force_active_until = force_until
                        salary_record.save()

                        if was_locked and reactivate:
                            reactivated += 1
                        elif not was_locked:
                            generated += 1
                except Exception as exc:
                    _log.exception(
                        "generate_payslips_batch failed for emp=%s cycle=%s",
                        emp.pk, target_date.isoformat(),
                    )
                    errors.append({
                        'employee_id': emp.pk,
                        'employee_name': emp.name,
                        'error': str(exc),
                    })

            # Monthly salary expense generation retired — expense rows are
            # now written per-day by attendance/signals.py against each
            # AttendanceRecord. This batch endpoint only locks payslips.
            already_generated = (generated == 0 and reactivated == 0 and not errors)

            msg = (
                f"Payslips generated for {cycle['label']}. "
                f"{generated} newly generated"
                + (f", {reactivated} reactivated" if reactivated else '')
                + "."
            )
            if errors:
                sample = ', '.join(
                    f"{e['employee_name']} ({e['error'][:60]})" for e in errors[:3]
                )
                more = f" and {len(errors) - 3} more" if len(errors) > 3 else ''
                msg += f" — {len(errors)} FAILED: {sample}{more}. See server logs."

            return JsonResponse({
                'success': True,
                'already_generated': already_generated,
                'generated': generated,
                'reactivated': reactivated,
                'recomputed_deltas': recomputed_deltas,
                'cycle_label': cycle['label'],
                'cycle_month_key': cycle['month_key'],
                'errors': errors,
                'message': msg,
            })

        # 1b. GENERATE PAYSLIP ACTION (Task 4)
        # Body: { "action": "generate_payslip", "id": <employee_pk>, "month": "January", "year": "2026" }
        # Flips the payslip lock on the SalaryUpdate row for the given cycle.
        # Employee can only download once this is set.
        if data.get('action') == 'generate_payslip':
            employee = get_object_or_404(Employee, pk=emp_id, admin_id=get_admin_id(request.user))
            month_name = (data.get('month') or '').strip()
            year_str   = (data.get('year') or '').strip()
            if not month_name or not year_str:
                return JsonResponse({'success': False, 'error': 'month and year are required.'}, status=400)
            try:
                target_date = datetime.date(int(year_str), timezone_month_map(month_name), 1)
            except (ValueError, TypeError):
                return JsonResponse({'success': False, 'error': 'Invalid month/year.'}, status=400)
            salary_record = SalaryUpdate.objects.filter(
                employee=employee, month=target_date,
            ).first()
            if not salary_record:
                return JsonResponse({
                    'success': False,
                    'error': 'No salary record for this cycle. Process salary before generating the payslip.',
                }, status=400)
            salary_record.is_payslip_generated = True
            salary_record.payslip_generated_at = timezone.now()
            salary_record.payslip_generated_by = request.user
            salary_record.save(update_fields=[
                'is_payslip_generated', 'payslip_generated_at', 'payslip_generated_by',
            ])
            return JsonResponse({'success': True, 'payslip_id': salary_record.id, 'is_generated': True})
            
        # 2. BANK UPDATE
        if 'bank' in data:
            employee = get_object_or_404(Employee, pk=emp_id, admin_id=get_admin_id(request.user))
            bank_data = data['bank']
            bank_detail, _ = BankDetail.objects.get_or_create(employee=employee)
            bank_detail.bank_name = bank_data.get('bank_name', '')
            bank_detail.account_holder = bank_data.get('holder', '')
            bank_detail.account_number = bank_data.get('account', '')
            bank_detail.ifsc_code = bank_data.get('ifsc', '')
            bank_detail.save()
            return JsonResponse({'success': True})

        # 3. PF UPDATE
        if 'pf' in data:
            employee = get_object_or_404(Employee, pk=emp_id, admin_id=get_admin_id(request.user))
            pf_data = data['pf']
            pf_detail, _ = PFDetail.objects.get_or_create(employee=employee)
            pf_detail.pf_number = pf_data.get('pf_number', '')
            pf_detail.esic_number = pf_data.get('esic', '')
            pf_detail.save()
            return JsonResponse({'success': True})

        # 4. SALARY ADJUSTMENT (or Create New)
        if 'salary' in data:
            employee = get_object_or_404(Employee, pk=emp_id, admin_id=get_admin_id(request.user))

            # Persist Role + Level back to the Employee record so the SPIM Lite
            # profile and salary chain stay in sync (Task 1 + Task 2). The
            # Salary Edit modal previously kept these in localStorage only,
            # which is why Level disappeared in SPIM Lite even though the
            # admin had "set" it via the salary modal.
            posted_role  = (data.get('role')  or '').strip()
            posted_level = (data.get('level') or '').strip()
            emp_dirty_fields = []
            if posted_role and posted_role != (employee.designation or ''):
                employee.designation = posted_role
                emp_dirty_fields.append('designation')
            if posted_level and posted_level != (employee.level or ''):
                employee.level = posted_level
                emp_dirty_fields.append('level')
            # Salary Type persistence. Blank / unknown values fall back to
            # 'base_salary' so the existing payroll flow is preserved. Only
            # a real change writes to the DB so untouched rows stay clean.
            posted_stype = (data.get('salary_type') or '').strip().lower()
            if posted_stype not in ('base_salary', 'daily_basis'):
                posted_stype = ''
            if posted_stype and posted_stype != (getattr(employee, 'salary_type', 'base_salary') or 'base_salary'):
                employee.salary_type = posted_stype
                emp_dirty_fields.append('salary_type')
            try:
                posted_base = Decimal(str(data.get('salary') or 0))
            except Exception:
                posted_base = Decimal('0')
            if posted_base > 0 and posted_base != employee.base_salary:
                employee.base_salary = posted_base
                emp_dirty_fields.append('base_salary')
                # Manual-override detection (Case 3). Compare the typed value
                # against the current Salary Config for this Role + Level. If
                # it differs, mark the employee as overridden so neither future
                # Salary Config edits nor role/level changes overwrite it.
                # If it matches the config exactly, clear the override flag so
                # the row resumes following the centralized mapping.
                admin_id = get_admin_id(request.user)
                cfg_base = None
                if employee.job_role_id and (employee.level or '').strip():
                    s = SalaryStructure.objects.filter(
                        admin_id=admin_id, job_role_id=employee.job_role_id,
                        level=employee.level.strip(),
                    ).first()
                    if s:
                        cfg_base = Decimal(str(s.base_salary))
                if cfg_base is None and (employee.designation or '').strip() and (employee.level or '').strip():
                    s = SalaryStructure.objects.filter(
                        admin_id=admin_id,
                        job_role__admin_id=admin_id,
                        job_role__name__iexact=employee.designation.strip(),
                        level=employee.level.strip(),
                    ).first()
                    if s:
                        cfg_base = Decimal(str(s.base_salary))
                new_override = (cfg_base is None) or (posted_base != cfg_base)
                if new_override != employee.salary_is_custom_override:
                    employee.salary_is_custom_override = new_override
                    emp_dirty_fields.append('salary_is_custom_override')
            if emp_dirty_fields:
                employee.save(update_fields=emp_dirty_fields)

            # Use provided month/year or default to current
            month_name = data.get('month', '')
            year = data.get('year', '')
            
            if month_name and year:
                m_idx = timezone_month_map(month_name)
                target_date = datetime.date(int(year), m_idx, 1)
            else:
                target_date = timezone.now().date().replace(day=1)

            salary_record, created = SalaryUpdate.objects.get_or_create(
                employee=employee,
                month=target_date,
                defaults={'created_by': request.user}
            )
            
            def _parse_decimal(val):
                if val is None or str(val).strip() == '':
                    return Decimal('0')
                try:
                    return Decimal(str(val))
                except Exception:
                    return Decimal('0')

            basic          = _parse_decimal(data.get('salary', 0))
            ot             = _parse_decimal(data.get('ot', 0))
            advance        = _parse_decimal(data.get('advance', 0))
            deduction      = _parse_decimal(data.get('deduction', 0))
            pf_amount      = _parse_decimal(data.get('pf_amount', 0))
            food_allowance = _parse_decimal(data.get('food_allowance', 0))
            food_usage     = _parse_decimal(data.get('food_usage', 0))

            # food_adjustment: positive → added to pay, negative → deducted
            food_adjustment = food_allowance - food_usage

            salary_record.basic_salary         = basic
            salary_record.ot_allowance         = ot
            salary_record.advance_pay          = advance
            salary_record.total_deduction      = deduction
            salary_record.pf_employee_snapshot = pf_amount
            salary_record.food_allowance       = food_allowance
            salary_record.food_usage           = food_usage
            # Only overwrite extra_allowance when the payload explicitly
            # sends it. The salary_manager modal doesn't expose this field
            # today, so the DB value is preserved for existing rows and the
            # model default (0) applies for newly created ones.
            if 'extra_allowance' in data and data.get('extra_allowance') not in (None, ''):
                salary_record.extra_allowance = _parse_decimal(data.get('extra_allowance'))

            # Compute the payable base server-side from AttendanceRecord rows.
            # The client-sent attendance_earnings is discarded — the server
            # derives the value directly from the database to prevent payload
            # manipulation.
            payable_base = _compute_attendance_earnings(employee, target_date, basic)
            salary_record.net_pay = payable_base + ot - advance - deduction + food_adjustment
            salary_record.save()
            
            return JsonResponse({'success': True})

        # 5. ADD NEW EMPLOYEE
        if 'name' in data:
            employee = Employee.objects.create(
                name=data['name'],
                designation=data.get('role', ''),
                location=data.get('location', ''),
                site=data.get('site', ''),
                base_salary=data.get('salary', 0),
                admin_id=get_admin_id(request.user),
                created_by=request.user
            )
            # Handle nested bank/pf if provided
            if 'bank' in data:
                b = data['bank']
                BankDetail.objects.create(
                    employee=employee,
                    bank_name=b.get('bank_name', ''),
                    account_holder=b.get('holder', ''),
                    account_number=b.get('account', ''),
                    ifsc_code=b.get('ifsc', '')
                )
            if 'pf' in data:
                p = data['pf']
                PFDetail.objects.create(
                    employee=employee,
                    pf_number=p.get('pf_number', ''),
                    esic_number=p.get('esic', '')
                )
            return JsonResponse({'success': True, 'id': employee.id})

        return JsonResponse({'success': False, 'error': 'Invalid request data'})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def add_salary(request):
    if request.method == 'POST':
        employee_id = request.POST.get('employee_id')
        employee = get_object_or_404(Employee, pk=employee_id, admin_id=get_admin_id(request.user))
        form = SalaryForm(request.POST)
        if form.is_valid():
            salary = form.save(commit=False)
            salary.employee = employee
            salary.created_by = request.user
            salary.save()
            messages.success(request, f"Salary record added for {employee.name}.")
            return redirect('employees:salary_management', pk=employee.pk)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.capitalize()}: {error}")
            return redirect('employees:salary_management', pk=employee.pk)
    return redirect('employees:list')


@login_required
def generate_payslip(request, pk):
    from dashboard.models import CompanySettings
    admin_id = get_admin_id(request.user)
    salary   = get_object_or_404(SalaryUpdate, pk=pk, employee__admin_id=admin_id)
    return render(request, 'employees/payslip.html', {
        'salary':   salary,
        'employee': salary.employee,
        'company':  CompanySettings.get_settings(admin_id),
    })


@login_required
@require_POST
def generate_salary_expenses(request):
    """
    Generate one expense Transaction per employee salary record for the given
    month/year.  Duplicate generation is blocked via a stable reference key
    stored on the Transaction (SAL-{employee_pk}-{YYYY-MM}).  Employee-level
    payroll details are NOT written to any expense field shown in the UI.
    """
    from finance.models import Transaction, Category
    from django.db import IntegrityError
    from django.utils.text import slugify as _slugify
    from attendance.models import AttendanceRecord
    from attendance.site_utils import resolve_working_site, get_or_create_office_site, OFFICE_SITE_NAME
    import logging as _sal_logging
    _sal_log = _sal_logging.getLogger(__name__)

    try:
        data = json.loads(request.body)
        month_name = data.get('month', '').strip()
        year_str   = data.get('year', '').strip()

        if not month_name or not year_str:
            return JsonResponse({'success': False, 'error': 'Month and year are required.'})

        m_idx       = timezone_month_map(month_name)
        year        = int(year_str)
        target_date = datetime.date(year, m_idx, 1)
        admin_id    = get_admin_id(request.user)

        # Optional: income source and account for balance combo linking.
        # When provided, generated salary expenses will deduct from the matching
        # income source/account balance.  When omitted, expenses are recorded
        # without combo linkage (no balance deduction), which is the legacy
        # behaviour.
        income_source_val = data.get('income_source', '').strip()
        account_val       = data.get('account', '').strip()

        # Resolve all employees belonging to this tenant
        tenant_employees = Employee.objects.filter(admin_id=admin_id)

        salary_records = SalaryUpdate.objects.filter(
            employee__in=tenant_employees,
            month__year=target_date.year,
            month__month=target_date.month,
        ).select_related('employee')

        if not salary_records.exists():
            return JsonResponse({
                'success': False,
                'error': f'No salary records found for {month_name} {year}. Process salaries first.',
            })

        # Get or create the "Salary" expense category for this tenant
        salary_category, _ = Category.objects.get_or_create(
            admin_id=admin_id,
            name='Salary',
            type='expense',
            defaults={'created_by': request.user, 'modified_by': request.user},
        )

        # Expense date = last day of the payroll month
        last_day     = calendar.monthrange(year, m_idx)[1]
        expense_date = datetime.date(year, m_idx, last_day)
        description  = f"Salary payout for {month_name} {year}"

        created_count = 0
        skipped_count = 0

        # Attendance window for the payroll month → drives per-day
        # working-site attribution. One SalaryUpdate row → possibly many
        # Transaction rows (one per distinct working site that month).
        month_start = target_date
        month_end   = datetime.date(year, m_idx, last_day)

        for record in salary_records:
            emp        = record.employee
            month_slug = target_date.strftime('%Y-%m')
            ref_prefix = f"SAL-{emp.pk}-{month_slug}"

            # Idempotent: drop any stale rows from a previous click (old
            # single-row shape SAL-{emp}-{YYYY-MM}, and any per-site rows
            # SAL-{emp}-{YYYY-MM}-{slug}). Do NOT touch [AUTO-SAL:...] —
            # the per-day live sync owns those.
            Transaction.objects.filter(
                admin_id=admin_id,
                type='expense',
                reference__startswith=ref_prefix,
            ).delete()

            # Group attendance days by working site.
            attendances = list(
                AttendanceRecord.objects
                .filter(admin_id=admin_id, employee=emp,
                        date__gte=month_start, date__lte=month_end)
                .select_related('site_ref')
            )
            site_days = {}
            for att in attendances:
                site_name = resolve_working_site(att)
                site_days[site_name] = site_days.get(site_name, 0) + 1

            net_pay = Decimal(record.net_pay or 0)

            if not site_days:
                # No attendance rows for the month — book the full amount
                # to OFFICE so the number still shows up on a card.
                _sal_log.info(
                    "generate_salary_expenses: no attendance for emp=%s "
                    "month=%s; booking full amount to OFFICE.",
                    emp.pk, month_slug,
                )
                site_days = {OFFICE_SITE_NAME: 1}

            if OFFICE_SITE_NAME in site_days:
                try:
                    get_or_create_office_site(admin_id, request.user)
                except Exception:
                    _sal_log.exception(
                        "generate_salary_expenses: failed to ensure "
                        "OFFICE site for tenant=%s", admin_id,
                    )

            total_days = sum(site_days.values())
            # Deterministic ordering so the "last row absorbs rounding"
            # rule is stable across re-clicks.
            ordered_sites = sorted(site_days.items(), key=lambda kv: kv[0])

            splits = []
            allocated = Decimal('0.00')
            for idx, (site_name, days) in enumerate(ordered_sites):
                if idx == len(ordered_sites) - 1:
                    # Last row absorbs sub-paisa rounding drift so the
                    # sum matches net_pay exactly.
                    share = (net_pay - allocated).quantize(Decimal('0.01'))
                else:
                    share = (net_pay * Decimal(days) / Decimal(total_days)) \
                        .quantize(Decimal('0.01'))
                    allocated += share
                splits.append((site_name, days, share))

            for site_name, days, share in splits:
                if share <= 0:
                    continue
                site_slug = _slugify(site_name) or 'nosite'
                ref_key = f"{ref_prefix}-{site_slug}"
                per_site_desc = (
                    f"{description} — {site_name} ({days}/{total_days} days)"
                )
                try:
                    Transaction.objects.create(
                        user          = request.user,
                        type          = 'expense',
                        category      = salary_category,
                        amount        = share,
                        description   = per_site_desc,
                        date          = expense_date,
                        purpose       = 'Employee Salary',
                        location_site = site_name,
                        reference     = ref_key,
                        income_source = income_source_val,
                        payment_mode  = account_val,
                        admin_id      = admin_id,
                        created_by    = request.user,
                    )
                    created_count += 1
                except IntegrityError:
                    # Partial UniqueConstraint on (admin_id, reference,
                    # type) for SAL-% caught a concurrent write — the
                    # prior delete + this create raced with a second
                    # click. Treat as skip.
                    _sal_log.warning(
                        "IntegrityError on salary expense insert; "
                        "treating as duplicate skip. ref=%s admin=%s",
                        ref_key, admin_id,
                    )
                    skipped_count += 1

        msg = f"{created_count} expense(s) generated for {month_name} {year}."
        if skipped_count:
            msg += f" {skipped_count} skipped (already exist)."

        return JsonResponse({'success': True, 'created': created_count, 'skipped': skipped_count, 'message': msg})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
