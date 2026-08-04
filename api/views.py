"""
SPIM Suite mobile API.

Public routes (no auth):
    POST /api/mobile/login/
    POST /api/mobile/reset-password/   (admin-side reset; uses Django session auth)

Authenticated routes (Bearer <token>):
    POST /api/mobile/change-password/
    GET  /api/mobile/profile/
    GET  /api/mobile/attendance/
    GET  /api/mobile/payslips/
    GET  /api/mobile/worklogs/

All authenticated routes are filtered to the employee that owns the token.
There is no way to query another employee's data through these endpoints.
"""
import json
import logging
import traceback
from functools import wraps
from datetime import date, datetime, timedelta

from accounts.date_utils import today_ist, validate_not_future
from django.core.exceptions import ValidationError as _DjValidationError

from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods, require_POST
from django.contrib.auth.decorators import login_required
from django.contrib.auth.hashers import check_password, make_password
from django.db.models import Sum, Count

from employees.models import Employee, SalaryUpdate, BankDetail
from attendance.models import AttendanceRecord
from attendance.utils import ensure_sunday_holidays, display_status
from projects.models import WorkLog, MachineLocation
from .models import MobileAuthToken
from .version_check import require_app_version

# Dedicated logger for the SPIM Lite mobile API. Surfaces silent save
# failures (admin_id mismatch, DB errors, lock conflicts) in Railway logs
# so the intermittent attendance-sync bug is debuggable. Configure once at
# module load — Django's root logger settings forward records to console.
_log = logging.getLogger('spim.api.mobile')


# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------

def _json_body(request):
    try:
        return json.loads(request.body or b'{}')
    except (ValueError, TypeError):
        return {}


def _extract_token(request):
    """
    Pull a bearer token from (in order):
      1. Authorization: Bearer <token> header
      2. X-Auth-Token header
      3. ?token=<token> query string  — only used for browser-opened links
         such as payslip downloads where custom headers cannot be set.
    """
    auth = request.headers.get('Authorization', '')
    if auth.lower().startswith('bearer '):
        return auth[7:].strip()
    header_tok = (request.headers.get('X-Auth-Token') or '').strip()
    if header_tok:
        return header_tok
    return (request.GET.get('token') or '').strip()


def mobile_auth_required(view_func):
    """
    Decorator: resolve the bearer token, attach `request.employee`, or 401.
    Inactive employee accounts (mobile_account_active=False) are rejected.
    """
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        key = _extract_token(request)
        if not key:
            return JsonResponse({'success': False, 'error': 'Authentication required.'}, status=401)
        try:
            tok = MobileAuthToken.objects.select_related('employee').get(key=key)
        except MobileAuthToken.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Invalid or expired token.'}, status=401)
        emp = tok.employee
        if not emp.mobile_account_active:
            return JsonResponse({'success': False, 'error': 'Account disabled.'}, status=403)
        # Touch last_used (auto_now updates on save)
        tok.save(update_fields=['last_used'])
        request.employee = emp
        request.mobile_token = tok
        return view_func(request, *args, **kwargs)
    return _wrapped


# ---------------------------------------------------------------------------
# Mobile auth endpoints
# ---------------------------------------------------------------------------

@csrf_exempt
@require_app_version
@require_POST
def mobile_login(request):
    """
    POST { "login_id": "SPIM001", "password": "...", "device_info": "optional" }
    -> { success, token, employee: {id, employee_id, name, ...}, password_reset_required }
    """
    data       = _json_body(request)
    login_id   = (data.get('login_id') or '').strip()
    password   = data.get('password') or ''
    device     = (data.get('device_info') or '')[:255]

    if not login_id or not password:
        return JsonResponse({'success': False, 'error': 'login_id and password required.'}, status=400)

    # Look up by employee_login_id first, fall back to employee_id for back-compat.
    emp = Employee.objects.filter(employee_login_id=login_id).first() \
        or Employee.objects.filter(employee_id=login_id).first()

    if not emp or not emp.mobile_account_active:
        return JsonResponse({'success': False, 'error': 'Invalid credentials.'}, status=401)

    # Verify against hash; if hash is empty (legacy row), fall back to plaintext
    # compare against mobile_app_password, then upgrade to a hash.
    valid = False
    if emp.mobile_password_hash:
        valid = check_password(password, emp.mobile_password_hash)
    elif emp.mobile_app_password and password == emp.mobile_app_password:
        valid = True
        emp.mobile_password_hash = make_password(password)
        emp.save(update_fields=['mobile_password_hash'])

    if not valid:
        return JsonResponse({'success': False, 'error': 'Invalid credentials.'}, status=401)

    token = MobileAuthToken.objects.create(employee=emp, device_info=device)

    return JsonResponse({
        'success': True,
        'token': token.key,
        'password_reset_required': emp.mobile_password_reset_required,
        'employee': {
            'id':            emp.id,
            'employee_id':   emp.employee_id,
            'login_id':      emp.employee_login_id or emp.employee_id,
            'name':          emp.name,
            'designation':   emp.designation,
            'department':    emp.department,
            'location':      emp.location,
            'site':          emp.site,
            # Surfaced at login so SPIM Lite clients that cache the login
            # response (and don't refetch /profile/) still see Level / Mobile
            # / Branch / Base Salary in the profile screen.
            'level':         emp.level,
            'mobile':        emp.mobile,
            'branch':        emp.branch,
            'base_salary':   str(emp.base_salary),
        },
    })


@csrf_exempt
@require_app_version
@require_POST
@mobile_auth_required
def mobile_change_password(request):
    """
    POST { "current_password": "...", "new_password": "...", "confirm_password": "..." }
    Authenticated employee changes their own password.
    """
    data    = _json_body(request)
    current = data.get('current_password') or ''
    new_pw  = data.get('new_password') or ''
    confirm = data.get('confirm_password') or ''

    if not current or not new_pw or not confirm:
        return JsonResponse({'success': False, 'error': 'All password fields are required.'}, status=400)
    if new_pw != confirm:
        return JsonResponse({'success': False, 'error': 'New password and confirmation do not match.'}, status=400)
    if len(new_pw) < 8:
        return JsonResponse({'success': False, 'error': 'Password must be at least 8 characters.'}, status=400)

    emp = request.employee
    # Verify current password (hash, then legacy plaintext fallback)
    if emp.mobile_password_hash:
        ok = check_password(current, emp.mobile_password_hash)
    else:
        ok = bool(emp.mobile_app_password) and current == emp.mobile_app_password
    if not ok:
        return JsonResponse({'success': False, 'error': 'Current password is incorrect.'}, status=400)

    emp.mobile_password_hash           = make_password(new_pw)
    emp.mobile_app_password            = ''  # never retain plaintext after a self-change
    emp.mobile_password_reset_required = False
    emp.save(update_fields=['mobile_password_hash', 'mobile_app_password', 'mobile_password_reset_required'])

    # Invalidate all other tokens for this employee; keep the current one.
    MobileAuthToken.objects.filter(employee=emp).exclude(pk=request.mobile_token.pk).delete()

    return JsonResponse({'success': True})


@csrf_exempt
@login_required
@require_POST
def mobile_reset_password(request):
    """
    Admin-side reset. Uses Django session auth (admin logged into SPIM Suite).
    POST { "employee_id": <pk>, "new_password": "optional — auto-generated if blank" }
    -> { success, new_password }   (plaintext returned ONCE for the admin)
    """
    # Defer import to avoid circulars / unrelated coupling
    from accounts.views import get_admin_id
    from employees.views import _generate_app_password

    data    = _json_body(request)
    emp_pk  = data.get('employee_id')
    new_pw  = (data.get('new_password') or '').strip()

    if not emp_pk:
        return JsonResponse({'success': False, 'error': 'employee_id required.'}, status=400)
    try:
        emp = Employee.objects.get(pk=emp_pk, admin_id=get_admin_id(request.user))
    except Employee.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Employee not found.'}, status=404)

    if not new_pw:
        new_pw = _generate_app_password()
    elif len(new_pw) < 8:
        return JsonResponse({'success': False, 'error': 'Password must be at least 8 characters.'}, status=400)

    emp.mobile_app_password            = new_pw  # shown once to admin
    emp.mobile_password_hash           = make_password(new_pw)
    emp.mobile_password_reset_required = True    # employee should change on next login
    emp.mobile_account_active          = True
    emp.save(update_fields=[
        'mobile_app_password', 'mobile_password_hash',
        'mobile_password_reset_required', 'mobile_account_active',
    ])
    # Force re-login on all devices
    MobileAuthToken.objects.filter(employee=emp).delete()

    return JsonResponse({'success': True, 'new_password': new_pw})


# ---------------------------------------------------------------------------
# Mobile data endpoints — STRICTLY scoped to request.employee
# ---------------------------------------------------------------------------

@require_app_version
@mobile_auth_required
@require_http_methods(['GET'])
def mobile_profile(request):
    """
    Full employee snapshot. Strictly scoped to request.employee — there is no
    employee-id query parameter and no listing variant.
    """
    emp  = request.employee
    bank = getattr(emp, 'bank_details', None)
    pf   = getattr(emp, 'pf_details', None)
    job  = emp.job_role

    # Latest payslip (most recent SalaryUpdate row)
    latest = SalaryUpdate.objects.filter(employee=emp).order_by('-month').first()
    latest_salary = None
    if latest:
        latest_salary = {
            'month':           latest.month.strftime('%Y-%m'),
            'basic_salary':    str(latest.basic_salary),
            'extra_allowance': str(latest.extra_allowance),
            'ot_allowance':    str(latest.ot_allowance),
            'advance_pay':     str(latest.advance_pay),
            'total_deduction': str(latest.total_deduction),
            'food_allowance':  str(latest.food_allowance),
            'food_usage':      str(latest.food_usage),
            'pf_employee':     str(latest.pf_employee_snapshot),
            'pf_employer':     str(latest.pf_employer_snapshot),
            'net_pay':         str(latest.net_pay),
        }

    # Company block — sourced from dashboard.CompanySettings for the
    # employee's tenant so SPIM Lite renders the same Company Name / Logo /
    # Address / Phone / Email as the Suite. Deferred import keeps this
    # module loadable when dashboard is not installed in some environments.
    company_block = None
    try:
        from dashboard.models import CompanySettings  # deferred to avoid cycles
        cs = CompanySettings.get_settings(emp.admin_id)
        if cs:
            logo_url = ''
            try:
                if cs.logo and hasattr(cs.logo, 'url'):
                    logo_url = request.build_absolute_uri(cs.logo.url)
            except Exception:
                logo_url = ''
            company_block = {
                'name':              cs.name or '',
                'logo_url':          logo_url,
                'address':           cs.address or '',
                'contact_number':    cs.contact_number or '',
                'email':             cs.email or '',
                'gst_number':        cs.gst_number or '',
                'managing_director': cs.managing_director or '',
            }
    except Exception:
        company_block = None

    # Attendance summary for the ACTIVE PAYROLL CYCLE (26th → 25th window
    # whose end date falls in the current calendar month), not the raw
    # calendar month. Employees expect the mobile summary to match their
    # payslip period — a raw month view undercounted days between the 1st
    # and 25th (missing the cycle's opening 26th-31st of the prior month)
    # and inflated days between the 26th and end-of-month (which belong to
    # the NEXT payroll cycle).
    # Live view: current cycle. Payslip endpoints intentionally use get_last_completed_cycle().
    from accounts.cycle_utils import get_salary_cycle
    from accounts.date_utils import today_ist
    _cycle = get_salary_cycle(today_ist())
    cycle_qs = AttendanceRecord.objects.filter(
        employee=emp,
        date__gte=_cycle['start'],
        date__lte=_cycle['end'],
    )
    # `absent` status has been retired and merged into `no_week_off`
    # (migration 0009). The `absent` key here is DEPRECATED — kept only
    # so older APKs continue to render this tile after the migration.
    # Remove once the SPIM Lite rollout is complete.
    _nwo_count = cycle_qs.filter(status='no_week_off').count()
    attendance_summary = {
        # `month` kept as the payroll month_key ('YYYY-MM') so the response
        # shape stays byte-identical for existing APK builds; only the
        # aggregated counts change from calendar-month to cycle-window.
        'month':       _cycle['month_key'],
        'present':     cycle_qs.filter(status='present').count(),
        'half_day':    cycle_qs.filter(status='half_day').count(),
        'no_week_off': _nwo_count,
        'absent':      _nwo_count,  # DEPRECATED: dual-emit for legacy APKs
        'leave':       cycle_qs.filter(status='leave').count(),
    }

    return JsonResponse({
        'success': True,
        'profile': {
            'id':                 emp.id,
            'employee_id':        emp.employee_id,
            'login_id':           emp.employee_login_id or emp.employee_id,
            'name':               emp.name,
            'designation':        emp.designation,
            'department':         emp.department,
            'location':           emp.location,
            'site':               emp.site,
            # New fields (Task 2 + Task 5)
            'level':              emp.level,
            'mobile':             emp.mobile,
            'branch':             emp.branch,
            'base_salary':        str(emp.base_salary),
            'salary_is_custom_override': bool(emp.salary_is_custom_override),
            'fixed_allowance':    str(emp.fixed_allowance),
            'joining_date':       str(emp.joining_date) if emp.joining_date else None,
            'status':             emp.status,
            'job_role': {
                'name':         job.name,
                'salary_type':  job.salary_type,
                'base_salary':  str(job.base_salary),
            } if job else None,
            'bank': {
                'bank_name':      bank.bank_name      if bank else '',
                'account_holder': bank.account_holder if bank else '',
                'account_number': bank.account_number if bank else '',
                'ifsc_code':      bank.ifsc_code      if bank else '',
                'branch':         bank.branch         if bank else '',
            } if bank else None,
            'pf': {
                'pf_number':             pf.pf_number             if pf else '',
                'uan_number':            pf.uan_number            if pf else '',
                'esic_number':           pf.esic_number           if pf else '',
                'employee_contribution': str(pf.employee_contribution) if pf else '0',
                'employer_contribution': str(pf.employer_contribution) if pf else '0',
            } if pf else None,
            'latest_salary':      latest_salary,
            'attendance_summary': attendance_summary,
            # Company details for the SPIM Lite dashboard. Frontend should
            # safely skip individual fields that are empty strings.
            'company':            company_block,
        },
    })


@csrf_exempt
@require_app_version
@mobile_auth_required
@require_http_methods(['GET', 'POST'])
def mobile_attendance(request):
    """
    GET  -> own attendance records (optional ?month=YYYY-MM filter).
    POST -> create own attendance for a single date (locked once created).
            Body: { "date": "YYYY-MM-DD" (optional, defaults today), "status": "present|no_week_off|half_day|leave" }
            (Legacy 'absent' from older APKs is coerced to 'no_week_off' — see coercion block below.)

    Locking rules (Suite ⇄ Lite sync hardening):
      * If a record already exists with source='admin', the employee cannot
        overwrite it. Response: HTTP 409 with the locked record.
      * If a record already exists with source='employee' and the status
        matches, the response is idempotent success (safe to retry).
      * If a record already exists with source='employee' but the status
        differs, the employee cannot self-correct — HTTP 409. Admin can
        still override via the Suite admin save endpoint, which writes
        source='admin' and re-locks the row.
      * No existing record → create a new employee-sourced row.

    The presence of a record (any source) is the lock. Every record in the
    response carries `locked: true` so SPIM Lite can render the lock state
    without inferring it from `source`.
    """
    emp = request.employee

    if request.method == 'POST':
        # ─── Parse + validate request body ──────────────────────────────
        data         = _json_body(request)
        # Anchor default "today" to IST — server runs UTC (Railway) but the
        # workforce is in India, so `date.today()` would flip a day 5.5 h
        # early and reject valid same-day marks between 00:00–05:30 IST.
        date_str     = (data.get('date') or '').strip() or today_ist().isoformat()
        status_val   = (data.get('status') or 'present').strip().lower()
        # Site / Working Site fields (new — Mod 2/3). Both optional. When
        # the SPIM Lite client doesn't send them (older builds), the
        # employee's home-site is used as a sensible default so the Suite
        # registry still gets a populated Site column.
        site_val         = (data.get('site') or '').strip()
        working_site_val = (data.get('working_site') or '').strip()
        if not site_val:
            site_val = (emp.site or '').strip()

        # Inbound compat shim: retired 'absent' → 'no_week_off' (migration
        # 0009). Older APKs in the field still POST 'absent'; coerce here
        # so the validator below doesn't 400 them. Logged so we can track
        # APK-rollout adoption and remove the shim later.
        if status_val == 'absent':
            _log.warning(
                "mobile_attendance coerced legacy status='absent' to "
                "'no_week_off' emp=%s date=%s",
                emp.pk, date_str,
            )
            status_val = 'no_week_off'

        valid_status = {s for s, _ in AttendanceRecord.STATUS_CHOICES}
        if status_val not in valid_status:
            _log.warning(
                'mobile_attendance POST rejected: invalid status emp=%s status=%s',
                emp.pk, status_val,
            )
            return JsonResponse(
                {'success': False, 'error': f'Invalid status. Allowed: {sorted(valid_status)}'},
                status=400,
            )
        try:
            d = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            _log.warning(
                'mobile_attendance POST rejected: bad date emp=%s date=%r',
                emp.pk, date_str,
            )
            return JsonResponse({'success': False, 'error': 'date must be YYYY-MM-DD.'}, status=400)
        # Don't allow marking attendance for future dates (IST-anchored so
        # the check matches web-app validation and the SPIM Lite user's own
        # clock). Message intentionally identical to accounts.date_utils so
        # web + mobile show the same string.
        try:
            validate_not_future(d, "Attendance date")
        except _DjValidationError as e:
            return JsonResponse({'success': False, 'error': '; '.join(e.messages)}, status=400)

        # ─── Resolve admin_id robustly ──────────────────────────────────
        # Bug 1 root cause: when emp.admin_id was 'PENDING' / blank /
        # mismatched, the AttendanceRecord landed in a tenancy bucket the
        # Suite's `filter(admin_id=admin_id)` queries never touched, so
        # the mark "saved successfully" from the mobile side but vanished
        # in the Suite registry. Fall back to the employee's creator's
        # admin_id when the direct field is unusable.
        effective_admin_id = (emp.admin_id or '').strip()
        if not effective_admin_id or effective_admin_id == 'PENDING':
            creator = getattr(emp, 'created_by', None)
            if creator is not None and getattr(creator, 'admin_id', None):
                effective_admin_id = creator.admin_id
                _log.info(
                    'mobile_attendance: emp=%s admin_id was %r, resolved via creator → %r',
                    emp.pk, emp.admin_id, effective_admin_id,
                )
        if not effective_admin_id:
            _log.error(
                'mobile_attendance: emp=%s has no usable admin_id; mark will be invisible to Suite',
                emp.pk,
            )
            return JsonResponse({
                'success': False,
                'error': 'Employee account is missing tenant scope. Please ask your admin to refresh your profile.',
            }, status=500)

        # ─── Lock enforcement ───────────────────────────────────────────
        # Never destructively overwrite an existing row from the mobile
        # path. This stops APK background-sync POSTs from silently
        # overwriting admin marks. EXCEPTION: a Sunday-auto-Holiday row
        # (source='admin' but stamped purely by the lazy backfill) should
        # NOT prevent the employee from claiming Present for that Sunday —
        # the previous behaviour was the root cause of the intermittent
        # "marks vanish" symptom on Sundays touched by ensure_sunday_holidays.
        existing = AttendanceRecord.objects.filter(employee=emp, date=d).first()
        if existing is not None:
            payload_record = {
                'id':           existing.id,
                'date':         str(existing.date),
                'status':       existing.status,
                'source':       existing.source,
                'site':         getattr(existing, 'site', '') or '',
                'working_site': getattr(existing, 'working_site', '') or '',
                'locked':       True,
            }
            # Detect the auto-Sunday-Holiday case: source='admin' + status='holiday'
            # + the date IS a Sunday + the row has no human-set audit fields.
            # The lazy ensure_sunday_holidays helper creates rows with
            # created_by=NULL and status='holiday'. Those should be
            # overwritable by the employee mark.
            is_auto_sunday = (
                existing.source == 'admin'
                and existing.status == 'holiday'
                and existing.created_by_id is None
                and d.weekday() == 6  # Sunday
            )
            if existing.source == 'admin' and not is_auto_sunday:
                return JsonResponse({
                    'success': False,
                    'error':   'Attendance for this date has been set by the admin and is locked.',
                    'locked':  True,
                    'record':  payload_record,
                }, status=409)
            if existing.source == 'employee':
                if existing.status == status_val:
                    # Idempotent retry — same value the employee already marked.
                    return JsonResponse({
                        'success': True,
                        'created': False,
                        'locked':  True,
                        'record':  payload_record,
                    })
                return JsonResponse({
                    'success': False,
                    'error':   'You have already marked attendance for this date.',
                    'locked':  True,
                    'record':  payload_record,
                }, status=409)
            # is_auto_sunday → fall through to upsert below

        # ─── Save (or upsert auto-Sunday-Holiday) ────────────────────────
        # Wrapped in try/except so transient DB errors surface in Railway
        # logs and return a meaningful error to the mobile client instead
        # of the generic 500 the previous code emitted.
        try:
            rec, created = AttendanceRecord.objects.update_or_create(
                employee=emp,
                date=d,
                defaults={
                    'status':       status_val,
                    'source':       'employee',
                    'admin_id':     effective_admin_id,
                    'site':         site_val,
                    'working_site': working_site_val,
                },
            )
        except Exception as exc:
            _log.exception(
                'mobile_attendance save failed: emp=%s date=%s status=%s admin_id=%s err=%s',
                emp.pk, d, status_val, effective_admin_id, exc,
            )
            return JsonResponse({
                'success': False,
                'error':   f'Could not save attendance: {exc.__class__.__name__}: {exc}',
            }, status=500)

        _log.info(
            'mobile_attendance saved: emp=%s date=%s status=%s site=%r ws=%r admin_id=%s created=%s',
            emp.pk, d, status_val, site_val, working_site_val, effective_admin_id, created,
        )
        return JsonResponse({
            'success': True,
            'created': created,
            'locked':  True,
            'record': {
                'id':           rec.id,
                'date':         str(rec.date),
                'status':       rec.status,
                'source':       rec.source,
                'site':         rec.site or '',
                'working_site': rec.working_site or '',
                'locked':       True,
            },
        })

    # GET
    qs    = AttendanceRecord.objects.filter(employee=emp).order_by('-date')
    month = (request.GET.get('month') or '').strip()
    if month:
        try:
            y, m = month.split('-')
            year_i, month_i = int(y), int(m)
            qs = qs.filter(date__year=year_i, date__month=month_i)
            # Lazy Sunday auto-Holiday backfill scoped to this employee for
            # the requested month — guarantees that when the Lite app
            # fetches a month's attendance, every Sunday already exists
            # as a 'holiday' row (created here if missing, untouched if
            # the admin has already overridden it).
            try:
                import calendar as _cal
                last_day = _cal.monthrange(year_i, month_i)[1]
                ensure_sunday_holidays(
                    emp.admin_id,
                    date(year_i, month_i, 1),
                    date(year_i, month_i, last_day),
                    employees=[emp],
                )
                # Re-query so the freshly-created rows are included in the
                # response. Cheap — same indexes as the original query.
                qs = AttendanceRecord.objects.filter(
                    employee=emp,
                    date__year=year_i,
                    date__month=month_i,
                ).order_by('-date')
            except (ValueError, TypeError):
                pass
        except (ValueError, IndexError):
            pass
    records = [{
        'id':     r.id,
        'date':   str(r.date),
        'status': r.status,
        'source': r.source,
        # Site / Working Site — defensive getattr keeps the GET working
        # even before migration 0004 has been applied to the Lite-facing DB.
        'site':         getattr(r, 'site',         '') or '',
        'working_site': getattr(r, 'working_site', '') or '',
        # Presence of a record == locked. Explicit flag so SPIM Lite can
        # render the lock state without inferring it from `source`.
        'locked': True,
    } for r in qs[:200]]
    return JsonResponse({'success': True, 'attendance': records})


@require_app_version
@mobile_auth_required
@require_http_methods(['GET'])
def mobile_payslips(request):
    """
    List payslips for the authenticated employee. Every row is returned
    (admins may need visibility of pending months too), but only rows with
    is_generated=True are downloadable — see mobile_payslip_download.

    Suite is the single source of truth for *which* payslip SPIM Lite must
    display. The 26→25 attendance cycle is anchored on today and the
    corresponding SalaryUpdate is identified by (year, month) of cycle-end
    — the same convention `mobile_salary` uses. Two per-row flags are
    emitted so the mobile client never compares months, sorts, or
    otherwise derives payroll state:

      * is_current_cycle    — the row whose month matches today's cycle-end
      * is_latest_generated — the most-recent SalaryUpdate with is_payslip_generated=True

    Plus a top-level `current_payslip_id` that picks:
      1. the current-cycle payslip if it exists AND has been generated, OR
      2. the latest generated previous payslip.

    SPIM Lite renders exactly that id — no client-side selection logic.
    """
    emp = request.employee

    # Active payroll cycle — same rule mobile_salary uses so the payslips
    # tab and the salary tab agree on which SalaryUpdate row is "current".
    # STRICT last-completed cycle: a cycle whose end-date has not yet
    # passed is never surfaced to the employee — even if HR generated the
    # SalaryUpdate row early, it would represent an in-progress window.
    from accounts.cycle_utils import get_last_completed_cycle
    from accounts.date_utils import today_ist
    cycle_end = get_last_completed_cycle(today_ist())['end']

    rows = list(SalaryUpdate.objects.filter(employee=emp).order_by('-month')[:60])

    # Identify the current-cycle row (may or may not be generated yet).
    current_row = next(
        (s for s in rows
         if s.month.year == cycle_end.year and s.month.month == cycle_end.month),
        None,
    )

    # Latest generated payslip anywhere in the history slice above.
    latest_generated_row = next(
        (s for s in rows if s.is_payslip_generated),
        None,
    )

    # Which payslip should SPIM Lite render?
    if current_row and current_row.is_payslip_generated:
        current_payslip_id = current_row.id
    elif latest_generated_row:
        current_payslip_id = latest_generated_row.id
    else:
        current_payslip_id = None

    payslips = [{
        'id':                  s.id,
        'is_generated':        s.is_payslip_generated,
        'generated_at':        s.payslip_generated_at.isoformat() if s.payslip_generated_at else None,
        'month':               s.month.strftime('%Y-%m'),
        'basic_salary':        str(s.basic_salary),
        'ot_allowance':        str(s.ot_allowance),
        'advance_pay':         str(s.advance_pay),
        'total_deduction':     str(s.total_deduction),
        'food_allowance':      str(s.food_allowance),
        'food_usage':          str(s.food_usage),
        'net_pay':             str(s.net_pay),
        'is_current_cycle':    bool(current_row and s.id == current_row.id),
        'is_latest_generated': bool(latest_generated_row and s.id == latest_generated_row.id),
    } for s in rows]
    return JsonResponse({
        'success':            True,
        'payslips':           payslips,
        'current_payslip_id': current_payslip_id,
    })


@csrf_exempt
@require_app_version
@mobile_auth_required
@require_http_methods(['GET', 'POST'])
def mobile_worklogs(request):
    """
    GET  -> work logs where this employee is in the M2M `employees` set.
    POST -> employee submits an end-of-day machine log; reflects directly into
            SPIM Suite Projects → Machine Work Summary.

            Body: {
              machine_no: "G001",            # required, must already exist as a MachineLocation
              status:     "Painting",         # written to WorkLog.work_details
              remarks:    "...",
              tmp:        4,                  # Total Man Power
              date:       "2026-05-24"        # optional, defaults today
            }

            Single source of truth: writes the existing `projects.WorkLog`
            row (same one the admin reads in the Suite). Upsert keyed on
            (admin_id, date, location) — the same 3-tuple used by the admin
            `work_log_upsert` endpoint in projects/views.py. `site` is now
            merged data (last-write-wins on non-empty value), not part of
            identity, so two employees on the same machine/date with
            different stored sites still collapse into one WorkLog row.
            The employee is added to the M2M `employees` set, TMP grows
            via max(existing, posted) so multiple punches do not double
            count. created_by stays NULL (mobile-originated row).
    """
    emp = request.employee

    if request.method == 'POST':
        data        = _json_body(request)
        machine_no  = (data.get('machine_no') or '').strip()
        # Accept multiple field names for work status: SPIM Lite may send
        # 'status' (original spec), 'work_status', or 'work_details' (the
        # DB column name mirrored in GET responses). Fall through in priority
        # order so whichever the client sends is captured.
        status_val  = (data.get('status') or data.get('work_status') or data.get('work_details') or '').strip()
        remarks_val = (data.get('remarks')    or '').strip()
        tmp_val     = data.get('tmp')
        date_str    = (data.get('date')       or '').strip() or today_ist().isoformat()

        if not machine_no:
            return JsonResponse({'success': False, 'error': 'machine_no is required.'}, status=400)
        try:
            d = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'success': False, 'error': 'date must be YYYY-MM-DD.'}, status=400)
        try:
            validate_not_future(d, "Work log date")
        except _DjValidationError as e:
            return JsonResponse({'success': False, 'error': '; '.join(e.messages)}, status=400)

        machine = MachineLocation.objects.filter(admin_id=emp.admin_id, name=machine_no).first()
        if not machine:
            return JsonResponse(
                {'success': False, 'error': f'Machine "{machine_no}" not found for this tenant.'},
                status=400,
            )

        try:
            tmp_int = int(tmp_val) if tmp_val not in (None, '') else 0
            if tmp_int < 0:
                tmp_int = 0
        except (TypeError, ValueError):
            tmp_int = 0

        site_val = (emp.site or '').strip()

        # Identity is (admin_id, date, location) — matches projects.work_log_upsert.
        # site is merged into defaults / overwritten on collision when non-empty.
        wl, created = WorkLog.objects.get_or_create(
            admin_id=emp.admin_id,
            date=d,
            location=machine,
            defaults={
                'site':         site_val,
                'work_details': status_val,
                'remarks':      remarks_val,
                'tmp':          tmp_int,
            },
        )
        if not created:
            # Respect the admin lock — an admin-saved (locked) row is
            # read-only from every path. The mobile employee can still be
            # associated via the M2M (they *did* work that machine), but
            # site / work_details / remarks / tmp are preserved as-is.
            if getattr(wl, 'locked', False):
                wl.employees.add(emp)
                return JsonResponse({
                    'success': True,
                    'created': False,
                    'locked':  True,
                    'worklog': {
                        'id':           wl.id,
                        'date':         str(wl.date),
                        'location':     wl.location.name if wl.location_id else '',
                        'site':         wl.site,
                        'work_details': wl.work_details,
                        'remarks':      wl.remarks,
                        'tmp':          wl.tmp,
                    },
                })
            # Merge — non-empty values overwrite existing, TMP grows.
            dirty = []
            if site_val and site_val != (wl.site or ''):
                wl.site = site_val
                dirty.append('site')
            if status_val and status_val != (wl.work_details or ''):
                wl.work_details = status_val
                dirty.append('work_details')
            if remarks_val and remarks_val != (wl.remarks or ''):
                wl.remarks = remarks_val
                dirty.append('remarks')
            if tmp_int > (wl.tmp or 0):
                wl.tmp = tmp_int
                dirty.append('tmp')
            if dirty:
                dirty.append('updated_at')
                wl.save(update_fields=dirty)
        wl.employees.add(emp)

        # Best-effort: if the mobile POST introduced a new freeform site
        # string, upsert it into projects.Site (under the "Unassigned"
        # ProjectClient) so it shows up in the desktop tree on next load.
        # Failure here MUST NOT affect the mobile response contract.
        _mobile_upsert_site_registry(emp.admin_id, site_val)

        return JsonResponse({
            'success': True,
            'created': created,
            'worklog': {
                'id':           wl.id,
                'date':         str(wl.date),
                'location':     wl.location.name if wl.location_id else '',
                'site':         wl.site,
                'work_details': wl.work_details,
                'remarks':      wl.remarks,
                'tmp':          wl.tmp,
            },
        })


def _mobile_upsert_site_registry(admin_id, site_str):
    """
    Best-effort: ensure `site_str` exists as a projects.Site row for this
    tenant, parked under the "Unassigned" ProjectClient. Never raises.

    Isolated helper so the mobile write path's response shape never depends
    on the desktop-registry side effect succeeding.
    """
    site_str = (site_str or '').strip()
    if not site_str:
        return
    try:
        from projects.models import ProjectClient, Site
        client, _ = ProjectClient.objects.get_or_create(
            admin_id=admin_id, name='Unassigned',
            defaults={'is_active': True},
        )
        existing = Site.objects.filter(
            admin_id=admin_id, name__iexact=site_str,
        ).first()
        if not existing:
            Site.objects.create(
                admin_id=admin_id, name=site_str,
                client=client, is_active=True,
            )
    except Exception:
        return

    # GET
    qs   = WorkLog.objects.select_related('location').filter(employees=emp).order_by('-date')[:200]
    logs = [{
        'id':           w.id,
        'date':         str(w.date),
        'location':     w.location.name if w.location_id else '',
        'site':         w.site,
        'work_details': w.work_details,
        'remarks':      w.remarks,
        'tmp':          w.tmp,
    } for w in qs]
    return JsonResponse({'success': True, 'worklogs': logs})


@require_app_version
@mobile_auth_required
@require_http_methods(['GET'])
def mobile_machines(request):
    """
    Return every MachineLocation belonging to this employee's tenant (Task 1).

    Business rule: admin only registers Machine Numbers / Locations under the
    Suite's Projects module. They do NOT assign machines to specific
    employees. Each employee picks their machine themselves in SPIM Lite.
    """
    emp = request.employee
    qs  = MachineLocation.objects.filter(admin_id=emp.admin_id).order_by('name')
    machines = [{'id': m.id, 'machine_no': m.name} for m in qs]
    return JsonResponse({'success': True, 'machines': machines})


@require_app_version
@mobile_auth_required
@require_http_methods(['GET'])
def mobile_sites(request):
    """
    Return every site known to this employee's tenant (Mod 2/3).

    Source list is the union of:
      * `branches.LocationSite.name` for the tenant (admin-managed registry
        used by the Suite's Site / Working Site dropdowns).
      * Distinct `Employee.site` values currently configured for the tenant.
      * The employee's own `site` field — guarantees the dropdown always
        contains their default even when the admin hasn't catalogued it yet.

    SPIM Lite calls this once at attendance-screen load time to populate
    the Site dropdown that drives `site` in the mobile_attendance POST.
    """
    emp = request.employee
    sites = set()

    try:
        from branches.models import LocationSite
        for name in LocationSite.objects.filter(admin_id=emp.admin_id).values_list('name', flat=True):
            n = (name or '').strip()
            if n:
                sites.add(n)
    except Exception as exc:
        _log.warning('mobile_sites: LocationSite lookup failed for emp=%s err=%s', emp.pk, exc)

    try:
        for name in Employee.objects.filter(admin_id=emp.admin_id).values_list('site', flat=True):
            n = (name or '').strip()
            if n:
                sites.add(n)
    except Exception as exc:
        _log.warning('mobile_sites: Employee.site sweep failed for emp=%s err=%s', emp.pk, exc)

    own = (emp.site or '').strip()
    if own:
        sites.add(own)

    return JsonResponse({
        'success': True,
        'sites':   sorted(sites),
        'default': own,
    })


@require_http_methods(['GET'])
def mobile_payslip_download(request, pk):
    """
    Dual-mode endpoint driven by the request's `Accept` header:

      * `Accept: application/json` (SPIM Lite's `apiGet` — authenticated
        via `Authorization: Bearer <token>`) → returns
        `{"download_url": "<absolute>/api/mobile/payslips/<pk>/download/?token=<key>"}`.
        The APK hands that URL to `Linking.openURL(...)`, which opens the
        system browser. Browsers can't attach custom auth headers, so the
        token rides in the query string — see `_extract_token`'s
        `?token=` fallback.

      * Anything else (system browser hitting the URL above) → renders
        the payslip HTML from `templates/employees/payslip.html`. The
        user prints to PDF from there.

    Cross-employee access is blocked because the query joins on
    `request.employee`. Locked until admin clicks "Generate Payslip":
    ungenerated payslips return 403.

    Version gate note: `@require_app_version` is intentionally OMITTED
    here. External browser hits carry no `App-Version` header, and
    leaving the gate on would 426-block every payslip download.

    Auth note: `@mobile_auth_required` is applied inline (not as a
    decorator) so we can emit one greppable log line per attempt BEFORE
    auth runs — otherwise failed downloads (missing token, expired
    token) silently 401 with no forensics.
    """
    _log.info(
        "PAYSLIP_DOWNLOAD_ATTEMPT: url=%s auth_header=%s token_qp=%s referer=%s",
        request.get_full_path(),
        'present' if request.headers.get('Authorization') else 'MISSING',
        'present' if request.GET.get('token') else 'MISSING',
        request.headers.get('Referer', 'MISSING'),
    )

    # Inline the same checks @mobile_auth_required performs, so the log
    # above always fires. Behavior is byte-identical to the decorator.
    key = _extract_token(request)
    if not key:
        return JsonResponse({'success': False, 'error': 'Authentication required.'}, status=401)
    try:
        tok = MobileAuthToken.objects.select_related('employee').get(key=key)
    except MobileAuthToken.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Invalid or expired token.'}, status=401)
    emp = tok.employee
    if not emp.mobile_account_active:
        return JsonResponse({'success': False, 'error': 'Account disabled.'}, status=403)
    tok.save(update_fields=['last_used'])
    request.employee = emp
    request.mobile_token = tok

    salary = get_object_or_404(SalaryUpdate, pk=pk, employee=request.employee)
    if not salary.is_payslip_generated:
        return JsonResponse(
            {'success': False, 'error': 'Payslip not yet generated by admin.'},
            status=403,
        )

    # JSON resolver: SPIM Lite's `payslipDownloadUrl(id)` calls this via
    # `apiGet` with `Accept: application/json`. Hand back a fully-qualified
    # URL with the caller's token embedded so `Linking.openURL(url)` opens
    # the payslip in the system browser without needing custom headers.
    accept = (request.headers.get('Accept') or '').lower()
    if 'application/json' in accept:
        from django.urls import reverse
        base = request.build_absolute_uri('/').rstrip('/')
        path = reverse('api:mobile_payslip_download', args=[salary.pk])
        download_url = f"{base}{path}?token={request.mobile_token.key}"
        return JsonResponse({'success': True, 'download_url': download_url})

    # Non-JSON = the system browser opened the tokenized URL. Stream a
    # standalone PDF so the phone triggers a real file download instead
    # of rendering the site-wrapped HTML template (which has no auth
    # session anyway and would render blank/broken).
    return _render_payslip_pdf(salary)


def _render_payslip_pdf(salary):
    """Build a single-page A4 payslip PDF via reportlab and return it as
    an attachment. Reads the same fields `PayslipGenerator` uses so the
    numbers match the on-screen payslip exactly."""
    from io import BytesIO
    from decimal import Decimal
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )

    from dashboard.models import CompanySettings
    from accounts.cycle_utils import get_cycle_ending_in_month

    emp     = salary.employee
    bank    = getattr(emp, 'bank_details', None)
    company = CompanySettings.get_settings(emp.admin_id)
    cycle   = get_cycle_ending_in_month(salary.month)

    def d(v):
        return Decimal(str(v or 0))

    basic       = d(salary.basic_salary)
    extra       = d(salary.extra_allowance)
    ot          = d(salary.ot_allowance)
    food_allow  = d(salary.food_allowance)
    advance     = d(salary.advance_pay)
    total_ded   = d(salary.total_deduction)
    food_used   = d(salary.food_usage)
    net_pay     = d(salary.net_pay)

    earnings_total = basic + extra + ot + food_allow

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=15 * mm, bottomMargin=15 * mm,
        title=f"Payslip {emp.employee_id or emp.pk} {salary.month.strftime('%Y-%m')}",
    )
    styles = getSampleStyleSheet()
    h_style   = ParagraphStyle('H',    parent=styles['Title'],  fontSize=16, alignment=1, spaceAfter=2)
    sub_style = ParagraphStyle('Sub',  parent=styles['Normal'], fontSize=9,  alignment=1, textColor=colors.grey)
    lbl_style = ParagraphStyle('Lbl',  parent=styles['Heading4'], fontSize=11, spaceBefore=6, spaceAfter=4)
    small     = ParagraphStyle('Sm',   parent=styles['Normal'], fontSize=8,  textColor=colors.grey, alignment=1)

    watermark = _load_payslip_watermark()

    def _draw_watermark(canvas, _doc):
        """Draws the faded logo centered on every page BEFORE story renders.
        Mirrors the web payslip's `.ps-watermark` (static/img/Logo.png, low
        opacity, centered)."""
        if not watermark:
            return
        try:
            pw, ph = A4
            iw, ih = watermark.getSize()
            w = pw * 0.50
            h = w * (float(ih) / float(iw)) if iw else w
            canvas.saveState()
            canvas.drawImage(
                watermark, (pw - w) / 2, (ph - h) / 2,
                width=w, height=h, mask='auto',
            )
            canvas.restoreState()
        except Exception:
            # Watermark is decorative — never let it break the payslip.
            pass

    story = []
    story.append(Paragraph((company.name if company else 'Company').upper(), h_style))
    if company and company.address:
        story.append(Paragraph(company.address.replace('\n', ' · '), sub_style))
    story.append(Paragraph(f"Payslip for {cycle['label']} ({cycle['start'].strftime('%d %b')} – {cycle['end'].strftime('%d %b %Y')})", sub_style))
    story.append(Spacer(1, 8))

    # Employee + bank block (two columns)
    emp_rows = [
        ['Employee ID',  emp.employee_id or '—', 'Bank',      (bank.bank_name if bank else '—') or '—'],
        ['Name',         emp.name or '—',         'A/C Holder', (bank.account_holder if bank else '—') or '—'],
        ['Designation',  emp.designation or '—',  'A/C No.',   (bank.account_number if bank else '—') or '—'],
        ['Level',        emp.level or '—',        'IFSC',      (bank.ifsc_code if bank else '—') or '—'],
        ['Site',         emp.site or '—',         'Branch',    (bank.branch if bank else '—') or '—'],
    ]
    emp_tbl = Table(emp_rows, colWidths=[28 * mm, 55 * mm, 25 * mm, 55 * mm])
    emp_tbl.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.grey),
        ('TEXTCOLOR', (2, 0), (2, -1), colors.grey),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING',    (0, 0), (-1, -1), 3),
    ]))
    story.append(emp_tbl)
    story.append(Spacer(1, 10))

    def _money(v):
        return '{:,.2f}'.format(float(v))

    # Earnings + Deductions side-by-side
    earn_data = [
        ['Earnings', 'Amount (INR)'],
        ['Base Salary',              _money(basic)],
        ['Attendance / Extra',       _money(extra)],
        ['OT Allowance',             _money(ot)],
        ['Food Allowance',           _money(food_allow)],
        ['Total Earnings',           _money(earnings_total)],
    ]
    ded_data = [
        ['Deductions', 'Amount (INR)'],
        ['Advance Pay',              _money(advance)],
        ['Food Used',                _money(food_used)],
        ['Other Deductions',         _money(max(Decimal('0'), total_ded - advance))],
        ['',                         ''],
        ['Total Deductions',         _money(total_ded)],
    ]

    def _amount_table(data):
        t = Table(data, colWidths=[52 * mm, 30 * mm])
        t.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F1F1F1')),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('LINEBELOW', (0, 0), (-1, 0), 0.5, colors.grey),
            ('LINEABOVE', (0, -1), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING',    (0, 0), (-1, -1), 4),
        ]))
        return t

    side_by_side = Table(
        [[_amount_table(earn_data), _amount_table(ded_data)]],
        colWidths=[85 * mm, 85 * mm],
    )
    side_by_side.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING',  (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))
    story.append(side_by_side)
    story.append(Spacer(1, 12))

    # Net pay banner
    net_tbl = Table(
        [['NET PAY', f'INR {_money(net_pay)}']],
        colWidths=[100 * mm, 70 * mm],
    )
    net_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0B5FFF')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, 0), 13),
        ('ALIGN',      (1, 0), (1, 0),  'RIGHT'),
        ('TOPPADDING',    (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('LEFTPADDING',   (0, 0), (0, 0),  10),
        ('RIGHTPADDING',  (1, 0), (1, 0),  10),
    ]))
    story.append(net_tbl)
    story.append(Spacer(1, 30))

    # Signature line + footer
    sig = Table(
        [['_________________________', '_________________________'],
         ['Employee Signature',        'Authorised Signatory']],
        colWidths=[85 * mm, 85 * mm],
    )
    sig.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN',    (0, 0), (-1, -1), 'CENTER'),
        ('TEXTCOLOR',(0, 1), (-1, 1),  colors.grey),
    ]))
    story.append(sig)
    story.append(Spacer(1, 10))
    story.append(Paragraph(
        "This is a system-generated payslip and does not require a physical signature.",
        small,
    ))

    doc.build(story, onFirstPage=_draw_watermark, onLaterPages=_draw_watermark)
    pdf = buf.getvalue()
    buf.close()

    filename = f"Payslip_{emp.employee_id or emp.pk}_{salary.month.strftime('%Y-%m')}.pdf"
    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    resp['Content-Length']      = str(len(pdf))
    return resp


def _load_payslip_watermark():
    """Return a ReportLab ImageReader for `static/img/Logo.png` pre-faded to
    ~10% alpha (matches the web payslip's `.ps-watermark` treatment), or
    None if the file cannot be located or PIL fails to open it.

    ReportLab's `setFillAlpha` does not reliably affect raster images, so
    the transparency is baked into the PNG's alpha channel via Pillow —
    Pillow is already a project dependency for other image handling."""
    try:
        from django.contrib.staticfiles import finders
        path = finders.find('img/Logo.png')
        if not path:
            return None
        from io import BytesIO
        from PIL import Image
        from reportlab.lib.utils import ImageReader
        im = Image.open(path).convert('RGBA')
        alpha = im.split()[3].point(lambda p: int(p * 0.10))
        im.putalpha(alpha)
        buf = BytesIO()
        im.save(buf, format='PNG')
        buf.seek(0)
        return ImageReader(buf)
    except Exception:
        return None


@csrf_exempt
@require_app_version
@mobile_auth_required
@require_http_methods(['GET', 'POST'])
def mobile_bank_details(request):
    """
    GET  -> the authenticated employee's bank details (matches the `bank`
            block in /api/mobile/profile/ but in its own focused payload).
    POST -> upsert the authenticated employee's bank details.
            Body: { bank_name, account_holder, account_number, ifsc_code, branch }
            Whichever fields are missing keep their current value. The row
            is the same BankDetail OneToOne that the Suite already reads,
            so updates from mobile are visible immediately to admin.
    """
    emp = request.employee

    if request.method == 'POST':
        data = _json_body(request)
        # Pull current row (or create empty) so partial payloads work.
        bank, _ = BankDetail.objects.get_or_create(
            employee=emp,
            defaults={
                'bank_name':      '',
                'account_holder': '',
                'account_number': '',
                'ifsc_code':      '',
            },
        )
        for src, attr in (
            ('bank_name', 'bank_name'),
            ('account_holder', 'account_holder'),
            ('account_number', 'account_number'),
            ('ifsc_code', 'ifsc_code'),
            ('branch', 'branch'),
        ):
            if src in data and data[src] is not None:
                setattr(bank, attr, str(data[src]).strip())
        # An employee-edited row is no longer "verified" until admin re-checks.
        if bank.status == 'verified':
            bank.status = 'modified'
        bank.save()

    bank = getattr(emp, 'bank_details', None)
    return JsonResponse({
        'success': True,
        'bank': {
            'bank_name':      bank.bank_name      if bank else '',
            'account_holder': bank.account_holder if bank else '',
            'account_number': bank.account_number if bank else '',
            'ifsc_code':      bank.ifsc_code      if bank else '',
            'branch':         bank.branch         if bank else '',
            'status':         bank.status         if bank else 'pending',
        },
    })


@require_app_version
@mobile_auth_required
@require_POST
def mobile_logout(request):
    """Revoke the current device token."""
    request.mobile_token.delete()
    return JsonResponse({'success': True})


# ---------------------------------------------------------------------------
# Salary cycle summary (26th prev month -> 25th current month)
# ---------------------------------------------------------------------------

@require_app_version
@mobile_auth_required
@require_http_methods(['GET'])
def mobile_salary(request):
    """
    Current salary cycle summary for the authenticated employee.

    Cycle window: 26th of previous month -> 25th of current month, anchored
    to today. The SalaryUpdate row whose `month` matches the cycle-end's
    year+month is used (SPIM Suite convention). If no row exists yet, all
    monetary fields default to '0.00' so SPIM Lite can render the cycle
    without a separate empty state.

    Notes:
      * `hra` is not tracked on SalaryUpdate today; surfaced as '0.00' to
        keep the response shape stable. Update only the one line below if a
        column is added later.
      * `allowances` = extra_allowance + ot_allowance + food_allowance
        (the three additive components SalaryUpdate already stores).
      * `deductions` = total_deduction (kept as a single rolled-up figure
        to match how the Suite UI shows it).
    """
    # -- TEMPORARY DEBUG INSTRUMENTATION (Phase: regression triage) ----
    # Wrap the entire body so any exception is emitted with a full
    # traceback + the exact employee / cycle context. Nothing else in
    # the function is altered. Remove this whole wrapper once the
    # regression is identified.
    import traceback as _tb
    try:
        return _mobile_salary_impl(request)
    except Exception as _exc:  # noqa: BLE001
        emp_dbg = getattr(request, 'employee', None)
        _log.error(
            '[MOBILE_SALARY_500] emp_pk=%s emp_id=%s name=%s exc=%r\n%s',
            getattr(emp_dbg, 'pk', None),
            getattr(emp_dbg, 'employee_id', None),
            getattr(emp_dbg, 'name', None),
            _exc,
            _tb.format_exc(),
        )
        raise


def _mobile_salary_impl(request):
    """Original mobile_salary body — kept intact so the try/except wrapper
    above logs any regression without altering behavior."""
    emp = request.employee

    # Live view: current cycle. Payslip endpoints intentionally use get_last_completed_cycle().
    # The Salary tab shows a live estimate for the cycle the employee is
    # currently living in (26th of prev month → 25th of this month), with
    # attendance counts and net-pay progressing as the cycle fills. Closed
    # payslips continue to be served by mobile_payslips /
    # mobile_payslip_download, which stay on get_last_completed_cycle so
    # they never surface an in-progress window as a finalised document.
    from accounts.cycle_utils import get_salary_cycle
    from accounts.date_utils import today_ist
    _cycle    = get_salary_cycle(today_ist())
    cycle_start = _cycle['start']
    cycle_end   = _cycle['end']

    sal = SalaryUpdate.objects.filter(
        employee=emp,
        month__year=cycle_end.year,
        month__month=cycle_end.month,
    ).first()

    # -- TEMPORARY DEBUG: per-step state snapshot ----------------------
    _log.info(
        '[MOBILE_SALARY_STEP1] emp_pk=%s emp_id=%s name=%s '
        'cycle_start=%s cycle_end=%s sal_present=%s sal_pk=%s base_salary=%r',
        emp.pk, emp.employee_id, emp.name,
        cycle_start, cycle_end,
        sal is not None, getattr(sal, 'pk', None), emp.base_salary,
    )

    # Attendance counts inside the cycle window (drives present/no-week-off days).
    att_qs = AttendanceRecord.objects.filter(
        employee=emp,
        date__gte=cycle_start,
        date__lte=cycle_end,
    )
    present_days     = att_qs.filter(status='present').count()
    no_week_off_days = att_qs.filter(status='no_week_off').count()

    # Live net salary — same helper the Suite Salary dashboard uses, so SPIM
    # Lite sees the current attendance-prorated payable base instead of a
    # stale SalaryUpdate.net_pay snapshot (which is stamped at Save time and
    # can be 0 for cycles with no manual save yet).
    # Deferred import: employees.views imports from attendance.models, which
    # already depends on employees.models — keeping this import inside the
    # function avoids any chance of an import-time cycle through api.views.
    from employees.views import _compute_attendance_breakdown
    _log.info(
        '[MOBILE_SALARY_STEP2] emp_pk=%s pre_compute base_salary_type=%s '
        'value=%r present_days=%s no_week_off_days=%s',
        emp.pk, type(emp.base_salary).__name__, emp.base_salary,
        present_days, no_week_off_days,
    )
    net_salary_amount, paid_days_dec = _compute_attendance_breakdown(
        emp, cycle_end.replace(day=1), float(emp.base_salary),
    )
    _log.info(
        '[MOBILE_SALARY_STEP3] emp_pk=%s post_compute '
        'net_salary_amount=%r paid_days_dec=%r',
        emp.pk, net_salary_amount, paid_days_dec,
    )
    paid_days    = float(paid_days_dec)
    basic_salary = str(emp.base_salary)
    hra          = '0.00'  # not modelled on SalaryUpdate

    # Cycle-length denominator — same divisor _compute_attendance_breakdown
    # uses (calendar days in the 26→25 window). Surfacing it here means the
    # mobile client displays the exact Suite figure instead of guessing 30.
    total_working_days = (cycle_end - cycle_start).days + 1

    # Attendance earnings = attendance-prorated basic (basic × paid ÷ cycle).
    # This is the number the client's "Attendance Earnings" row displays;
    # returning it explicitly removes the client-side fallback formula and
    # keeps the client purely a pass-through.
    attendance_earnings = str(round(net_salary_amount, 2))
    daily_rate = (
        str(round(float(emp.base_salary) / total_working_days, 2))
        if total_working_days else '0.00'
    )

    # Allowances + deductions still come from the SalaryUpdate row (the
    # admin-entered figures); fall back to 0.00 when no row exists yet.
    if sal:
        # -- TEMPORARY DEBUG: dump raw SalaryUpdate field values so a
        #    None on any nullable-in-DB Decimal is visible in the log. --
        _log.info(
            '[MOBILE_SALARY_STEP4a] emp_pk=%s sal_pk=%s '
            'extra_allowance=%r ot_allowance=%r food_allowance=%r '
            'total_deduction=%r advance_pay=%r',
            emp.pk, sal.pk,
            sal.extra_allowance, sal.ot_allowance, sal.food_allowance,
            sal.total_deduction, sal.advance_pay,
        )
        allowances         = str(sal.extra_allowance + sal.ot_allowance + sal.food_allowance)
        deductions         = str(sal.total_deduction)
        ot_allowance       = str(sal.ot_allowance)
        food_allowance     = str(sal.food_allowance)
        # SalaryUpdate persists the advance figure under `advance_pay`.
        # The old `sal.advance_deduction` reference raised AttributeError
        # (500) for every employee that had a SalaryUpdate row for the
        # current cycle — mobile then rendered every salary field as "—".
        advance_deduction  = str(sal.advance_pay)
    else:
        allowances         = '0.00'
        deductions         = '0.00'
        ot_allowance       = '0.00'
        food_allowance     = '0.00'
        advance_deduction  = '0.00'

    # -- TEMPORARY DEBUG: values entering the float() calls below -----
    _log.info(
        '[MOBILE_SALARY_STEP4b] emp_pk=%s branch=%s '
        'attendance_earnings=%r ot_allowance=%r food_allowance=%r '
        'deductions=%r advance_deduction=%r',
        emp.pk, 'if_sal' if sal else 'else',
        attendance_earnings, ot_allowance, food_allowance,
        deductions, advance_deduction,
    )

    # Live net = attendance earnings + OT + food − deductions (advance is
    # already inside total_deduction on SalaryUpdate; see manage_ajax save).
    net_salary_final = round(
        float(attendance_earnings)
        + float(ot_allowance)
        + float(food_allowance)
        - float(deductions),
        2,
    )
    net_salary = str(net_salary_final)

    _log.info(
        '[MOBILE_SALARY_STEP5] emp_pk=%s net_salary=%s',
        emp.pk, net_salary,
    )

    return JsonResponse({
        'success': True,
        'salary': {
            'basic_salary':        basic_salary,
            'hra':                 hra,
            'allowances':          allowances,
            'deductions':          deductions,
            'net_salary':          net_salary,
            'paid_days':           paid_days,
            'present_days':        present_days,
            'no_week_off_days':    no_week_off_days,
            # DEPRECATED: legacy alias for older APKs. Same value as
            # no_week_off_days. Remove once the SPIM Lite rollout completes.
            'absent_days':         no_week_off_days,
            'cycle_start':         str(cycle_start),
            'cycle_end':           str(cycle_end),
            'attendance_earnings': attendance_earnings,
            'daily_rate':          daily_rate,
            'total_working_days':  total_working_days,
            'overtime_allowance':  ot_allowance,
            'food_allowance':      food_allowance,
            'advance_deduction':   advance_deduction,
        },
    })


# ---------------------------------------------------------------------------
# Dashboard summary (lightweight; for SPIM Lite home tab)
# ---------------------------------------------------------------------------

@require_app_version
@mobile_auth_required
@require_http_methods(['GET'])
def mobile_dashboard(request):
    """
    Focused dashboard payload for the authenticated employee. Mirrors fields
    already surfaced by `mobile_profile` so SPIM Lite's home tab can render
    in a single round-trip instead of pulling the full profile snapshot.

    Fields:
      * employee identity (name, designation, department)
      * today_attendance_status  -> AttendanceRecord.status for today, or None
      * current_month_present_days / current_month_no_week_off_days
        (also current_month_absent_days, DEPRECATED alias for older APKs)
      * latest_net_salary  -> most recent SalaryUpdate.net_pay (or '0.00')
      * company.name / company.logo_url  -> dashboard.CompanySettings
    """
    emp   = request.employee
    today = date.today()

    # Today's attendance (None if not yet marked)
    todays_rec   = AttendanceRecord.objects.filter(employee=emp, date=today).first()
    today_status = todays_rec.status if todays_rec else None

    # Live view: current cycle. Payslip endpoints intentionally use get_last_completed_cycle().
    # Home tab and Salary tab share the same active-cycle window (the 26→25
    # cycle CONTAINING today) so counters progress live as attendance is
    # marked. The `current_month_*` JSON keys are legacy misnomers — the
    # actual filter is the cycle window, not a calendar month.
    from accounts.cycle_utils import get_salary_cycle
    from accounts.date_utils import today_ist
    _cycle      = get_salary_cycle(today_ist())
    cycle_start = _cycle['start']
    cycle_end   = _cycle['end']

    # Cycle attendance counts, capped at today so future-dated rows (auto
    # Sunday holidays etc.) do not inflate the present/no-week-off figures.
    cycle_qs = AttendanceRecord.objects.filter(
        employee=emp,
        date__gte=cycle_start,
        date__lte=cycle_end,
    ).filter(date__lte=today)
    present_days     = cycle_qs.filter(status='present').count()
    # Unpaid days: leave + no_week_off (the retired 'absent' status was
    # merged into 'no_week_off' by migration 0009).
    no_week_off_days = cycle_qs.filter(status__in=['leave', 'no_week_off']).count()

    # Live net salary — same helper the Suite Salary dashboard uses, so the
    # SPIM Lite home tab no longer surfaces a stale SalaryUpdate.net_pay.
    # Deferred import to avoid any import-time cycle through api.views.
    from employees.views import _compute_attendance_breakdown
    latest_net_salary_amount, _paid_dec = _compute_attendance_breakdown(
        emp, cycle_end.replace(day=1), float(emp.base_salary),
    )
    latest_net_salary = str(round(latest_net_salary_amount, 2))

    # Company block — deferred import (matches mobile_profile's pattern).
    company_name = ''
    company_logo = ''
    try:
        from dashboard.models import CompanySettings  # deferred to avoid cycles
        cs = CompanySettings.get_settings(emp.admin_id)
        if cs:
            company_name = cs.name or ''
            try:
                if cs.logo and hasattr(cs.logo, 'url'):
                    company_logo = request.build_absolute_uri(cs.logo.url)
            except Exception:
                company_logo = ''
    except Exception:
        pass

    return JsonResponse({
        'success': True,
        'dashboard': {
            'name':                       emp.name,
            'designation':                emp.designation,
            'department':                 emp.department,
            'today_attendance_status':       today_status,
            'current_month_present_days':    present_days,
            'current_month_no_week_off_days': no_week_off_days,
            # DEPRECATED: legacy alias for older APKs. Same value as
            # current_month_no_week_off_days. Remove once the SPIM Lite
            # rollout completes.
            'current_month_absent_days':     no_week_off_days,
            'latest_net_salary':             latest_net_salary,
            'company': {
                'name':     company_name,
                'logo_url': company_logo,
            },
        },
    })


# ---------------------------------------------------------------------------
# Legacy stub endpoints (kept for backward-compatibility; not wired into urls
# unless the consumer app exists). Imports are deferred so this module always
# loads cleanly even when optional apps are absent.
# ---------------------------------------------------------------------------

@login_required
def api_expenses(request):
    from finance.models import Transaction
    user = request.user
    if request.method == 'GET':
        qs = Transaction.objects.filter(user=user, type='expense').select_related('category').order_by('-date')[:50]
        data = [{
            'id':       t.id,
            'amount':   str(t.amount),
            'category': t.category.name if t.category else None,
            'date':     str(t.date),
        } for t in qs]
        return JsonResponse({'success': True, 'expenses': data})
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@login_required
def api_income(request):
    from income.models import Income
    user = request.user
    if request.method == 'GET':
        qs = Income.objects.filter(user=user).order_by('-date')[:50]
        data = [{
            'id':     i.id,
            'amount': str(i.amount),
            'date':   str(i.date),
        } for i in qs]
        return JsonResponse({'success': True, 'income': data})
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@login_required
def api_dashboard(request):
    from finance.models import Transaction
    from income.models import Income
    user  = request.user
    today = date.today()
    total_income  = Income.objects.filter(user=user).aggregate(total=Sum('amount'))['total'] or 0
    total_expense = Transaction.objects.filter(user=user, type='expense').aggregate(total=Sum('amount'))['total'] or 0
    return JsonResponse({
        'success': True,
        'total_income':  str(total_income),
        'total_expense': str(total_expense),
        'balance':       str(total_income - total_expense),
    })


# ---------------------------------------------------------------------------
# SPIM Lite HR endpoints — read-only, HR-gated, tenant-scoped
# ---------------------------------------------------------------------------
#
# These endpoints exist because every other /api/mobile/* endpoint is
# self-scoped to request.employee. HR needs to view other employees under
# the SAME admin_id (tenant). Every endpoint below is:
#   * bearer-token authenticated (via mobile_hr_required → mobile_auth_required)
#   * HR-gated (403 if the caller isn't HR)
#   * admin_id-scoped (404 if the target employee belongs to another tenant)
#   * read-only (GET only)
#
# Business logic is REUSED from existing helpers — no duplication:
#   * ensure_sunday_holidays  (attendance.utils)
#   * display_status          (attendance.utils — single source of truth)
#   * _compute_attendance_breakdown  (employees.views — same helper mobile_salary uses)
# ---------------------------------------------------------------------------

_HR_KEYWORD = 'hr'


def _is_hr_employee(emp):
    """
    HR detection mirrors SPIM Lite's utils/permissions.ts:isHrUser — scan
    designation / department / level (trim + lowercase) for the substring
    'hr'. Kept as a substring match so both server and client agree on
    who is HR without a schema change.
    """
    if emp is None:
        return False
    for field in (
        getattr(emp, 'designation', '') or '',
        getattr(emp, 'department', '') or '',
        getattr(emp, 'level', '') or '',
    ):
        if _HR_KEYWORD in field.strip().lower():
            return True
    return False


def _hr_effective_admin_id(emp):
    """
    Resolve the effective admin_id (tenant) for an HR mobile caller.

    Mirrors the fallback already used by mobile_attendance for the same
    reason: some Employee rows land with admin_id='PENDING' or blank
    (created before the tenant was finalised). In that case the row's
    creator (Employee.created_by, a Django User) still carries the real
    admin_id, so we prefer that. Without this resolver, HR endpoints
    would filter on 'PENDING' and either return legacy noise or nothing.

    Returns the resolved admin_id string, or '' if unresolvable.
    """
    if emp is None:
        return ''
    direct = (getattr(emp, 'admin_id', '') or '').strip()
    if direct and direct != 'PENDING':
        return direct
    creator = getattr(emp, 'created_by', None)
    if creator is not None:
        creator_admin = (getattr(creator, 'admin_id', '') or '').strip()
        if creator_admin:
            return creator_admin
    return direct  # possibly '' — caller decides how to handle


def mobile_hr_required(view_func):
    """
    Decorator: mobile_auth_required + HR gate. Layers on top of the
    existing token auth without modifying it. Non-HR callers get 403.
    """
    @wraps(view_func)
    @mobile_auth_required
    def _wrapped(request, *args, **kwargs):
        if not _is_hr_employee(request.employee):
            return JsonResponse(
                {'success': False, 'error': 'HR privileges required.'},
                status=403,
            )
        return view_func(request, *args, **kwargs)
    return _wrapped


@require_app_version
@mobile_hr_required
@require_http_methods(['GET'])
def mobile_hr_employees(request):
    """
    GET /api/mobile/hr/employees/

    Returns every employee under the HR caller's admin_id. Mirrors the
    dict shape of employees.views.employee_list_json (so a future Lite
    client can share a picker component with the Suite web view). No
    calculation, no aggregation — just the reused queryset.
    """
    hr_admin_id = _hr_effective_admin_id(request.employee)
    qs = Employee.objects.filter(admin_id=hr_admin_id).order_by('name')
    employees = [{
        'id':           emp.employee_id or '',
        'pk':           emp.pk,
        'name':         emp.name,
        'dept':         emp.department or '',
        'role':         emp.designation or '',
        'mainLocation': emp.location or '',
        'site':         emp.site or '',
        'baseSalary':   float(emp.base_salary) if emp.base_salary else 0,
        'salaryType':   getattr(emp, 'salary_type', 'base_salary') or 'base_salary',
    } for emp in qs]
    return JsonResponse({'success': True, 'employees': employees})


@require_app_version
@mobile_hr_required
@require_http_methods(['GET'])
def mobile_hr_attendance(request):
    """
    GET /api/mobile/hr/attendance/?employee_id=<pk>&date_from=YYYY-MM-DD&date_to=YYYY-MM-DD

    Returns attendance rows for the target employee across the requested
    window. Reuses:
      * ensure_sunday_holidays()  — same Sunday backfill the Suite uses
      * display_status()          — same Sunday-display rule the Suite uses
      * AttendanceRecord queryset — same admin_id-scoped filter shape

    404 if the target employee belongs to another admin_id (tenant isolation).
    """
    hr_admin_id = _hr_effective_admin_id(request.employee)
    emp_pk = (request.GET.get('employee_id') or '').strip()
    date_from = (request.GET.get('date_from') or '').strip()
    date_to = (request.GET.get('date_to') or '').strip()

    if not emp_pk:
        return JsonResponse(
            {'success': False, 'error': 'employee_id is required.'},
            status=400,
        )

    try:
        target = Employee.objects.get(pk=emp_pk, admin_id=hr_admin_id)
    except (Employee.DoesNotExist, ValueError, TypeError):
        # Never leak "wrong tenant" vs "doesn't exist" — both return 404.
        return JsonResponse(
            {'success': False, 'error': 'Employee not found.'},
            status=404,
        )

    # Sunday backfill scoped to this target so the response matches the
    # Suite web view for the same window.
    if date_from:
        backfill_to = date_to
        if not backfill_to:
            try:
                df = datetime.strptime(date_from, '%Y-%m-%d').date()
                next_month = df.replace(day=28) + timedelta(days=4)
                backfill_to = (
                    next_month.replace(day=1) - timedelta(days=1)
                ).isoformat()
            except ValueError:
                backfill_to = ''
        if backfill_to:
            ensure_sunday_holidays(
                hr_admin_id, date_from, backfill_to, employees=[target],
            )

    # `employee=target` is already tenant-scoped — `target` was fetched with
    # `admin_id=hr_admin_id` above. Filtering AttendanceRecord.admin_id here
    # is redundant AND silently drops legacy / mobile-stamped rows whose
    # row-level admin_id is 'PENDING' or blank (same rows the employee's own
    # `mobile_attendance` GET shows, since that view queries only `employee=`).
    qs = AttendanceRecord.objects.filter(
        employee=target,
    ).select_related('employee')
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)
    qs = qs.order_by('-date')

    records = [{
        'empId':       r.employee.employee_id or '',
        'empPk':       r.employee.pk,
        'empName':     r.employee.name,
        'date':        r.date.isoformat(),
        'status':      display_status(r),
        'source':      r.source,
        'site':        getattr(r, 'site', '') or '',
        'workingSite': getattr(r, 'working_site', '') or '',
    } for r in qs]

    return JsonResponse({'success': True, 'records': records})


@require_app_version
@mobile_hr_required
@require_http_methods(['GET'])
def mobile_hr_salary(request):
    """
    GET /api/mobile/hr/salary/?employee_id=<pk>

    Current 26→25 cycle salary summary for the target employee. Reuses
    the same helper mobile_salary uses (_compute_attendance_breakdown),
    so no salary calculation is duplicated.

    404 if the target employee belongs to another admin_id (tenant isolation).
    """
    hr_admin_id = _hr_effective_admin_id(request.employee)
    emp_pk = (request.GET.get('employee_id') or '').strip()

    if not emp_pk:
        return JsonResponse(
            {'success': False, 'error': 'employee_id is required.'},
            status=400,
        )

    try:
        target = Employee.objects.get(pk=emp_pk, admin_id=hr_admin_id)
    except (Employee.DoesNotExist, ValueError, TypeError):
        return JsonResponse(
            {'success': False, 'error': 'Employee not found.'},
            status=404,
        )

    # Live view: current cycle. Payslip endpoints intentionally use get_last_completed_cycle().
    # HR "view employee salary" mirrors mobile_salary's live current-cycle
    # window so the two surfaces agree on the same in-progress figures.
    from accounts.cycle_utils import get_salary_cycle
    from accounts.date_utils import today_ist
    _cycle      = get_salary_cycle(today_ist())
    cycle_start = _cycle['start']
    cycle_end   = _cycle['end']

    sal = SalaryUpdate.objects.filter(
        employee=target,
        month__year=cycle_end.year,
        month__month=cycle_end.month,
    ).first()

    att_qs = AttendanceRecord.objects.filter(
        employee=target,
        date__gte=cycle_start,
        date__lte=cycle_end,
    )
    present_days     = att_qs.filter(status='present').count()
    no_week_off_days = att_qs.filter(status='no_week_off').count()

    # Reuse the exact same helper mobile_salary uses.
    from employees.views import _compute_attendance_breakdown
    net_salary_amount, paid_days_dec = _compute_attendance_breakdown(
        target, cycle_end.replace(day=1), float(target.base_salary),
    )
    net_salary = str(round(net_salary_amount, 2))
    paid_days = float(paid_days_dec)
    basic_salary = str(target.base_salary)
    hra = '0.00'

    if sal:
        allowances = str(sal.extra_allowance + sal.ot_allowance + sal.food_allowance)
        deductions = str(sal.total_deduction)
    else:
        allowances = '0.00'
        deductions = '0.00'

    return JsonResponse({
        'success': True,
        'employee': {
            'pk': target.pk,
            'employee_id': target.employee_id or '',
            'name': target.name,
        },
        'salary': {
            'basic_salary': basic_salary,
            'hra':          hra,
            'allowances':   allowances,
            'deductions':   deductions,
            'net_salary':   net_salary,
            'paid_days':        paid_days,
            'present_days':     present_days,
            'no_week_off_days': no_week_off_days,
            # DEPRECATED: legacy alias for older APKs. Same value as
            # no_week_off_days. Remove once the SPIM Lite rollout completes.
            'absent_days':      no_week_off_days,
            'cycle_start':      str(cycle_start),
            'cycle_end':        str(cycle_end),
        },
    })


# ---------------------------------------------------------------------------
# SPIM Lite HR Income endpoints — read/write, HR-gated, tenant-scoped
# ---------------------------------------------------------------------------
#
# These endpoints wrap the existing Suite Income module:
#   * validation      → income.forms.IncomeForm   (single source of truth)
#   * serialization   → income.views._income_to_dict
#   * queryset shape  → Income.objects.filter(admin_id=…)
#   * location auto-register → income.views._ensure_location_site
#
# Nothing new is duplicated. Income has no employee FK (schema decision —
# search happens on the free-text fields title / source / description /
# payment_by; searching by Employee ID or Name is not supported).
#
# Every mutating endpoint sets Income.user to the Employee's creator
# Django User (Employee.created_by). This mirrors the fallback that
# api.views.mobile_attendance already uses to resolve admin_id for
# mobile-originated writes, so HR-created rows land in the same tenant
# bucket the Suite admin sees.
# ---------------------------------------------------------------------------


def _hr_creator_user(emp):
    """
    Resolve the Django User to stamp on rows created by an HR mobile
    call. Uses Employee.created_by, which is how the Suite already
    ties Employee rows to a tenant. Returns None if unresolvable;
    callers must 500 in that case rather than write a bad FK.
    """
    creator = getattr(emp, 'created_by', None)
    return creator if creator is not None else None


def _hr_income_payload_dict(income):
    """
    Thin wrapper over income.views._income_to_dict so the mobile
    response uses the exact same serializer as the Suite web AJAX
    endpoint. Deferred import avoids pulling income into api's
    module-load path.
    """
    from income.views import _income_to_dict
    return _income_to_dict(income)


def _hr_income_filtered_queryset(hr_admin_id, get_params):
    """
    Build the Income queryset for the caller's tenant using the same
    filter rule as `mobile_hr_income_list`. Extracted so the list and
    the report share one source of truth for `search / category /
    date_from / date_to` — no filter logic is duplicated.

    `get_params` is any dict-like object with .get() (typically
    request.GET). Returns an ordered queryset (Income.Meta.ordering
    already sorts -date, -created_at).
    """
    from income.models import Income
    from django.db.models import Q

    qs = Income.objects.filter(admin_id=hr_admin_id).select_related('category', 'user')

    search    = (get_params.get('search')    or '').strip()
    category  = (get_params.get('category')  or '').strip()
    date_from = (get_params.get('date_from') or '').strip()
    date_to   = (get_params.get('date_to')   or '').strip()

    if search:
        qs = qs.filter(
            Q(title__icontains=search)
            | Q(source__icontains=search)
            | Q(description__icontains=search)
            | Q(payment_by__icontains=search)
        )
    if category:
        try:
            qs = qs.filter(category_id=int(category))
        except (ValueError, TypeError):
            pass
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)

    return qs


@require_app_version
@mobile_hr_required
@require_http_methods(['GET'])
def mobile_hr_income_categories(request):
    """
    GET /api/mobile/hr/income/categories/

    Reused IncomeCategory queryset — scoped to the tenant via the
    category creator's admin_id, matching income.views.income_list.
    """
    from categories.models import IncomeCategory
    hr_admin_id = _hr_effective_admin_id(request.employee)
    qs = IncomeCategory.objects.filter(
        created_by__admin_id=hr_admin_id,
    ).order_by('name')
    categories = [{
        'id':   c.id,
        'name': c.name,
    } for c in qs]
    return JsonResponse({'success': True, 'categories': categories})


@csrf_exempt
@require_app_version
@mobile_hr_required
@require_http_methods(['GET', 'POST'])
def mobile_hr_income_list(request):
    """
    GET  /api/mobile/hr/income/       — list (with filters)
    POST /api/mobile/hr/income/       — create

    List filters (all optional, mirror income.views.income_list):
      * search       — matches title / source / description / payment_by
      * category     — IncomeCategory pk
      * date_from    — YYYY-MM-DD
      * date_to      — YYYY-MM-DD

    Sort order: same as Income.Meta.ordering (-date, -created_at) —
    "latest first" out of the box; no client-side sort needed.

    Search notes: Income is NOT linked to an Employee row (no FK). The
    Suite's income_list already searches title / source / description;
    payment_by is added here so free-text party names (which HR often
    treats as the "person") are matched too. Employee ID / Employee Name
    lookup is not possible against this schema.

    Response:
      { success: True, incomes: [ _income_to_dict(row), … ] }
    """
    from income.forms import IncomeForm
    from income.views import _ensure_location_site

    hr_admin_id = _hr_effective_admin_id(request.employee)

    if request.method == 'GET':
        qs = _hr_income_filtered_queryset(hr_admin_id, request.GET)
        # Income.Meta.ordering already sorts -date, -created_at.
        incomes = [_hr_income_payload_dict(i) for i in qs]
        return JsonResponse({'success': True, 'incomes': incomes})

    # POST — create. Reuses IncomeForm verbatim.
    creator = _hr_creator_user(request.employee)
    if creator is None:
        return JsonResponse({
            'success': False,
            'error': 'HR account is missing a tenant creator. Please contact your admin.',
        }, status=500)

    data = _json_body(request) or {}
    form = IncomeForm(creator, data)
    if not form.is_valid():
        errors = {f: [str(e) for e in errs] for f, errs in form.errors.items()}
        first = next(iter(errors.values()), ['Invalid data.'])[0]
        return JsonResponse(
            {'success': False, 'errors': errors, 'message': first},
            status=400,
        )

    income = form.save(commit=False)
    income.user     = creator
    income.admin_id = hr_admin_id
    income.save()
    _ensure_location_site(hr_admin_id, income.location_site, creator)

    return JsonResponse({
        'success': True,
        'income':  _hr_income_payload_dict(income),
    }, status=201)


@csrf_exempt
@require_app_version
@mobile_hr_required
@require_http_methods(['GET', 'PUT', 'DELETE'])
def mobile_hr_income_detail(request, pk):
    """
    GET    /api/mobile/hr/income/<pk>/  — retrieve
    PUT    /api/mobile/hr/income/<pk>/  — update
    DELETE /api/mobile/hr/income/<pk>/  — delete (mirrors the Suite web
                                          admin behavior; Suite already
                                          supports Income deletion).

    Cross-tenant lookups return 404 to avoid leaking existence.
    """
    from income.models import Income
    from income.forms import IncomeForm
    from income.views import _ensure_location_site

    hr_admin_id = _hr_effective_admin_id(request.employee)

    try:
        income = Income.objects.get(pk=pk, admin_id=hr_admin_id)
    except (Income.DoesNotExist, ValueError, TypeError):
        return JsonResponse(
            {'success': False, 'error': 'Income not found.'},
            status=404,
        )

    if request.method == 'GET':
        return JsonResponse({
            'success': True,
            'income':  _hr_income_payload_dict(income),
        })

    if request.method == 'DELETE':
        income.delete()
        return JsonResponse({'success': True})

    # PUT — update via the same form the Suite uses.
    creator = _hr_creator_user(request.employee)
    if creator is None:
        return JsonResponse({
            'success': False,
            'error': 'HR account is missing a tenant creator. Please contact your admin.',
        }, status=500)

    data = _json_body(request) or {}
    form = IncomeForm(creator, data, instance=income)
    if not form.is_valid():
        errors = {f: [str(e) for e in errs] for f, errs in form.errors.items()}
        first = next(iter(errors.values()), ['Invalid data.'])[0]
        return JsonResponse(
            {'success': False, 'errors': errors, 'message': first},
            status=400,
        )

    income = form.save(commit=False)
    income.admin_id = hr_admin_id  # never let the payload widen scope
    income.save()
    _ensure_location_site(hr_admin_id, income.location_site, creator)

    return JsonResponse({
        'success': True,
        'income':  _hr_income_payload_dict(income),
    })


# ---------------------------------------------------------------------------
# SPIM Lite HR Expense endpoints — read/write, HR-gated, tenant-scoped
# ---------------------------------------------------------------------------
#
# These endpoints wrap the existing Suite Expense module. The Suite stores
# expenses as finance.Transaction rows with type='expense', so:
#   * validation      → finance.forms.TransactionForm   (single source of truth)
#   * serialization   → finance.views._expense_to_dict
#   * queryset shape  → Transaction.objects.filter(admin_id=…, type='expense')
#
# Same admin_id-resolution + creator-user pattern as the HR Income endpoints
# above; nothing new is duplicated.
# ---------------------------------------------------------------------------


def _hr_expense_payload_dict(t):
    """
    Thin wrapper over finance.views._expense_to_dict so the mobile response
    uses the exact same serializer as the Suite web AJAX endpoint. Deferred
    import avoids pulling finance into api's module-load path.
    """
    from finance.views import _expense_to_dict
    return _expense_to_dict(t)


def _hr_expense_filtered_queryset(hr_admin_id, get_params):
    """
    Build the Transaction (type='expense') queryset for the caller's tenant
    using the same filter rule as `mobile_hr_expense_list`. Extracted so
    the list and the report share one source of truth for `search /
    category / date_from / date_to` — no filter logic is duplicated.

    `get_params` is any dict-like object with .get() (typically request.GET
    or a plain dict passed from the report view). Returns an ordered
    queryset (Transaction.Meta.ordering already sorts -date, -created_at).
    """
    from finance.models import Transaction
    from django.db.models import Q

    qs = Transaction.objects.filter(
        admin_id=hr_admin_id, type='expense',
    ).select_related('category', 'branch')

    search    = (get_params.get('search')    or '').strip()
    category  = (get_params.get('category')  or '').strip()
    date_from = (get_params.get('date_from') or '').strip()
    date_to   = (get_params.get('date_to')   or '').strip()

    if search:
        qs = qs.filter(
            Q(description__icontains=search)
            | Q(vendor__icontains=search)
            | Q(reference__icontains=search)
            | Q(purpose__icontains=search)
            | Q(payment_by__icontains=search)
            | Q(income_source__icontains=search)
        )
    if category:
        try:
            qs = qs.filter(category_id=int(category))
        except (ValueError, TypeError):
            pass
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)

    return qs


@require_app_version
@mobile_hr_required
@require_http_methods(['GET'])
def mobile_hr_expense_categories(request):
    """
    GET /api/mobile/hr/expense/categories/

    Reused finance.Category queryset — filtered to type='expense' and the
    caller's tenant, matching finance.views.transaction_list.
    """
    from finance.models import Category
    hr_admin_id = _hr_effective_admin_id(request.employee)
    qs = Category.objects.filter(
        admin_id=hr_admin_id, type='expense',
    ).order_by('name')
    categories = [{
        'id':   c.id,
        'name': c.name,
    } for c in qs]
    return JsonResponse({'success': True, 'categories': categories})


@csrf_exempt
@require_app_version
@mobile_hr_required
@require_http_methods(['GET', 'POST'])
def mobile_hr_expense_list(request):
    """
    GET  /api/mobile/hr/expense/       — list (with filters)
    POST /api/mobile/hr/expense/       — create

    List filters (all optional, mirror finance.views.transaction_list):
      * search       — matches description / vendor / reference / purpose
                        / payment_by / income_source
      * category     — finance.Category pk
      * date_from    — YYYY-MM-DD
      * date_to      — YYYY-MM-DD

    Sort order: same as Transaction.Meta.ordering (-date, -created_at) —
    "latest first" out of the box.

    Response:
      { success: True, expenses: [ _expense_to_dict(row), … ] }
    """
    from finance.forms import TransactionForm
    from branches.models import LocationSite

    hr_admin_id = _hr_effective_admin_id(request.employee)

    if request.method == 'GET':
        qs = _hr_expense_filtered_queryset(hr_admin_id, request.GET)
        expenses = [_hr_expense_payload_dict(t) for t in qs]
        return JsonResponse({'success': True, 'expenses': expenses})

    # POST — create. Reuses TransactionForm verbatim; type is forced server-side.
    creator = _hr_creator_user(request.employee)
    if creator is None:
        return JsonResponse({
            'success': False,
            'error': 'HR account is missing a tenant creator. Please contact your admin.',
        }, status=500)

    data = _json_body(request) or {}
    # Mirror finance.add_transaction: pin type='expense' regardless of payload.
    data = dict(data)
    data['type'] = 'expense'

    form = TransactionForm(creator, data)
    if not form.is_valid():
        errors = {f: [str(e) for e in errs] for f, errs in form.errors.items()}
        first = next(iter(errors.values()), ['Invalid data.'])[0]
        return JsonResponse(
            {'success': False, 'errors': errors, 'message': first},
            status=400,
        )

    t = form.save(commit=False)
    t.user     = creator
    t.type     = 'expense'
    t.admin_id = hr_admin_id
    t.save()
    # Auto-register the LocationSite so it shows up in Suite dropdowns —
    # same behavior as finance.add_transaction.
    if t.location_site:
        name = t.location_site.strip()
        if name and not LocationSite.objects.filter(
            admin_id=hr_admin_id, name__iexact=name,
        ).exists():
            LocationSite.objects.create(
                admin_id=hr_admin_id, name=name, created_by=creator,
            )

    return JsonResponse({
        'success': True,
        'expense': _hr_expense_payload_dict(t),
    }, status=201)


@csrf_exempt
@require_app_version
@mobile_hr_required
@require_http_methods(['GET', 'PUT', 'DELETE'])
def mobile_hr_expense_detail(request, pk):
    """
    GET    /api/mobile/hr/expense/<pk>/  — retrieve
    PUT    /api/mobile/hr/expense/<pk>/  — update
    DELETE /api/mobile/hr/expense/<pk>/  — delete

    Cross-tenant lookups return 404 to avoid leaking existence.
    """
    from finance.models import Transaction
    from finance.forms import TransactionForm
    from branches.models import LocationSite

    hr_admin_id = _hr_effective_admin_id(request.employee)

    try:
        t = Transaction.objects.get(
            pk=pk, admin_id=hr_admin_id, type='expense',
        )
    except (Transaction.DoesNotExist, ValueError, TypeError):
        return JsonResponse(
            {'success': False, 'error': 'Expense not found.'},
            status=404,
        )

    if request.method == 'GET':
        return JsonResponse({
            'success': True,
            'expense': _hr_expense_payload_dict(t),
        })

    if request.method == 'DELETE':
        t.delete()
        return JsonResponse({'success': True})

    # PUT — update via the same form the Suite uses.
    creator = _hr_creator_user(request.employee)
    if creator is None:
        return JsonResponse({
            'success': False,
            'error': 'HR account is missing a tenant creator. Please contact your admin.',
        }, status=500)

    data = _json_body(request) or {}
    data = dict(data)
    data['type'] = 'expense'

    form = TransactionForm(creator, data, instance=t)
    if not form.is_valid():
        errors = {f: [str(e) for e in errs] for f, errs in form.errors.items()}
        first = next(iter(errors.values()), ['Invalid data.'])[0]
        return JsonResponse(
            {'success': False, 'errors': errors, 'message': first},
            status=400,
        )

    obj = form.save(commit=False)
    obj.type     = 'expense'
    obj.admin_id = hr_admin_id  # never let the payload widen scope
    obj.save()
    if obj.location_site:
        name = obj.location_site.strip()
        if name and not LocationSite.objects.filter(
            admin_id=hr_admin_id, name__iexact=name,
        ).exists():
            LocationSite.objects.create(
                admin_id=hr_admin_id, name=name, created_by=creator,
            )

    return JsonResponse({
        'success': True,
        'expense': _hr_expense_payload_dict(obj),
    })


# ---------------------------------------------------------------------------
# SPIM Lite HR Attendance Report — PDF / XLSX download, HR-gated, tenant-scoped
# ---------------------------------------------------------------------------
#
# Streams a binary file (PDF default, XLSX on request) built from the same
# attendance rows the HR Attendance Viewer already reads:
#   * ensure_sunday_holidays  — Sunday backfill (unchanged)
#   * display_status          — Sunday/label mapping (unchanged)
#   * AttendanceRecord queryset — admin_id-scoped filter (unchanged)
#   * openpyxl / reportlab    — same libs the Suite's salary report uses
#
# No new attendance / employee models, forms, or serializers were added.
# ---------------------------------------------------------------------------


_ATTENDANCE_REPORT_SUMMARY_LABELS = (
    'Present',
    'Half Day',
    'Leave',
    'Holiday',
    'Sunday',
    'Weekly Off',
    'No Week Off',
)


def _hr_attendance_report_cycle_default():
    """
    Fallback attendance window when the caller does not supply date_from /
    date_to. Returns the CURRENT 26→25 cycle (the one containing today) —
    when HR opens the report without specifying a range, the default is
    the cycle they are currently operating in.
    """
    # Live view: current cycle. Payslip endpoints intentionally use get_last_completed_cycle().
    from accounts.cycle_utils import get_salary_cycle
    from accounts.date_utils import today_ist
    _cycle = get_salary_cycle(today_ist())
    return _cycle['start'], _cycle['end']


def _hr_attendance_report_rows(hr_admin_id, target_employees, date_from, date_to):
    """
    Backfill Sundays, then stream attendance rows (dicts) + a per-label
    summary counter — no display transform beyond display_status(). The
    resulting list is what both the PDF and XLSX renderers consume.
    """
    if target_employees:
        ensure_sunday_holidays(
            hr_admin_id,
            date_from.isoformat(),
            date_to.isoformat(),
            employees=list(target_employees),
        )

    # `employee__in=target_employees` is already tenant-scoped — every
    # target Employee was fetched with `admin_id=hr_admin_id`. Filtering
    # AttendanceRecord.admin_id here would silently drop legacy or
    # mobile-stamped rows carrying a mismatched row-level admin_id
    # (same divergence fixed in `mobile_hr_attendance`).
    qs = AttendanceRecord.objects.filter(
        employee__in=target_employees,
        date__gte=date_from,
        date__lte=date_to,
    ).select_related('employee').order_by('employee__name', 'date')

    rows = []
    summary = {lbl: 0 for lbl in _ATTENDANCE_REPORT_SUMMARY_LABELS}
    for r in qs:
        status_label = display_status(r)
        rows.append({
            'emp_id':       r.employee.employee_id or '',
            'emp_name':     r.employee.name or '',
            'date':         r.date.isoformat(),
            'date_display': r.date.strftime('%d %b, %Y'),
            'status':       status_label,
            'site':         getattr(r, 'site', '') or '',
            'working_site': getattr(r, 'working_site', '') or '',
        })
        if status_label in summary:
            summary[status_label] += 1
    return rows, summary


def _render_hr_attendance_xlsx(rows, summary, period_label, filename_base, scope_label):
    """XLSX HttpResponse — same styling recipe as _render_salary_xlsx."""
    from io import BytesIO
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = 'Attendance Report'

    headers = ['EMP ID', 'EMP NAME', 'DATE', 'STATUS', 'SITE', 'WORKING SITE']
    ws.append(headers)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1E293B')
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')

    for r in rows:
        ws.append([
            r['emp_id'],
            r['emp_name'],
            r['date_display'],
            r['status'],
            r['site'],
            r['working_site'],
        ])

    ws.append([])
    ws.append([f'Scope: {scope_label}', '', '', '', '', ''])
    ws.append([f'Period: {period_label}', '', '', '', '', ''])
    ws.append([f'Total Rows: {len(rows)}', '', '', '', '', ''])
    ws.append([])
    ws.append(['SUMMARY', '', '', '', '', ''])
    for lbl in _ATTENDANCE_REPORT_SUMMARY_LABELS:
        ws.append([lbl, summary.get(lbl, 0), '', '', '', ''])

    widths = [14, 26, 16, 14, 22, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
    return response


# ---------------------------------------------------------------------------
# Monthly Attendance Register (matrix) — server-side ReportLab twin of the
# client-side html2pdf export in attendance/templates/attendance/index.html
# (see exportRegisterPDF / _buildRegisterMatrix / REGISTER_STATUS_CODE).
#
# Suite generates the register client-side in the browser; the SPIM Lite HR
# report endpoint has no browser to run that JS in, so this helper reproduces
# the same matrix, headers, codes, and layout on the server. Keep it in step
# with the Suite JS — the two PDFs must remain visually identical.
# ---------------------------------------------------------------------------

_REGISTER_STATUS_CODE = {
    # Keys match display_status() output; values match REGISTER_STATUS_CODE
    # in attendance/templates/attendance/index.html.
    'Present':      'P',
    'Half Day':     'HD',
    'Leave':        'L',
    'Holiday':      'H',
    'Sunday':       'S',
    'Weekly Off':   'WO',
    'No Week Off':  'NWO',
}


def _hr_attendance_register_matrix(hr_admin_id, target_employees, date_from, date_to):
    """Mirror of _buildRegisterMatrix() in the Suite web template.

    Returns (dates, rows):
      dates = [{'iso': 'YYYY-MM-DD', 'dom': int}, ...] every day in the window.
      rows  = [{'label': 'EMP_ID / EMP_NAME', 'cells': ['P', '', ...],
                'present': int}, ...] one entry per employee in the order
              the caller passed them in (mirroring Suite's loadEmployees()
              order, which the mobile endpoint sorts by employee name).
    """
    if target_employees:
        ensure_sunday_holidays(
            hr_admin_id,
            date_from.isoformat(),
            date_to.isoformat(),
            employees=list(target_employees),
        )

    qs = AttendanceRecord.objects.filter(
        employee__in=target_employees,
        date__gte=date_from,
        date__lte=date_to,
    ).select_related('employee')

    rec_map = {}
    for r in qs:
        rec_map[(r.employee_id, r.date.isoformat())] = display_status(r)

    dates = []
    cur = date_from
    while cur <= date_to:
        dates.append({'iso': cur.isoformat(), 'dom': cur.day})
        cur += timedelta(days=1)

    rows = []
    for emp in target_employees:
        present = 0
        cells = []
        for d in dates:
            label = rec_map.get((emp.pk, d['iso']))
            if label == 'Present':
                present += 1
            cells.append(_REGISTER_STATUS_CODE.get(label or '', ''))
        rows.append({
            'label':   f"{emp.employee_id or ''} / {emp.name or ''}",
            'cells':   cells,
            'present': present,
        })
    return dates, rows


def _render_hr_attendance_pdf(dates, matrix_rows, cycle_label, filename_base):
    """PDF HttpResponse — Monthly Attendance Register matrix.

    Layout mirrors exportRegisterPDF() in Suite's attendance/templates/
    attendance/index.html so the mobile-downloaded PDF is
    indistinguishable from the one Suite produces client-side.
    """
    from io import BytesIO
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=22, rightMargin=22, topMargin=22, bottomMargin=22,
        title='Monthly Attendance Register',
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'RegisterTitle', parent=styles['Title'],
        fontName='Helvetica-Bold', fontSize=14, alignment=TA_CENTER,
        leading=16, spaceAfter=4,
    )
    cycle_style = ParagraphStyle(
        'RegisterCycle', parent=styles['Normal'],
        fontName='Helvetica', fontSize=10, alignment=TA_CENTER,
        spaceAfter=10,
    )

    story = [
        Paragraph('MONTHLY ATTENDANCE REGISTER', title_style),
        Paragraph(f'Attendance Cycle: {cycle_label}', cycle_style),
    ]

    num_date_cols = len(dates)
    remarks_col = num_date_cols + 1  # 0 = label, 1..N = dates, N+1 = remarks

    # Two-row header — matches Suite JS rowspan/colspan layout:
    #   ┌───────────────┬──── Attendance Dates ────┬──────────┐
    #   │ EMP ID / NAME ├── 26 │ 27 │ … │ 25 ─────┤ Remarks  │
    header_a = ['EMP ID / EMP NAME', 'Attendance Dates']
    header_a.extend([''] * (num_date_cols - 1))
    header_a.append('Remarks')

    header_b = ['']
    header_b.extend(str(d['dom']) for d in dates)
    header_b.append('')

    data = [header_a, header_b]
    if matrix_rows:
        for r in matrix_rows:
            data.append([
                r['label'],
                *r['cells'],
                f"Total Present Days: {r['present']}",
            ])
    else:
        data.append(['-'] + [''] * num_date_cols + ['Total Present Days: 0'])

    # Column widths — mirrors Suite JS colWidths (26/5/…/5/26 wch):
    # label & remarks wide, date cells narrow. Scaled to landscape A4's
    # usable width so the register fits regardless of cycle length.
    label_w, remarks_w = 130, 130
    usable = landscape(A4)[0] - 44  # subtract left+right margins
    day_w = max(14, (usable - label_w - remarks_w) / max(num_date_cols, 1))
    col_widths = [label_w] + [day_w] * num_date_cols + [remarks_w]

    tbl = Table(data, colWidths=col_widths, repeatRows=2)
    tbl.setStyle(TableStyle([
        # 1px #333 grid — Suite HTML uses `border:1px solid #333`.
        ('GRID',       (0, 0), (-1, -1), 0.5, colors.HexColor('#333333')),
        # Header backgrounds — #f1f5f9 (row 0) and #f8fafc (day-number row).
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F1F5F9')),
        ('BACKGROUND', (1, 1), (num_date_cols, 1), colors.HexColor('#F8FAFC')),
        # Merges (rowspan / colspan equivalents from the Suite HTML):
        #   EMP ID / EMP NAME spans both header rows.
        ('SPAN', (0, 0), (0, 1)),
        #   Attendance Dates spans every day column in row 0.
        ('SPAN', (1, 0), (num_date_cols, 0)),
        #   Remarks spans both header rows.
        ('SPAN', (remarks_col, 0), (remarks_col, 1)),
        # Header typography — bold, centered, size 8 to match the compact
        # register look Suite JS produces (9px HTML rendered at scale 2).
        ('FONTNAME', (0, 0), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN',    (0, 0), (-1, 1), 'CENTER'),
        ('VALIGN',   (0, 0), (-1, -1), 'MIDDLE'),
        # Body: label left-aligned & bold, date codes centered, remarks left.
        ('FONTNAME', (0, 2), (0, -1), 'Helvetica-Bold'),
        ('ALIGN',    (0, 2), (0, -1), 'LEFT'),
        ('ALIGN',    (1, 2), (num_date_cols, -1), 'CENTER'),
        ('ALIGN',    (remarks_col, 2), (remarks_col, -1), 'LEFT'),
        # Comfortable padding — matches the 2–4px padding on the HTML cells.
        ('LEFTPADDING',   (0, 0), (-1, -1), 3),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 3),
        ('TOPPADDING',    (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    story.append(tbl)

    doc.build(story)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
    return response


@require_app_version
@mobile_hr_required
@require_http_methods(['GET'])
def mobile_hr_attendance_report(request):
    """
    GET /api/mobile/hr/attendance/report/
        ?format=pdf|xlsx        (default: pdf)
        &employee_id=<pk>|all   (default: all — every employee in the tenant)
        &date_from=YYYY-MM-DD   (default: current 26→25 cycle start)
        &date_to=YYYY-MM-DD     (default: current 26→25 cycle end)

    Bearer token accepted via `Authorization: Bearer` header OR
    `?token=<>` query string — same _extract_token rule the Suite already
    uses for the payslip download link, so the mobile client can open the
    URL directly with `Linking.openURL(...)`.

    404 if a target employee belongs to another admin_id (tenant isolation).
    """
    fmt = (request.GET.get('format') or 'pdf').strip().lower()
    if fmt not in ('pdf', 'xlsx'):
        fmt = 'pdf'

    hr_admin_id = _hr_effective_admin_id(request.employee)

    # ---- Resolve the target employees (single or all) ----
    emp_param = (request.GET.get('employee_id') or 'all').strip().lower()
    if emp_param and emp_param != 'all':
        try:
            target = Employee.objects.get(pk=emp_param, admin_id=hr_admin_id)
        except (Employee.DoesNotExist, ValueError, TypeError):
            return JsonResponse(
                {'success': False, 'error': 'Employee not found.'},
                status=404,
            )
        target_employees = [target]
        scope_label = f"{target.name} ({target.employee_id or target.pk})"
        scope_slug  = target.employee_id or f'emp{target.pk}'
    else:
        target_employees = list(
            Employee.objects.filter(admin_id=hr_admin_id).order_by('name'),
        )
        scope_label = 'All Employees'
        scope_slug  = 'AllEmployees'

    # ---- Resolve the window (custom range or current cycle) ----
    df = (request.GET.get('date_from') or '').strip()
    dt = (request.GET.get('date_to')   or '').strip()
    try:
        date_from = datetime.strptime(df, '%Y-%m-%d').date() if df else None
    except ValueError:
        date_from = None
    try:
        date_to = datetime.strptime(dt, '%Y-%m-%d').date() if dt else None
    except ValueError:
        date_to = None
    if date_from is None or date_to is None:
        cs, ce = _hr_attendance_report_cycle_default()
        date_from = date_from or cs
        date_to   = date_to   or ce
    if date_from > date_to:
        return JsonResponse(
            {'success': False, 'error': 'date_from must be on or before date_to.'},
            status=400,
        )

    period_label  = f"{date_from.isoformat()} to {date_to.isoformat()}"
    filename_base = f"Attendance_Report_{scope_slug}_{date_from.isoformat()}_to_{date_to.isoformat()}"

    if fmt == 'xlsx':
        # XLSX stays row-wise (existing behavior) — the register-matrix
        # parity work is scoped to the PDF, which is what the SPIM Lite
        # HR download screen presents as the primary format.
        rows, summary = _hr_attendance_report_rows(
            hr_admin_id, target_employees, date_from, date_to,
        )
        return _render_hr_attendance_xlsx(
            rows, summary, period_label, filename_base, scope_label,
        )

    # PDF: build the Monthly Attendance Register matrix — same shape as
    # Suite's client-side exportRegisterPDF() so the two PDFs match.
    dates, matrix_rows = _hr_attendance_register_matrix(
        hr_admin_id, target_employees, date_from, date_to,
    )
    cycle_label = (
        f"{date_from.strftime('%d %b %Y')} – "  # en-dash matches Suite
        f"{date_to.strftime('%d %b %Y')}"
    )
    return _render_hr_attendance_pdf(dates, matrix_rows, cycle_label, filename_base)


# ---------------------------------------------------------------------------
# SPIM Lite HR Income Report — PDF / XLSX download, HR-gated, tenant-scoped
# ---------------------------------------------------------------------------
#
# Filters (search / category / date_from / date_to) are applied by
# `_hr_income_filtered_queryset` — the exact same helper `mobile_hr_income_list`
# uses. Row-level fields are read via `_income_to_dict` (existing serializer)
# so the report never diverges from the list. openpyxl / reportlab style
# recipe mirrors the attendance report added in Phase 8.1.
#
# NOTE: Income has no Employee FK. The `search` param is the same substring
# rule the mobile HR Income screen exposes today — it matches `payment_by`
# (which is the "Party / Payer" field HR uses as the person tag). Passing no
# `search` yields "all employees" (every income row in the tenant).
# ---------------------------------------------------------------------------


def _render_hr_income_xlsx(rows, total_amount, period_label, filename_base, scope_label):
    """XLSX HttpResponse — same styling recipe as _render_hr_attendance_xlsx."""
    from io import BytesIO
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = 'Income Report'

    headers = [
        'DATE', 'PARTY (PAYMENT_BY)', 'CATEGORY',
        'LOCATION / SITE', 'PAYMENT MODE', 'DESCRIPTION', 'AMOUNT (Rs)',
    ]
    ws.append(headers)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1E293B')
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')

    for r in rows:
        ws.append([
            r['date_display'],
            r['payment_by'],
            r['category'],
            r['location'],
            r['payment_mode'],
            r['description'],
            float(r['amount'] or 0),
        ])

    ws.append([])
    ws.append([f'Scope: {scope_label}',   '', '', '', '', '', ''])
    ws.append([f'Period: {period_label}', '', '', '', '', '', ''])
    ws.append([f'Total Rows: {len(rows)}', '', '', '', '', 'Total Amount', float(total_amount or 0)])

    widths = [16, 26, 20, 22, 18, 30, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
    return response


def _render_hr_income_pdf(rows, total_amount, period_label, filename_base, scope_label):
    """PDF HttpResponse — same styling recipe as _render_hr_attendance_pdf."""
    from io import BytesIO
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title14', parent=styles['Title'], fontSize=14, alignment=0)
    sub_style   = ParagraphStyle('Sub8',    parent=styles['Normal'], fontSize=8)

    story = [
        Paragraph(f"Income Report — {period_label}", title_style),
        Paragraph(f"Scope: {scope_label}", sub_style),
        Paragraph(
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            sub_style,
        ),
        Spacer(1, 6),
    ]

    table_data = [[
        'DATE', 'PARTY (PAYMENT_BY)', 'CATEGORY',
        'LOCATION / SITE', 'PAYMENT MODE', 'DESCRIPTION', 'AMOUNT (Rs)',
    ]]
    for r in rows:
        table_data.append([
            r['date_display'] or '-',
            r['payment_by']   or '-',
            r['category']     or '-',
            r['location']     or '-',
            r['payment_mode'] or '-',
            r['description']  or '-',
            '{:,.2f}'.format(float(r['amount'] or 0)),
        ])
    if not rows:
        table_data.append(['-', '-', '-', '-', '-', '-', '-'])
    table_data.append([
        f'Total Rows: {len(rows)}', '', '', '', '', 'Total Amount',
        '{:,.2f}'.format(float(total_amount or 0)),
    ])

    col_widths = [70, 120, 80, 100, 75, 165, 75]
    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0),  colors.HexColor('#1E293B')),
        ('TEXTCOLOR',  (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',   (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 8),
        ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN',      (6, 1), (6, -1),  'RIGHT'),
        ('GRID',       (0, 0), (-1, -1), 0.25, colors.HexColor('#CBD5E1')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F0FDF4')),
        ('TEXTCOLOR',  (0, -1), (-1, -1), colors.HexColor('#059669')),
        ('FONTNAME',   (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    story.append(tbl)
    doc.build(story)

    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
    return response


@require_app_version
@mobile_hr_required
@require_http_methods(['GET'])
def mobile_hr_income_report(request):
    """
    GET /api/mobile/hr/income/report/
        ?format=pdf|xlsx        (default: pdf)
        &search=<substring>     (optional — party/payer text, matches
                                 title/source/description/payment_by)
        &category=<pk>          (optional — IncomeCategory pk)
        &date_from=YYYY-MM-DD   (default: current 26→25 cycle start)
        &date_to=YYYY-MM-DD     (default: current 26→25 cycle end)

    Bearer token accepted via `?token=<>` (same _extract_token rule the
    payslip + attendance report links use).
    """
    fmt = (request.GET.get('format') or 'pdf').strip().lower()
    if fmt not in ('pdf', 'xlsx'):
        fmt = 'pdf'

    hr_admin_id = _hr_effective_admin_id(request.employee)

    # ---- Resolve the window (custom range or current cycle) ----
    df = (request.GET.get('date_from') or '').strip()
    dt = (request.GET.get('date_to')   or '').strip()
    try:
        date_from = datetime.strptime(df, '%Y-%m-%d').date() if df else None
    except ValueError:
        date_from = None
    try:
        date_to = datetime.strptime(dt, '%Y-%m-%d').date() if dt else None
    except ValueError:
        date_to = None
    if date_from is None or date_to is None:
        cs, ce = _hr_attendance_report_cycle_default()
        date_from = date_from or cs
        date_to   = date_to   or ce
    if date_from > date_to:
        return JsonResponse(
            {'success': False, 'error': 'date_from must be on or before date_to.'},
            status=400,
        )

    # Reuse the exact same filter block the list endpoint uses. Override
    # date_from / date_to with the resolved ISO strings so the default
    # cycle also flows through the shared helper.
    filter_params = {
        'search':    (request.GET.get('search')   or '').strip(),
        'category':  (request.GET.get('category') or '').strip(),
        'date_from': date_from.isoformat(),
        'date_to':   date_to.isoformat(),
    }
    qs = _hr_income_filtered_queryset(hr_admin_id, filter_params)

    # Serialize + total via the same dict the list uses — no field access
    # duplication.
    rows = [_hr_income_payload_dict(i) for i in qs]
    total_amount = 0.0
    for r in rows:
        try:
            total_amount += float(r.get('amount') or 0)
        except (TypeError, ValueError):
            pass

    # Scope + filename
    search_txt = filter_params['search']
    if search_txt:
        scope_label = f'Party matches "{search_txt}"'
        safe_search = ''.join(ch if ch.isalnum() else '_' for ch in search_txt)[:32]
        scope_slug  = f'party_{safe_search}' if safe_search else 'party'
    else:
        scope_label = 'All Employees / Parties'
        scope_slug  = 'AllParties'

    period_label  = f"{date_from.isoformat()} to {date_to.isoformat()}"
    filename_base = f"Income_Report_{scope_slug}_{date_from.isoformat()}_to_{date_to.isoformat()}"

    if fmt == 'xlsx':
        return _render_hr_income_xlsx(rows, total_amount, period_label, filename_base, scope_label)
    return _render_hr_income_pdf(rows, total_amount, period_label, filename_base, scope_label)


# ---------------------------------------------------------------------------
# SPIM Lite HR Expense Report — PDF / XLSX download, HR-gated, tenant-scoped
# ---------------------------------------------------------------------------
#
# Filters (search / category / date_from / date_to) are applied by
# `_hr_expense_filtered_queryset` — the exact same helper `mobile_hr_expense_list`
# uses. Row-level fields are read via `_expense_to_dict` (existing serializer)
# so the report never diverges from the list. openpyxl / reportlab style
# recipe mirrors the attendance + income reports.
#
# Search behavior is identical to `mobile_hr_expense_list` — substring across
# description / vendor / reference / purpose / payment_by / income_source.
# ---------------------------------------------------------------------------


def _render_hr_expense_xlsx(rows, total_amount, period_label, filename_base, scope_label):
    """XLSX HttpResponse — same styling recipe as _render_hr_income_xlsx."""
    from io import BytesIO
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = 'Expense Report'

    headers = [
        'DATE', 'CATEGORY', 'EXPENSE TYPE', 'LOCATION / SITE',
        'FROM (PAYMENT_BY)', 'TO (VENDOR)', 'PAYMENT MODE',
        'DESCRIPTION', 'AMOUNT (Rs)',
    ]
    ws.append(headers)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1E293B')
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')

    for r in rows:
        ws.append([
            r['date_display'],
            r['category'],
            r['expense_type'],
            r['location'],
            r['payment_by'],
            r['payment_to'],
            r['payment_mode'],
            r['description'],
            float(r['amount'] or 0),
        ])

    ws.append([])
    ws.append([f'Scope: {scope_label}',   '', '', '', '', '', '', '', ''])
    ws.append([f'Period: {period_label}', '', '', '', '', '', '', '', ''])
    ws.append([
        f'Total Rows: {len(rows)}', '', '', '', '', '', '',
        'Total Amount', float(total_amount or 0),
    ])

    widths = [16, 18, 18, 22, 22, 22, 16, 30, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
    return response


def _render_hr_expense_pdf(rows, total_amount, period_label, filename_base, scope_label):
    """PDF HttpResponse — same styling recipe as _render_hr_income_pdf."""
    from io import BytesIO
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title14', parent=styles['Title'], fontSize=14, alignment=0)
    sub_style   = ParagraphStyle('Sub8',    parent=styles['Normal'], fontSize=8)

    story = [
        Paragraph(f"Expense Report — {period_label}", title_style),
        Paragraph(f"Scope: {scope_label}", sub_style),
        Paragraph(
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            sub_style,
        ),
        Spacer(1, 6),
    ]

    table_data = [[
        'DATE', 'CATEGORY', 'EXPENSE TYPE', 'LOCATION / SITE',
        'FROM (PAYMENT_BY)', 'TO (VENDOR)', 'MODE',
        'DESCRIPTION', 'AMOUNT (Rs)',
    ]]
    for r in rows:
        table_data.append([
            r['date_display'] or '-',
            r['category']     or '-',
            r['expense_type'] or '-',
            r['location']     or '-',
            r['payment_by']   or '-',
            r['payment_to']   or '-',
            r['payment_mode'] or '-',
            r['description']  or '-',
            '{:,.2f}'.format(float(r['amount'] or 0)),
        ])
    if not rows:
        table_data.append(['-', '-', '-', '-', '-', '-', '-', '-', '-'])
    table_data.append([
        f'Total Rows: {len(rows)}', '', '', '', '', '', '',
        'Total Amount',
        '{:,.2f}'.format(float(total_amount or 0)),
    ])

    col_widths = [55, 60, 65, 75, 75, 75, 50, 155, 65]
    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0),  colors.HexColor('#1E293B')),
        ('TEXTCOLOR',  (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',   (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 7),
        ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN',      (8, 1), (8, -1),  'RIGHT'),
        ('GRID',       (0, 0), (-1, -1), 0.25, colors.HexColor('#CBD5E1')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F0FDF4')),
        ('TEXTCOLOR',  (0, -1), (-1, -1), colors.HexColor('#059669')),
        ('FONTNAME',   (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    story.append(tbl)
    doc.build(story)

    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
    return response


@require_app_version
@mobile_hr_required
@require_http_methods(['GET'])
def mobile_hr_expense_report(request):
    """
    GET /api/mobile/hr/expense/report/
        ?format=pdf|xlsx        (default: pdf)
        &search=<substring>     (optional — matches description / vendor /
                                 reference / purpose / payment_by /
                                 income_source; same as the list endpoint)
        &category=<pk>          (optional — finance.Category pk, type='expense')
        &date_from=YYYY-MM-DD   (default: current 26→25 cycle start)
        &date_to=YYYY-MM-DD     (default: current 26→25 cycle end)

    Bearer token accepted via `?token=<>` (same _extract_token rule the
    payslip + attendance + income report links use).
    """
    fmt = (request.GET.get('format') or 'pdf').strip().lower()
    if fmt not in ('pdf', 'xlsx'):
        fmt = 'pdf'

    hr_admin_id = _hr_effective_admin_id(request.employee)

    # ---- Resolve the window (custom range or current cycle) ----
    df = (request.GET.get('date_from') or '').strip()
    dt = (request.GET.get('date_to')   or '').strip()
    try:
        date_from = datetime.strptime(df, '%Y-%m-%d').date() if df else None
    except ValueError:
        date_from = None
    try:
        date_to = datetime.strptime(dt, '%Y-%m-%d').date() if dt else None
    except ValueError:
        date_to = None
    if date_from is None or date_to is None:
        cs, ce = _hr_attendance_report_cycle_default()
        date_from = date_from or cs
        date_to   = date_to   or ce
    if date_from > date_to:
        return JsonResponse(
            {'success': False, 'error': 'date_from must be on or before date_to.'},
            status=400,
        )

    # Reuse the exact same filter block the list endpoint uses. Override
    # date_from / date_to with the resolved ISO strings so the default
    # cycle also flows through the shared helper.
    filter_params = {
        'search':    (request.GET.get('search')   or '').strip(),
        'category':  (request.GET.get('category') or '').strip(),
        'date_from': date_from.isoformat(),
        'date_to':   date_to.isoformat(),
    }
    qs = _hr_expense_filtered_queryset(hr_admin_id, filter_params)

    # Serialize + total via the same dict the list uses — no field access
    # duplication.
    rows = [_hr_expense_payload_dict(t) for t in qs]
    total_amount = 0.0
    for r in rows:
        try:
            total_amount += float(r.get('amount') or 0)
        except (TypeError, ValueError):
            pass

    # Scope + filename
    search_txt = filter_params['search']
    if search_txt:
        scope_label = f'Search matches "{search_txt}"'
        safe_search = ''.join(ch if ch.isalnum() else '_' for ch in search_txt)[:32]
        scope_slug  = f'search_{safe_search}' if safe_search else 'search'
    else:
        scope_label = 'All Expenses'
        scope_slug  = 'AllExpenses'

    period_label  = f"{date_from.isoformat()} to {date_to.isoformat()}"
    filename_base = f"Expense_Report_{scope_slug}_{date_from.isoformat()}_to_{date_to.isoformat()}"

    if fmt == 'xlsx':
        return _render_hr_expense_xlsx(rows, total_amount, period_label, filename_base, scope_label)
    return _render_hr_expense_pdf(rows, total_amount, period_label, filename_base, scope_label)


# ---------------------------------------------------------------------------
# SPIM Lite HR Dashboard — GET /api/mobile/hr/dashboard/today/
# ---------------------------------------------------------------------------
#
# Single aggregated call powering the mobile HR dashboard's attendance tiles.
# Reuses existing helpers only — no new attendance rules, no new querysets,
# no per-employee fan-out.
#
# Reused logic:
#   * mobile_hr_required        — auth + HR gate (unchanged)
#   * _hr_effective_admin_id    — tenant resolver with PENDING fallback (unchanged)
#   * Employee queryset filter  — same admin_id-scoped roster mobile_hr_employees uses
#   * AttendanceRecord queryset — same admin_id-scoped filter shape mobile_hr_attendance uses
#   * Raw status literals       — mapped by the existing STATUS_DISPLAY table
#                                 ('present' → Present, 'no_week_off' → No Week Off)
#
# The Suite schema has no "Late" concept (see _ATTENDANCE_REPORT_SUMMARY_LABELS
# — Late is not among the recognised display labels). The endpoint therefore
# returns "late": null, and the mobile client renders "—" for that tile.
#
# Efficiency: two queries total (Employee count, AttendanceRecord group-by-
# status). No Python-side loop over employees.
# ---------------------------------------------------------------------------


@require_app_version
@mobile_hr_required
@require_http_methods(['GET'])
def mobile_hr_dashboard_today(request):
    hr_admin_id = _hr_effective_admin_id(request.employee)
    today = date.today()

    total_employees = Employee.objects.filter(admin_id=hr_admin_id).count()

    # Group today's attendance rows by raw DB status. Labels pass through
    # display_status() unchanged, so aggregating on the raw column is
    # equivalent and lets the database do the counting.
    status_counts = dict(
        AttendanceRecord.objects
        .filter(admin_id=hr_admin_id, date=today)
        .values_list('status')
        .annotate(n=Count('id'))
    )

    # 'absent' status was retired and merged into 'no_week_off' (migration
    # 0009). 'absent' is retained as a DEPRECATED alias so older APKs keep
    # rendering this tile — remove once the SPIM Lite rollout completes.
    nwo_count = int(status_counts.get('no_week_off', 0))
    return JsonResponse({
        'success':         True,
        'date':            today.isoformat(),
        'total_employees': total_employees,
        'present':         int(status_counts.get('present', 0)),
        'no_week_off':     nwo_count,
        'absent':          nwo_count,  # DEPRECATED: dual-emit for legacy APKs
        # Suite has no "Late" attendance status — return null so the mobile
        # client renders "—" without inventing a rule server-side.
        'late':            None,
    })
