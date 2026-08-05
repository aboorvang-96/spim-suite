"""
Material Stock / Inventory backend views.

Two generations live here:

  * LEGACY  — the retired localStorage-sync `state` endpoint, kept verbatim
              (now backed by the renamed Legacy* models).  No longer routed
              from a page, but the endpoint stays so nothing 500s.

  * NEW     — the three-page register:
                Page 1  index()      → Stock       (default)
                Page 2  master()     → Master Control
                Page 3  movements()  → Stock In/Out
              plus tenant-scoped, admin-gated CRUD + autocomplete endpoints.

Every query is tenant-isolated via `get_admin_id(request.user)`.  All
mutations are gated by `projects.decorators.admin_required`.
"""
import json
from datetime import date as date_cls, datetime as datetime_cls
from decimal import Decimal, InvalidOperation

from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.db import transaction, IntegrityError
from django.db.models import Count, Q
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from django.views.decorators.http import require_http_methods, require_POST, require_GET

from accounts.views import get_admin_id, is_admin_user
from projects.decorators import admin_required
from projects.models import Site
from .models import (
    LegacyStockSite, LegacyStockItem, LegacyStockEntry,   # retired UI backend
    Brand, MaterialType, StockItem, StockMovement,
)


def _sites_payload(admin_id):
    """Active-tenant Site list [{id, name}] for header dropdowns."""
    return [{'id': s.id, 'name': s.name}
            for s in Site.objects.filter(admin_id=admin_id, is_active=True)
                                 .order_by('name')]


def _get_tenant_site(admin_id, site_id):
    """Return the Site iff it belongs to this tenant, else None."""
    if not site_id:
        return None
    try:
        return Site.objects.filter(admin_id=admin_id, id=int(site_id)).first()
    except (ValueError, TypeError):
        return None


# ===========================================================================
# Shared helpers
# ===========================================================================

def _to_decimal(v):
    if v is None or v == '':
        return Decimal('0')
    try:
        return Decimal(str(v))
    except (InvalidOperation, ValueError, TypeError):
        return Decimal('0')


def _to_date(v):
    if not v:
        return None
    if isinstance(v, date_cls):
        return v
    try:
        return datetime_cls.strptime(str(v)[:10], '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None


def _to_datetime(v):
    if not v:
        return None
    if isinstance(v, datetime_cls):
        return v
    s = str(v)
    for fmt in ('%Y-%m-%dT%H:%M:%S.%f', '%Y-%m-%dT%H:%M:%S',
                '%Y-%m-%dT%H:%M',       '%Y-%m-%d %H:%M:%S',
                '%Y-%m-%d'):
        try:
            return datetime_cls.strptime(s[:len(fmt) + 6], fmt)
        except (ValueError, TypeError):
            continue
    return None


def _body(request):
    """Parse a JSON request body, tolerating form-encoded posts."""
    try:
        if request.content_type and 'application/json' in request.content_type:
            return json.loads(request.body or b'{}')
    except (ValueError, TypeError):
        return {}
    return request.POST


def _err(msg, status=400):
    return JsonResponse({'success': False, 'error': msg}, status=status)


class _BatchError(Exception):
    """Raised inside a batch movement to force an all-or-nothing rollback."""
    def __init__(self, serial_id, serial, message):
        self.serial_id = serial_id
        self.serial    = serial
        self.message   = message
        super().__init__(message)


def _clean(v, maxlen=None):
    if v is None:
        return ''
    s = str(v).strip()
    return s[:maxlen] if maxlen else s


def _brands_payload(admin_id):
    """Brands + their material types, for client-side dropdown filtering."""
    out = []
    brands = (Brand.objects.filter(admin_id=admin_id, is_active=True)
              .prefetch_related('material_types'))
    for b in brands:
        out.append({
            'id':       b.id,
            'name':     b.name,
            'category': b.category,
            'types': [
                {'id': mt.id, 'name': mt.name, 'tracking_mode': mt.tracking_mode}
                for mt in b.material_types.all() if mt.is_active
            ],
        })
    return out


# ===========================================================================
# PAGE 1 — Stock  (default landing page)
# ===========================================================================

def _stock_report_rows(admin_id, params, today):
    """
    Shared summary-row builder used by index() and the PDF/Excel exports.
    Groups by (material_type + purchase_date). Site is NOT a grouping key:
    the column shows the site name only when every item in the group shares
    one site, otherwise it shows an em-dash.

    `params` is a plain dict of GET-style filters:
        serial, site_id, equipment_id, material_type_id,
        pdate_from, pdate_to, has_in ('1'), has_out ('1')
    """
    items_qs = (StockItem.objects.filter(admin_id=admin_id)
                .select_related('material_type', 'material_type__brand', 'site'))

    serial = (params.get('serial') or '').strip()
    if serial:
        items_qs = items_qs.filter(serial_no__icontains=serial)
    try:
        if params.get('site_id'):
            items_qs = items_qs.filter(site_id=int(params['site_id']))
    except (ValueError, TypeError):
        pass
    try:
        if params.get('equipment_id'):
            items_qs = items_qs.filter(material_type__brand_id=int(params['equipment_id']))
    except (ValueError, TypeError):
        pass
    try:
        if params.get('material_type_id'):
            items_qs = items_qs.filter(material_type_id=int(params['material_type_id']))
    except (ValueError, TypeError):
        pass
    pd_from = _to_date(params.get('pdate_from'))
    if pd_from:
        items_qs = items_qs.filter(purchase_date__gte=pd_from)
    pd_to = _to_date(params.get('pdate_to'))
    if pd_to:
        items_qs = items_qs.filter(purchase_date__lte=pd_to)

    items = list(items_qs)

    # Movement in/out counts per stock item (single query, unfiltered — we
    # want the item's real balance regardless of the row-level filters).
    mv_rows = (StockMovement.objects.filter(admin_id=admin_id)
               .values('stock_item_id')
               .annotate(
                   outs=Count('id', filter=Q(to_person__isnull=False) & ~Q(to_person='')),
                   ins=Count('id',  filter=Q(to_person__isnull=True) | Q(to_person='')),
               ))
    mv_map = {r['stock_item_id']: (r['ins'], r['outs']) for r in mv_rows}

    groups = {}
    for it in items:
        key = (it.material_type_id, it.purchase_date)
        g = groups.get(key)
        if g is None:
            g = groups[key] = {
                'material_type_id': it.material_type_id,
                'material_label':   f"{it.material_type.brand.name} · {it.material_type.name}",
                'purchase_date':    it.purchase_date,
                'is_new':           it.purchase_date == today,
                'quantity':         0,
                'stock_in':         0,
                'stock_out':        0,
                '_site_ids':        set(),
                '_site_name':       '',
            }
        g['quantity'] += 1
        ins, outs = mv_map.get(it.id, (0, 0))
        g['stock_in']  += ins
        g['stock_out'] += outs
        g['_site_ids'].add(it.site_id)
        if it.site_id and not g['_site_name']:
            g['_site_name'] = it.site.name

    for g in groups.values():
        # Single-site group iff exactly one non-null site is present.
        sids = {s for s in g['_site_ids'] if s is not None}
        g['site_label'] = g['_site_name'] if len(sids) == 1 else ''
        g['balance'] = g['quantity'] + g['stock_in'] - g['stock_out']
        del g['_site_ids']
        del g['_site_name']

    group_list = sorted(groups.values(),
                        key=lambda g: (g['material_label'],
                                       g['purchase_date'] or date_cls.min))

    # Post-aggregation filters (need summed values).
    if params.get('has_in') in ('1', 'true', 'on'):
        group_list = [g for g in group_list if g['stock_in'] > 0]
    if params.get('has_out') in ('1', 'true', 'on'):
        group_list = [g for g in group_list if g['stock_out'] > 0]

    for i, g in enumerate(group_list, start=1):
        g['sl'] = i
    return group_list


@login_required
def index(request):
    admin_id = get_admin_id(request.user)
    today    = timezone.localdate()

    filters = {
        'serial':           request.GET.get('serial', ''),
        'site_id':          request.GET.get('site_id', ''),
        'equipment_id':     request.GET.get('equipment_id', ''),
        'material_type_id': request.GET.get('material_type_id', ''),
        'pdate_from':       request.GET.get('pdate_from', ''),
        'pdate_to':         request.GET.get('pdate_to', ''),
        'has_in':           request.GET.get('has_in', ''),
        'has_out':          request.GET.get('has_out', ''),
    }
    group_list = _stock_report_rows(admin_id, filters, today)

    # Filter-bar dropdown data.
    equipments = list(Brand.objects.filter(admin_id=admin_id, is_active=True)
                      .order_by('name').values('id', 'name'))
    mtypes = list(MaterialType.objects.filter(admin_id=admin_id, is_active=True)
                  .select_related('brand').order_by('brand__name', 'name'))
    mtype_rows = [{'id': m.id, 'brand_id': m.brand_id,
                   'label': f'{m.brand.name} · {m.name}'} for m in mtypes]

    return render(request, 'material_stock/page_stock.html', {
        'active_page': 'stock',
        'is_admin':    is_admin_user(request.user),
        'groups':      group_list,
        'brands':      _brands_payload(admin_id),
        'today':       today,
        'sites':       _sites_payload(admin_id),
        'equipments':  equipments,
        'mtypes_all':  mtype_rows,
        'filters':     filters,
    })


def _report_filename(ext):
    now = timezone.localtime()
    return f'material_stock_report_{now.strftime("%Y-%m-%d_%H%M")}.{ext}'


@login_required
def stock_report_export(request):
    """
    GET /material-stock/report/download/?format=pdf|xlsx&<filters>
    Downloads the currently-filtered Stock Summary. Reuses the same
    reportlab + openpyxl patterns as projects.machine_history_download
    (commit ce3a6bb6).
    """
    admin_id = get_admin_id(request.user)
    today    = timezone.localdate()
    filters = {k: request.GET.get(k, '') for k in
               ('serial', 'site_id', 'equipment_id', 'material_type_id',
                'pdate_from', 'pdate_to', 'has_in', 'has_out')}
    rows = _stock_report_rows(admin_id, filters, today)
    fmt  = (request.GET.get('format') or 'pdf').lower()

    if fmt == 'xlsx':
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = 'Stock Summary'

        thin   = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill('solid', fgColor='F1F5F9')

        ws.cell(row=1, column=1, value='Material Stock Report').font = Font(bold=True, size=13)
        ws.cell(row=2, column=1, value=f'Generated: {today.strftime("%Y-%m-%d")}  •  Rows: {len(rows)}'
                ).font = Font(italic=True, color='64748B')

        headers = ['SL No', 'Purchase Date', 'Equipment', 'Material Type',
                   'Site', 'Quantity', 'Stock In', 'Stock Out', 'Balance']
        for ci, h in enumerate(headers, start=1):
            c = ws.cell(row=4, column=ci, value=h)
            c.font = Font(bold=True)
            c.fill = header_fill
            c.alignment = Alignment(vertical='center')

        for ri, g in enumerate(rows, start=5):
            # material_label = "Brand · Material" — split for display.
            brand_part, _, mt_part = g['material_label'].partition(' · ')
            ws.cell(row=ri, column=1, value=g['sl'])
            ws.cell(row=ri, column=2,
                    value=g['purchase_date'].strftime('%d %b %Y') if g['purchase_date'] else '')
            ws.cell(row=ri, column=3, value=brand_part)
            ws.cell(row=ri, column=4, value=mt_part or g['material_label'])
            ws.cell(row=ri, column=5, value=g['site_label'] or '—')
            ws.cell(row=ri, column=6, value=g['quantity']).alignment = Alignment(horizontal='center')
            ws.cell(row=ri, column=7, value=g['stock_in']).alignment = Alignment(horizontal='center')
            ws.cell(row=ri, column=8, value=g['stock_out']).alignment = Alignment(horizontal='center')
            ws.cell(row=ri, column=9, value=g['balance']).alignment = Alignment(horizontal='center')

        last_row = max(4, 4 + len(rows))
        for rr in range(4, last_row + 1):
            for cc in range(1, len(headers) + 1):
                ws.cell(row=rr, column=cc).border = border

        widths = [7, 14, 20, 22, 20, 10, 10, 11, 10]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        buf = BytesIO()
        wb.save(buf)
        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="{_report_filename("xlsx")}"'
        return resp

    if fmt == 'pdf':
        from io import BytesIO
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_LEFT

        buf = BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=landscape(A4),
            leftMargin=1*cm, rightMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm,
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'TitleBold', parent=styles['Normal'],
            fontName='Helvetica-Bold', fontSize=14, spaceAfter=4, alignment=TA_LEFT,
        )
        sub_style = ParagraphStyle(
            'Sub', parent=styles['Normal'],
            fontName='Helvetica', fontSize=9, textColor=colors.HexColor('#475569'), spaceAfter=8,
        )
        cell_style = ParagraphStyle(
            'Cell', parent=styles['Normal'], fontName='Helvetica', fontSize=8, leading=10,
        )

        def _wrap(txt):
            safe = (str(txt) if txt is not None else '').replace('&', '&amp;')\
                .replace('<', '&lt;').replace('>', '&gt;')
            return Paragraph(safe or '—', cell_style)

        story = [
            Paragraph('Material Stock Report', title_style),
            Paragraph(f'Generated: {today.strftime("%Y-%m-%d")} &nbsp;•&nbsp; Rows: {len(rows)}', sub_style),
            Spacer(1, 4),
        ]

        headers = ['SL', 'Purchase Date', 'Equipment', 'Material Type',
                   'Site', 'Qty', 'Stock In', 'Stock Out', 'Balance']
        data = [headers]
        for g in rows:
            brand_part, _, mt_part = g['material_label'].partition(' · ')
            data.append([
                str(g['sl']),
                g['purchase_date'].strftime('%d %b %Y') if g['purchase_date'] else '',
                _wrap(brand_part),
                _wrap(mt_part or g['material_label']),
                _wrap(g['site_label'] or '—'),
                str(g['quantity']),
                str(g['stock_in']),
                str(g['stock_out']),
                str(g['balance']),
            ])
        col_widths = [1.0*cm, 2.6*cm, 4.5*cm, 5.5*cm, 4.5*cm,
                      1.4*cm, 1.9*cm, 1.9*cm, 1.9*cm]
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 0), (-1, 0), 9),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F1F5F9')),
            ('GRID',       (0, 0), (-1, -1), 0.3, colors.HexColor('#CBD5E1')),
            ('VALIGN',     (0, 0), (-1, -1), 'TOP'),
            ('ALIGN',      (0, 0), (1, -1), 'CENTER'),
            ('ALIGN',      (5, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE',   (0, 1), (-1, -1), 8),
        ]))
        story.append(t)
        if not rows:
            story.append(Paragraph('No stock rows match the current filters.', sub_style))
        doc.build(story)
        resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="{_report_filename("pdf")}"'
        return resp

    return HttpResponse('Unsupported format. Use ?format=xlsx or pdf.', status=400)


# ===========================================================================
# PAGE 2 — Master Control
# ===========================================================================

@login_required
def master(request):
    admin_id = get_admin_id(request.user)
    today    = timezone.localdate()

    brands = list(Brand.objects.filter(admin_id=admin_id).order_by('name'))

    # Selected brand (?brand=<id>) or first.
    sel_id = request.GET.get('brand')
    selected = None
    if sel_id:
        selected = next((b for b in brands if str(b.id) == str(sel_id)), None)
    if selected is None and brands:
        selected = brands[0]

    ctx = {
        'active_page':        'master',
        'is_admin':           is_admin_user(request.user),
        'brands':             brands,
        'equipment_brands':   [b for b in brands if b.section != Brand.MACHINERIES],
        'machineries_brands': [b for b in brands if b.section == Brand.MACHINERIES],
        'selected':           selected,
        'today':              today,
    }

    if selected:
        mtypes = list(selected.material_types.filter(is_active=True).order_by('name'))
        # Payload for the in-page Add Stock Item form (material dropdown).
        ctx['mtype_options'] = [
            {'id': mt.id, 'name': mt.name, 'is_batched': mt.is_batched}
            for mt in mtypes
        ]
        items_by_type = {}
        items = (StockItem.objects.filter(admin_id=admin_id, material_type__in=mtypes)
                 .select_related('site')
                 .order_by('purchase_date', 'serial_no'))
        for it in items:
            it.is_new = (it.purchase_date == today)
            it.site_label = it.site.name if it.site_id else ''
            items_by_type.setdefault(it.material_type_id, []).append(it)
        # Attach items + a running SL to each material type for the template.
        blocks = []
        for mt in mtypes:
            blocks.append({'mtype': mt, 'items': items_by_type.get(mt.id, [])})
        ctx['blocks'] = blocks

    ctx['sites'] = _sites_payload(admin_id)
    return render(request, 'material_stock/page_master.html', ctx)


# ===========================================================================
# PAGE 3 — Stock In/Out
# ===========================================================================

@login_required
def movements(request):
    admin_id = get_admin_id(request.user)
    today    = timezone.localdate()

    # Selected month as YYYY-MM (default: current month).
    month = request.GET.get('month') or today.strftime('%Y-%m')
    try:
        y, m = (int(x) for x in month.split('-')[:2])
        month_start = date_cls(y, m, 1)
    except (ValueError, TypeError):
        month_start = today.replace(day=1)
        month = month_start.strftime('%Y-%m')
    # First day of next month.
    if month_start.month == 12:
        next_start = date_cls(month_start.year + 1, 1, 1)
    else:
        next_start = date_cls(month_start.year, month_start.month + 1, 1)

    # Bounds for the in-card Date input. Never past today (validate_not_future
    # would reject anyway). Default: today if we're on the current month,
    # otherwise the 1st of the selected month.
    from datetime import timedelta as _td
    month_end   = next_start - _td(days=1)
    is_current  = (today.year, today.month) == (month_start.year, month_start.month)
    date_max    = today if today <= month_end else month_end
    date_min    = month_start
    date_default = today if is_current else month_start

    history = (StockMovement.objects
               .filter(admin_id=admin_id,
                       movement_date__gte=month_start,
                       movement_date__lt=next_start)
               .select_related('source_site', 'destination_site',
                               'stock_item', 'stock_item__material_type',
                               'stock_item__material_type__brand'))

    # Material types that actually track stock (exclude catalog-only brands).
    mtypes = (MaterialType.objects
              .filter(admin_id=admin_id, is_active=True,
                      brand__category=Brand.TRACKED, brand__is_active=True)
              .select_related('brand').order_by('brand__name', 'name'))
    mtype_payload = [
        {'id': mt.id, 'brand_id': mt.brand_id, 'label': f"{mt.brand.name} · {mt.name}"}
        for mt in mtypes
    ]

    # Build the header Month dropdown: current + previous 12 + next 1 = 14
    # options, newest first. Uses YYYY-MM strings; the view resolves them.
    months = []
    # Next month first (top of list).
    if today.month == 12:
        ny, nm = today.year + 1, 1
    else:
        ny, nm = today.year, today.month + 1
    months.append(date_cls(ny, nm, 1).strftime('%Y-%m'))
    # Then current month and the previous 12.
    cy, cm = today.year, today.month
    for _ in range(13):
        months.append(date_cls(cy, cm, 1).strftime('%Y-%m'))
        cm -= 1
        if cm == 0:
            cm = 12
            cy -= 1

    return render(request, 'material_stock/page_movements.html', {
        'active_page':   'movements',
        'is_admin':      is_admin_user(request.user),
        'history':       history,
        'mtypes':        mtype_payload,
        'sites':         _sites_payload(admin_id),
        'month':         month,
        'months':        months,
        'date_default':  date_default,
        'date_min':      date_min,
        'date_max':      date_max,
        'today':         today,
    })


# ===========================================================================
# CRUD — Stock Items  (Page 1 quick-add, Page 2 edit/delete)
# ===========================================================================

@login_required
@admin_required
@require_POST
def item_add(request):
    admin_id = get_admin_id(request.user)
    data = _body(request)

    brand = get_object_or_404(Brand, id=data.get('brand_id') or 0, admin_id=admin_id)

    # All brands are serial-tracked → StockItem row.
    mt = get_object_or_404(MaterialType, id=data.get('material_type_id') or 0,
                           admin_id=admin_id, brand=brand)
    serial = _clean(data.get('serial_no'), 120)
    if not serial:
        return _err('Serial number is required.')

    purchase_date = _to_date(data.get('purchase_date')) or timezone.localdate()

    batch_no = mfg_date = None
    if mt.is_batched:
        batch_no = _clean(data.get('batch_no'), 120) or None
        mfg_date = _to_date(data.get('mfg_date'))

    # Site FK — optional at Stock-page add time (nullable after 0007).
    # The + Add Serial popover on Stock In/Out still enforces site via
    # quick_add_item, which validates before dispatching here.
    site = _get_tenant_site(admin_id, data.get('site_id'))
    if data.get('site_id') and site is None:
        return _err('Invalid site for this account.', status=400)

    # Friendly duplicate check before the DB constraint fires.
    if StockItem.objects.filter(admin_id=admin_id, material_type=mt,
                                serial_no=serial).exists():
        return _err(f"Serial '{serial}' already exists for {mt.name}.", status=409)

    try:
        item = StockItem.objects.create(
            admin_id=admin_id, material_type=mt, serial_no=serial,
            site=site,
            batch_no=batch_no, mfg_date=mfg_date, purchase_date=purchase_date,
            remarks=_clean(data.get('remarks')) or None,
            created_by=request.user if request.user.is_authenticated else None,
        )
    except IntegrityError:
        return _err(f"Serial '{serial}' already exists for {mt.name}.", status=409)
    except ValidationError as e:
        return _err('; '.join(e.messages))

    return JsonResponse({
        'success': True,
        'kind':    'item',
        'id':      item.id,
        'serial':  item.serial_no,
        'label':   f"{mt.brand.name} · {mt.name}",
    })


@login_required
@admin_required
@require_POST
def item_edit(request, pk):
    admin_id = get_admin_id(request.user)
    data = _body(request)
    item = get_object_or_404(StockItem, id=pk, admin_id=admin_id)

    serial = _clean(data.get('serial_no'), 120)
    if not serial:
        return _err('Serial number is required.')

    dup = (StockItem.objects
           .filter(admin_id=admin_id, material_type=item.material_type, serial_no=serial)
           .exclude(pk=item.pk).exists())
    if dup:
        return _err(f"Serial '{serial}' already exists for {item.material_type.name}.", status=409)

    pdate = _to_date(data.get('purchase_date'))
    if pdate:
        item.purchase_date = pdate
    item.serial_no = serial
    if item.material_type.is_batched:
        item.batch_no = _clean(data.get('batch_no'), 120) or None
        item.mfg_date = _to_date(data.get('mfg_date'))
    status = _clean(data.get('status'), 20)
    if status in dict(StockItem.STATUS_CHOICES):
        item.status = status
    if 'current_holder_name' in data:
        item.current_holder_name = _clean(data.get('current_holder_name'), 150) or None
    if 'current_site_name' in data:
        item.current_site_name = _clean(data.get('current_site_name'), 200) or None
    if 'site_id' in data:
        new_site = _get_tenant_site(admin_id, data.get('site_id'))
        if new_site is None:
            return _err('Site is required.', status=400)
        item.site = new_site
    item.remarks = _clean(data.get('remarks')) or None

    try:
        item.save()
    except IntegrityError:
        return _err(f"Serial '{serial}' already exists.", status=409)
    except ValidationError as e:
        return _err('; '.join(e.messages))
    return JsonResponse({'success': True})


@login_required
@admin_required
@require_POST
def item_delete(request, pk):
    admin_id = get_admin_id(request.user)
    item = get_object_or_404(StockItem, id=pk, admin_id=admin_id)

    n = StockMovement.objects.filter(admin_id=admin_id, stock_item=item).count()
    if n:
        return _err(f"Cannot delete — item has {n} movement record(s). "
                    f"This is historical data.", status=409)
    item.delete()
    return JsonResponse({'success': True})


# ===========================================================================
# CRUD — Brands & Material Types  (Page 2)
# ===========================================================================

@login_required
@admin_required
@require_POST
def brand_add(request):
    admin_id = get_admin_id(request.user)
    data = _body(request)
    name = _clean(data.get('name'), 120)
    if not name:
        return _err('Equipment name is required.')
    category = _clean(data.get('category'), 20)
    if category not in dict(Brand.CATEGORY_CHOICES):
        category = Brand.TRACKED
    section = _clean(data.get('section'), 20)
    if section not in dict(Brand.SECTION_CHOICES):
        section = Brand.EQUIPMENT
    if Brand.objects.filter(admin_id=admin_id, name__iexact=name).exists():
        return _err(f"Equipment '{name}' already exists.", status=409)
    try:
        b = Brand.objects.create(admin_id=admin_id, name=name,
                                 category=category, section=section)
    except IntegrityError:
        return _err(f"Equipment '{name}' already exists.", status=409)
    return JsonResponse({'success': True, 'id': b.id})


@login_required
@admin_required
@require_POST
def brand_delete(request, pk):
    admin_id = get_admin_id(request.user)
    brand = get_object_or_404(Brand, id=pk, admin_id=admin_id)
    n = MaterialType.objects.filter(admin_id=admin_id, brand=brand).count()
    if n:
        return _err(f"Cannot delete — equipment has {n} material type(s). "
                    f"Remove them first.", status=409)
    brand.delete()
    return JsonResponse({'success': True})


@login_required
@admin_required
@require_POST
def mtype_add(request):
    admin_id = get_admin_id(request.user)
    data = _body(request)
    brand = get_object_or_404(Brand, id=data.get('brand_id') or 0, admin_id=admin_id)
    name = _clean(data.get('name'), 120)
    if not name:
        return _err('Material type name is required.')
    mode = _clean(data.get('tracking_mode'), 20)
    if mode not in dict(MaterialType.TRACKING_CHOICES):
        mode = MaterialType.INDIVIDUAL
    if MaterialType.objects.filter(admin_id=admin_id, brand=brand, name__iexact=name).exists():
        return _err(f"'{name}' already exists under {brand.name}.", status=409)
    try:
        mt = MaterialType.objects.create(admin_id=admin_id, brand=brand,
                                         name=name, tracking_mode=mode)
    except IntegrityError:
        return _err(f"'{name}' already exists under {brand.name}.", status=409)
    return JsonResponse({'success': True, 'id': mt.id})


@login_required
@admin_required
@require_POST
def mtype_delete(request, pk):
    admin_id = get_admin_id(request.user)
    mt = get_object_or_404(MaterialType, id=pk, admin_id=admin_id)
    n = StockItem.objects.filter(admin_id=admin_id, material_type=mt).count()
    if n:
        return _err(f"Cannot delete — {n} item(s) use this material type.", status=409)
    mt.delete()
    return JsonResponse({'success': True})


# ===========================================================================
# CRUD — Electric Motor Catalog  (RETIRED)
#
# Electric Motor is now a normal serial-tracked brand (migration 0005 moved its
# catalog rows into StockItem).  The three endpoints below are kept ONLY as
# 410 Gone stubs so a stale browser tab posting to the old catalog URLs gets a
# clear message instead of a mystery 500.  The routes remain wired in urls.py.
# ===========================================================================

_MOTOR_GONE = ('Electric Motor is now managed as tracked stock items. '
               'Please use the new Add Stock Item flow.')


@login_required
@admin_required
@require_POST
def motor_add(request):
    return JsonResponse({'success': False, 'error': _MOTOR_GONE}, status=410)


@login_required
@admin_required
@require_POST
def motor_edit(request, pk):
    return JsonResponse({'success': False, 'error': _MOTOR_GONE}, status=410)


@login_required
@admin_required
@require_POST
def motor_delete(request, pk):
    return JsonResponse({'success': False, 'error': _MOTOR_GONE}, status=410)


# ===========================================================================
# CRUD — Stock Movements  (Page 3) — atomic in/out + reverse-on-delete
# ===========================================================================

def _collect_serial_ids(raw):
    """Normalise anything sane into a de-duped ordered list of positive ints."""
    if raw is None:
        return []
    if not isinstance(raw, (list, tuple)):
        raw = [raw]
    seen, out = set(), []
    for r in raw:
        try:
            n = int(r)
        except (ValueError, TypeError):
            continue
        if n > 0 and n not in seen:
            seen.add(n)
            out.append(n)
    return out


@login_required
@admin_required
@require_POST
def movement_batch_add(request):
    """Header-then-rows batch create for stock movements.

    Payload:
        {
          "movement_date": "YYYY-MM-DD",
          "site_id":       <int>,
          "from_person":   "...",   # optional (whole batch)
          "to_person":     "...",   # optional (whole batch)
          "rows": [
            {"material_type_id": N, "serial_ids": [id, ...], "remarks": "..."},
            ...
          ]
        }

    One StockMovement per selected serial. All-or-nothing — a single row
    error rolls back the whole batch.
    """
    admin_id = get_admin_id(request.user)
    data = _body(request)

    # ── Header ──────────────────────────────────────────────────────────
    mdate = _to_date(data.get('movement_date')) or timezone.localdate()
    try:
        from accounts.date_utils import validate_not_future
        validate_not_future(mdate, "Movement date")
    except ValidationError as e:
        return _err('; '.join(e.messages))

    # Source + destination sites are both optional. If provided, they must
    # belong to this tenant.
    source_site = _get_tenant_site(admin_id, data.get('source_site_id'))
    if data.get('source_site_id') and source_site is None:
        return _err('Invalid source site for this account.')
    dest_site = _get_tenant_site(admin_id, data.get('destination_site_id'))
    if data.get('destination_site_id') and dest_site is None:
        return _err('Invalid destination site for this account.')

    from_person = _clean(data.get('from_person'), 150) or None
    to_person   = _clean(data.get('to_person'), 150) or None

    raw_rows = data.get('rows')
    if not isinstance(raw_rows, list) or not raw_rows:
        return _err('Select at least one serial.')
    rows = []
    for idx, r in enumerate(raw_rows):
        if not isinstance(r, dict):
            continue
        sids = _collect_serial_ids(r.get('serial_ids'))
        if not sids:
            continue
        rows.append({
            'index':   idx,
            'sids':    sids,
            'remarks': _clean(r.get('remarks')) or None,
        })
    if not rows:
        return _err('Select at least one serial.')

    # ── Atomic write ────────────────────────────────────────────────────
    saved_ids, per_row_errors = [], {}
    try:
        with transaction.atomic():
            for row in rows:
                for sid in row['sids']:
                    item = (StockItem.objects.select_for_update()
                            .filter(id=sid, admin_id=admin_id).first())
                    if item is None:
                        per_row_errors.setdefault(row['index'], []).append({
                            'serial_id': sid,
                            'error':     'Serial not found for this account.',
                        })
                        raise _BatchError(sid, None,
                                          'Serial not found for this account.')

                    mv = StockMovement.objects.create(
                        admin_id=admin_id, stock_item=item, movement_date=mdate,
                        source_site=source_site, destination_site=dest_site,
                        from_person=from_person, to_person=to_person,
                        remarks=row['remarks'],
                        created_by=request.user if request.user.is_authenticated else None,
                    )

                    # Live pointer on the item. The item's own StockItem.site
                    # (home site) is NOT modified — sites recorded on the
                    # movement are ledger-only.
                    item.current_holder_name = to_person
                    # Prefer destination for the live "current site" pointer;
                    # fall back to source, else clear.
                    item.current_site_name = (
                        dest_site.name if dest_site
                        else (source_site.name if source_site else None)
                    )
                    item.status = StockItem.ISSUED if to_person else StockItem.AVAILABLE
                    item.save(update_fields=['current_holder_name', 'current_site_name',
                                             'status', 'updated_at'])
                    saved_ids.append(mv.id)
    except _BatchError as e:
        return JsonResponse({
            'success':      False,
            'saved_count':  0,
            'error':        f"{e.message} Nothing was saved.",
            'row_errors':   per_row_errors,
        }, status=409)

    return JsonResponse({
        'success':     True,
        'saved_count': len(saved_ids),
        'ids':         saved_ids,
    })


@login_required
@admin_required
@require_POST
def movement_delete(request, pk):
    admin_id = get_admin_id(request.user)

    with transaction.atomic():
        mv = get_object_or_404(
            StockMovement.objects.select_for_update(), id=pk, admin_id=admin_id)
        item = StockItem.objects.select_for_update().get(
            id=mv.stock_item_id, admin_id=admin_id)

        # Prior movement = newest movement strictly older than mv.
        prior = (StockMovement.objects
                 .filter(admin_id=admin_id, stock_item=item)
                 .exclude(pk=mv.pk)
                 .filter(Q(movement_date__lt=mv.movement_date) |
                         (Q(movement_date=mv.movement_date) &
                          Q(created_at__lt=mv.created_at)))
                 .order_by('-movement_date', '-created_at')
                 .first())

        if prior:
            item.current_holder_name = prior.to_person or None
            prior_site = prior.destination_site or prior.source_site
            item.current_site_name   = prior_site.name if prior_site else None
            item.status = StockItem.ISSUED if prior.to_person else StockItem.AVAILABLE
        else:
            # First-ever movement → back to unissued stock.
            item.current_holder_name = None
            item.current_site_name   = None
            item.status = StockItem.AVAILABLE

        item.save(update_fields=['current_holder_name', 'current_site_name',
                                 'status', 'updated_at'])
        mv.delete()

    return JsonResponse({'success': True})


# ===========================================================================
# Autocomplete endpoints
# ===========================================================================

@login_required
@require_GET
def persons_suggest(request):
    admin_id = get_admin_id(request.user)
    q = (request.GET.get('q') or '').strip()

    names = set()
    mv = StockMovement.objects.filter(admin_id=admin_id)
    it = StockItem.objects.filter(admin_id=admin_id)
    if q:
        mv_from = mv.filter(from_person__icontains=q).values_list('from_person', flat=True)
        mv_to   = mv.filter(to_person__icontains=q).values_list('to_person', flat=True)
        holders = it.filter(current_holder_name__icontains=q).values_list('current_holder_name', flat=True)
    else:
        mv_from = mv.values_list('from_person', flat=True)
        mv_to   = mv.values_list('to_person', flat=True)
        holders = it.values_list('current_holder_name', flat=True)
    for src in (mv_from, mv_to, holders):
        for n in src:
            if n and n.strip():
                names.add(n.strip())
    return JsonResponse({'results': sorted(names)[:10]})


@login_required
@require_GET
def sites_suggest(request):
    admin_id = get_admin_id(request.user)
    q = (request.GET.get('q') or '').strip()

    sites = set()
    it = StockItem.objects.filter(admin_id=admin_id)
    if q:
        it_sites = it.filter(current_site_name__icontains=q).values_list('current_site_name', flat=True)
    else:
        it_sites = it.values_list('current_site_name', flat=True)
    for s in it_sites:
        if s and s.strip():
            sites.add(s.strip())
    return JsonResponse({'results': sorted(sites)[:10]})


@login_required
@require_GET
def serials_suggest(request):
    admin_id = get_admin_id(request.user)
    q  = (request.GET.get('q') or '').strip()
    mt = request.GET.get('material_type_id')
    site_id = request.GET.get('site_id')

    # Only AVAILABLE units can be issued out, so the picker never lists
    # serials that are already issued / lost / damaged.
    qs = StockItem.objects.filter(admin_id=admin_id, status=StockItem.AVAILABLE)
    if mt:
        qs = qs.filter(material_type_id=mt)
    if site_id:
        try:
            qs = qs.filter(site_id=int(site_id))
        except (ValueError, TypeError):
            pass
    if q:
        qs = qs.filter(serial_no__icontains=q)
    qs = qs.order_by('serial_no')[:20]

    results = [{
        'id':       it.id,
        'serial':   it.serial_no,
        'status':   it.status,
        'holder':   it.current_holder_name or '',
        'site':     it.current_site_name or '',
    } for it in qs]
    return JsonResponse({'results': results})


@login_required
@require_GET
def items_by_site(request):
    """Available stock, grouped by material type.

    Used by the Stock In/Out header→batch UI. Only AVAILABLE items are
    returned. If `site_id` is given the list is scoped to items whose
    home site matches it OR items with no home site set (unassigned stock
    can be moved from any site). If `site_id` is omitted, ALL available
    items across the tenant are returned.
    """
    admin_id = get_admin_id(request.user)
    site_id  = request.GET.get('site_id')

    qs = (StockItem.objects
          .filter(admin_id=admin_id, status=StockItem.AVAILABLE)
          .select_related('material_type', 'material_type__brand')
          .order_by('material_type__brand__name',
                    'material_type__name', 'serial_no'))

    site = None
    if site_id:
        site = _get_tenant_site(admin_id, site_id)
        if site is None:
            return _err('Invalid site for this account.', status=400)
        qs = qs.filter(Q(site=site) | Q(site__isnull=True))

    groups = {}
    for it in qs:
        mt = it.material_type
        g = groups.get(mt.id)
        if g is None:
            g = groups[mt.id] = {
                'material_type_id': mt.id,
                'brand_id':         mt.brand_id,
                'label':            f"{mt.brand.name} · {mt.name}",
                'items':            [],
            }
        g['items'].append({
            'id':     it.id,
            'serial': it.serial_no,
            'status': it.status,
            'holder': it.current_holder_name or '',
        })

    return JsonResponse({
        'site':   {'id': site.id, 'name': site.name} if site else None,
        'groups': sorted(groups.values(), key=lambda g: g['label']),
    })


@login_required
@admin_required
@require_POST
def quick_add_item(request):
    """Inline stock-item creation from the + Add Serial popover.

    Thin passthrough to `item_add`. Site is optional (matches the top-level
    Master Control add form). If a site_id is supplied it must belong to
    this tenant.
    """
    data = _body(request)
    if data.get('site_id') and not _get_tenant_site(
            get_admin_id(request.user), data.get('site_id')):
        return _err('Invalid site for this account.', status=400)
    return item_add(request)


# ===========================================================================
# LEGACY — retired localStorage-sync endpoint (kept so old callers don't 500)
# ===========================================================================

def _site_to_dict(s):
    return {'id': s.client_id, 'name': s.name}


def _item_to_dict(i):
    return {
        'id':        i.client_id,
        'siteId':    i.site_client_id or '',
        'itemName':  i.item_name,
        'brand':     i.brand,
        'length':    i.length,
        'category':  i.category,
        'qty':       float(i.qty or 0),
        'condition': i.condition,
    }


def _entry_to_dict(e):
    return {
        'id':              e.client_id,
        'createdAt':       e.created_at.isoformat() if e.created_at else '',
        'updatedAt':       e.updated_at.isoformat() if e.updated_at else '',
        'date':            e.date.isoformat() if e.date else '',
        'fromSite':        e.from_site,
        'toSite':          e.to_site,
        'itemName':        e.item_name,
        'serialNumber':    e.serial_number,
        'typeBrand':       e.type_brand,
        'length':          e.length,
        'quantityOut':     float(e.quantity_out or 0),
        'quantityIn':      float(e.quantity_in or 0),
        'remarks':         e.remarks,
        'workerName':      e.worker_name,
        'dateTime':        e.date_time.isoformat() if e.date_time else '',
        'missingStock':    float(e.missing_stock or 0),
        'damageDeduction': float(e.damage_deduction or 0),
        'approvalBy':      e.approval_by,
    }


@login_required
@require_http_methods(['GET', 'POST'])
def state(request):
    admin_id = get_admin_id(request.user)

    if request.method == 'GET':
        return JsonResponse({
            'success': True,
            'sites':   [_site_to_dict(s)  for s in LegacyStockSite.objects.filter(admin_id=admin_id)],
            'items':   [_item_to_dict(i)  for i in LegacyStockItem.objects.filter(admin_id=admin_id)],
            'entries': [_entry_to_dict(e) for e in LegacyStockEntry.objects.filter(admin_id=admin_id)],
        })

    try:
        payload = json.loads(request.body or b'{}')
    except (ValueError, TypeError):
        return _err('Invalid JSON.')

    sites_in   = payload.get('sites')   if 'sites'   in payload else None
    items_in   = payload.get('items')   if 'items'   in payload else None
    entries_in = payload.get('entries') if 'entries' in payload else None

    with transaction.atomic():
        if isinstance(sites_in, list):
            _replace_sites(admin_id, request.user, sites_in)
        if isinstance(items_in, list):
            _replace_items(admin_id, request.user, items_in)
        if isinstance(entries_in, list):
            _replace_entries(admin_id, request.user, entries_in)

    return JsonResponse({
        'success': True,
        'sites':   [_site_to_dict(s)  for s in LegacyStockSite.objects.filter(admin_id=admin_id)],
        'items':   [_item_to_dict(i)  for i in LegacyStockItem.objects.filter(admin_id=admin_id)],
        'entries': [_entry_to_dict(e) for e in LegacyStockEntry.objects.filter(admin_id=admin_id)],
    })


def _replace_sites(admin_id, user, payload_list):
    incoming_ids = set()
    for raw in payload_list:
        if not isinstance(raw, dict):
            continue
        cid = (raw.get('id') or '').strip()
        if not cid:
            continue
        incoming_ids.add(cid)
        LegacyStockSite.objects.update_or_create(
            admin_id=admin_id, client_id=cid,
            defaults={
                'name':       (raw.get('name') or '').strip()[:200],
                'created_by': user if user.is_authenticated else None,
            },
        )
    LegacyStockSite.objects.filter(admin_id=admin_id).exclude(client_id__in=incoming_ids).delete()


def _replace_items(admin_id, user, payload_list):
    incoming_ids = set()
    for raw in payload_list:
        if not isinstance(raw, dict):
            continue
        cid = (raw.get('id') or '').strip()
        if not cid:
            continue
        incoming_ids.add(cid)
        LegacyStockItem.objects.update_or_create(
            admin_id=admin_id, client_id=cid,
            defaults={
                'site_client_id': (raw.get('siteId') or '').strip()[:80],
                'item_name':      (raw.get('itemName') or '').strip()[:150],
                'brand':          (raw.get('brand') or '').strip()[:150],
                'length':         (raw.get('length') or '').strip()[:50],
                'category':       (raw.get('category') or 'General').strip()[:100],
                'qty':            _to_decimal(raw.get('qty')),
                'condition':      (raw.get('condition') or 'Good').strip()[:50],
                'created_by':     user if user.is_authenticated else None,
            },
        )
    LegacyStockItem.objects.filter(admin_id=admin_id).exclude(client_id__in=incoming_ids).delete()


def _replace_entries(admin_id, user, payload_list):
    incoming_ids = set()
    for raw in payload_list:
        if not isinstance(raw, dict):
            continue
        cid = (raw.get('id') or '').strip()
        if not cid:
            continue
        incoming_ids.add(cid)
        LegacyStockEntry.objects.update_or_create(
            admin_id=admin_id, client_id=cid,
            defaults={
                'date':             _to_date(raw.get('date')),
                'from_site':        (raw.get('fromSite') or '').strip()[:200],
                'to_site':          (raw.get('toSite')   or '').strip()[:200],
                'item_name':        (raw.get('itemName') or '').strip()[:150],
                'serial_number':    (raw.get('serialNumber') or '').strip()[:100],
                'type_brand':       (raw.get('typeBrand') or '').strip()[:150],
                'length':           (raw.get('length') or '').strip()[:50],
                'quantity_out':     _to_decimal(raw.get('quantityOut')),
                'quantity_in':      _to_decimal(raw.get('quantityIn')),
                'remarks':          (raw.get('remarks') or ''),
                'worker_name':      (raw.get('workerName') or '').strip()[:150],
                'date_time':        _to_datetime(raw.get('dateTime')),
                'missing_stock':    _to_decimal(raw.get('missingStock')),
                'damage_deduction': _to_decimal(raw.get('damageDeduction')),
                'approval_by':      (raw.get('approvalBy') or '').strip()[:150],
                'created_by':       user if user.is_authenticated else None,
            },
        )
    LegacyStockEntry.objects.filter(admin_id=admin_id).exclude(client_id__in=incoming_ids).delete()
