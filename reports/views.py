from io import BytesIO
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Sum
from django.utils import timezone
from finance.models import Transaction, Category
from income.models import Income
from branches.models import Branch
from accounts.views import get_admin_id
from datetime import date, timedelta, datetime as dt
from dateutil.relativedelta import relativedelta
from calendar import month_abbr
import json


# ─── querysets ───────────────────────────────────────────────────────────────

def _base_qs(user):
    admin_id = get_admin_id(user)
    if user.is_admin:
        return (
            Transaction.objects.filter(admin_id=admin_id, type='expense'),
            Income.objects.filter(admin_id=admin_id),
        )
    return (
        Transaction.objects.filter(user=user, type='expense'),
        Income.objects.filter(user=user),
    )


def _apply_filters(txn_qs, inc_qs, branch_id, cat_param, min_amt, max_amt):
    """Apply filters. cat_param: 'f_<id>' finance.Category, 'i_<id>' IncomeCategory, 'e_<id>' ExpenseCategory."""
    if branch_id:
        try:
            txn_qs = txn_qs.filter(branch_id=int(branch_id))
        except (ValueError, TypeError):
            pass

    if cat_param:
        if cat_param.startswith('f_'):
            try:
                txn_qs = txn_qs.filter(category_id=int(cat_param[2:]))
            except (ValueError, TypeError):
                pass
        elif cat_param.startswith('i_'):
            try:
                inc_qs = inc_qs.filter(category_id=int(cat_param[2:]))
            except (ValueError, TypeError):
                pass
        elif cat_param.startswith('e_'):
            try:
                from categories.models import ExpenseCategory as ExpCat
                exp_cat = ExpCat.objects.get(pk=int(cat_param[2:]))
                txn_qs = txn_qs.filter(category__name__iexact=exp_cat.name)
            except Exception:
                pass
        else:
            try:
                txn_qs = txn_qs.filter(category_id=int(cat_param))
            except (ValueError, TypeError):
                pass

    if min_amt:
        try:
            v = float(min_amt)
            txn_qs = txn_qs.filter(amount__gte=v)
            inc_qs = inc_qs.filter(amount__gte=v)
        except (ValueError, TypeError):
            pass
    if max_amt:
        try:
            v = float(max_amt)
            txn_qs = txn_qs.filter(amount__lte=v)
            inc_qs = inc_qs.filter(amount__lte=v)
        except (ValueError, TypeError):
            pass

    return txn_qs, inc_qs


# ─── aggregation helpers ──────────────────────────────────────────────────────

def _sum_range(qs, start, end):
    return float(qs.filter(date__range=[start, end]).aggregate(s=Sum('amount'))['s'] or 0)


def _period(txn_qs, inc_qs, start, end):
    income  = _sum_range(inc_qs, start, end)
    expense = _sum_range(txn_qs, start, end)
    t_count = (
        txn_qs.filter(date__range=[start, end]).count() +
        inc_qs.filter(date__range=[start, end]).count()
    )
    net   = income - expense
    total = income + expense
    pct   = round((income / total * 100) if total else 0, 1)
    return {
        'income': income, 'expense': expense, 'net': net,
        'transactions': t_count, 'income_pct': pct,
        'start': start.isoformat(), 'end': end.isoformat(),
    }


def _monthly_rows(txn_qs, inc_qs, year):
    rows = []
    for m in range(1, 13):
        ms = date(year, m, 1)
        me = (ms + relativedelta(months=1)) - timedelta(days=1)
        d  = _period(txn_qs, inc_qs, ms, me)
        d['label'] = month_abbr[m] + f' {year}'
        rows.append(d)
    return rows


def _daily_rows(txn_qs, inc_qs, start, end):
    rows, cur = [], start
    while cur <= end:
        d = _period(txn_qs, inc_qs, cur, cur)
        d['label'] = cur.strftime('%a, %b %d')
        rows.append(d)
        cur += timedelta(days=1)
    return rows


def _weekly_rows_for_month(txn_qs, inc_qs, month_start, month_end):
    rows, ws, wn = [], month_start, 1
    while ws <= month_end:
        we = min(ws + timedelta(days=6), month_end)
        d  = _period(txn_qs, inc_qs, ws, we)
        d['label'] = f"Week {wn} ({ws.strftime('%b %d')}–{we.strftime('%d')})"
        rows.append(d)
        ws  = we + timedelta(days=1)
        wn += 1
    return rows


# ─── period boundaries ────────────────────────────────────────────────────────

def _period_boundaries(today):
    week_start  = today - timedelta(days=today.weekday())
    week_end    = week_start + timedelta(days=6)
    month_start = today.replace(day=1)
    month_end   = (month_start + relativedelta(months=1)) - timedelta(days=1)
    q_num       = (today.month - 1) // 3
    q_mo        = q_num * 3 + 1
    q_start     = date(today.year, q_mo, 1)
    q_end       = (q_start + relativedelta(months=3)) - timedelta(days=1)
    h_start     = date(today.year, 1, 1)  if today.month <= 6 else date(today.year, 7, 1)
    h_end       = date(today.year, 6, 30) if today.month <= 6 else date(today.year, 12, 31)
    year_start  = date(today.year, 1, 1)
    year_end    = date(today.year, 12, 31)
    return {
        'week_start': week_start, 'week_end': week_end,
        'month_start': month_start, 'month_end': month_end,
        'q_num': q_num, 'q_start': q_start, 'q_end': q_end,
        'h_start': h_start, 'h_end': h_end,
        'year_start': year_start, 'year_end': year_end,
    }


# ─── global categories ────────────────────────────────────────────────────────

def _all_categories(user):
    """Return a flat, alphabetically-sorted, deduplicated list of categories.

    Finance categories are preferred (directly linked to Transaction.category).
    Module categories fill in names not already covered by Finance.
    No groups / optgroups — clean names only.
    """
    from categories.models import IncomeCategory as IncomeCatModel, ExpenseCategory as ExpCatModel
    admin_id = get_admin_id(user)
    seen = {}   # lowercase name → {'id': str, 'name': str}

    for c in Category.objects.filter(admin_id=admin_id, type='expense').order_by('name'):
        seen[c.name.lower()] = {'id': f'f_{c.pk}', 'name': c.name}

    for c in Category.objects.filter(admin_id=admin_id, type='income').order_by('name'):
        key = c.name.lower()
        if key not in seen:
            seen[key] = {'id': f'f_{c.pk}', 'name': c.name}

    for c in IncomeCatModel.objects.filter(created_by__admin_id=admin_id).order_by('name'):
        key = c.name.lower()
        if key not in seen:
            seen[key] = {'id': f'i_{c.pk}', 'name': c.name}

    for c in ExpCatModel.objects.filter(created_by__admin_id=admin_id).order_by('name'):
        key = c.name.lower()
        if key not in seen:
            seen[key] = {'id': f'e_{c.pk}', 'name': c.name}

    return sorted(seen.values(), key=lambda c: c['name'].lower())


# ─── download generators ─────────────────────────────────────────────────────

def _generate_xlsx(rows, period_label, today):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Financial Report'

    ws['A1'] = 'SPIM Suite – Financial Report'
    ws['A1'].font = Font(bold=True, size=14, color='0F172A')
    ws['A2'] = period_label
    ws['A2'].font = Font(size=11, color='475569')
    ws['A3'] = f'Generated: {today.strftime("%B %d, %Y")}'
    ws['A3'].font = Font(size=9, color='94A3B8')
    ws.append([])

    headers = ['Period', 'Income (₹)', 'Expenses (₹)', 'Net Balance (₹)', 'Transactions']
    ws.append(headers)
    hdr_row  = ws.max_row
    hdr_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
    hdr_font = Font(color='FFFFFF', bold=True, size=10)
    thin     = Side(style='thin', color='E2E8F0')
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, 6):
        cell = ws.cell(row=hdr_row, column=col)
        cell.fill, cell.font = hdr_fill, hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    ws.row_dimensions[hdr_row].height = 22

    alt_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    num_fmt  = '#,##0.00'
    for ri, r in enumerate(rows):
        ws.append([r['label'], round(r['income'],2), round(r['expense'],2), round(r['net'],2), r['transactions']])
        rn = ws.max_row
        fill = alt_fill if ri % 2 == 1 else None
        for col in range(1, 6):
            cell = ws.cell(row=rn, column=col)
            cell.border = border
            if fill:
                cell.fill = fill
        ws.cell(row=rn, column=1).alignment = Alignment(horizontal='left')
        for col in range(2, 6):
            ws.cell(row=rn, column=col).alignment = Alignment(horizontal='right')
        ws.cell(row=rn, column=2).font        = Font(color='059669', bold=True)
        ws.cell(row=rn, column=2).number_format = num_fmt
        ws.cell(row=rn, column=3).font        = Font(color='DC2626', bold=True)
        ws.cell(row=rn, column=3).number_format = num_fmt
        nc = '059669' if r['net'] >= 0 else 'DC2626'
        ws.cell(row=rn, column=4).font        = Font(color=nc, bold=True)
        ws.cell(row=rn, column=4).number_format = num_fmt

    if rows:
        ti = sum(r['income'] for r in rows)
        te = sum(r['expense'] for r in rows)
        tn = sum(r['net'] for r in rows)
        tt = sum(r['transactions'] for r in rows)
        ws.append(['TOTAL', round(ti,2), round(te,2), round(tn,2), tt])
        tr = ws.max_row
        tf = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
        for col in range(1, 6):
            cell = ws.cell(row=tr, column=col)
            cell.fill, cell.font, cell.border = tf, Font(bold=True), border
            cell.alignment = Alignment(horizontal='right' if col > 1 else 'left')
        ws.cell(row=tr, column=2).font = Font(color='059669', bold=True)
        ws.cell(row=tr, column=2).number_format = num_fmt
        ws.cell(row=tr, column=3).font = Font(color='DC2626', bold=True)
        ws.cell(row=tr, column=3).number_format = num_fmt
        nc = '059669' if tn >= 0 else 'DC2626'
        ws.cell(row=tr, column=4).font = Font(color=nc, bold=True)
        ws.cell(row=tr, column=4).number_format = num_fmt

    for i, w in enumerate([26, 16, 16, 18, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _generate_pdf(rows, period_label, today):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    story = []

    story.append(Paragraph('SPIM Suite – Financial Report', ParagraphStyle(
        'T', fontSize=16, fontName='Helvetica-Bold',
        textColor=colors.HexColor('#0f172a'), spaceAfter=4)))
    story.append(Paragraph(period_label, ParagraphStyle(
        'S', fontSize=11, fontName='Helvetica',
        textColor=colors.HexColor('#475569'), spaceAfter=2)))
    story.append(Paragraph(f"Generated: {today.strftime('%B %d, %Y')}", ParagraphStyle(
        'D', fontSize=9, fontName='Helvetica',
        textColor=colors.HexColor('#94a3b8'), spaceAfter=12)))

    if not rows:
        story.append(Paragraph('No data available for this period.', getSampleStyleSheet()['Normal']))
    else:
        header = ['Period', 'Income (₹)', 'Expenses (₹)', 'Net Balance (₹)', 'Txns']
        data = [header]
        for r in rows:
            net_str = f"₹{r['net']:.2f}" if r['net'] >= 0 else f"-₹{abs(r['net']):.2f}"
            data.append([r['label'], f"₹{r['income']:.2f}", f"₹{r['expense']:.2f}", net_str, str(r['transactions'])])

        ti = sum(r['income'] for r in rows)
        te = sum(r['expense'] for r in rows)
        tn = sum(r['net'] for r in rows)
        tt = sum(r['transactions'] for r in rows)
        data.append(['TOTAL', f"₹{ti:.2f}", f"₹{te:.2f}",
                     f"₹{tn:.2f}" if tn >= 0 else f"-₹{abs(tn):.2f}", str(tt)])

        t = Table(data, colWidths=[6*cm, 4.5*cm, 4.5*cm, 5*cm, 2.5*cm], repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND',  (0,0), (-1,0),  colors.HexColor('#0f172a')),
            ('TEXTCOLOR',   (0,0), (-1,0),  colors.white),
            ('FONTNAME',    (0,0), (-1,0),  'Helvetica-Bold'),
            ('FONTSIZE',    (0,0), (-1,0),  10),
            ('TOPPADDING',  (0,0), (-1,0),  8),
            ('BOTTOMPADDING',(0,0),(-1,0),  8),
            ('FONTNAME',    (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE',    (0,1), (-1,-1), 9),
            ('ROWBACKGROUNDS',(0,1),(-1,-2),[colors.white, colors.HexColor('#f8fafc')]),
            ('TOPPADDING',  (0,1), (-1,-1), 6),
            ('BOTTOMPADDING',(0,1),(-1,-1), 6),
            ('BACKGROUND',  (0,-1),(-1,-1), colors.HexColor('#f1f5f9')),
            ('FONTNAME',    (0,-1),(-1,-1), 'Helvetica-Bold'),
            ('ALIGN',       (1,0), (-1,-1), 'RIGHT'),
            ('ALIGN',       (0,0), (0,-1),  'LEFT'),
            ('TEXTCOLOR',   (1,1), (1,-2),  colors.HexColor('#059669')),
            ('TEXTCOLOR',   (2,1), (2,-2),  colors.HexColor('#dc2626')),
            ('GRID',        (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ]))
        story.append(t)

    doc.build(story)
    return buf.getvalue()


def _get_download_rows(txn_qs, inc_qs, period, today, date_from='', date_to=''):
    """Return (rows, label) for the given period key."""
    b = _period_boundaries(today)

    if period == 'daily':
        return ([dict({'label': today.strftime('%A, %B %d, %Y')}, **_period(txn_qs, inc_qs, today, today))],
                f"Daily Report – {today.strftime('%B %d, %Y')}")

    if period == 'weekly':
        return (_daily_rows(txn_qs, inc_qs, b['week_start'], b['week_end']),
                'Weekly Report – This Week')

    if period == 'quarterly':
        all_m = _monthly_rows(txn_qs, inc_qs, today.year)
        q = b['q_num']
        return (all_m[q*3:(q+1)*3], f"Q{q+1} {today.year} Quarterly Report")

    if period == 'half_yearly':
        all_m = _monthly_rows(txn_qs, inc_qs, today.year)
        half  = all_m[:6] if today.month <= 6 else all_m[6:]
        label = f"{'H1' if today.month <= 6 else 'H2'} {today.year} Half-Yearly Report"
        return (half, label)

    if period == 'annual':
        return (_monthly_rows(txn_qs, inc_qs, today.year), f"Annual Report – {today.year}")

    if period == 'custom' and date_from and date_to:
        try:
            cs = dt.strptime(date_from, '%Y-%m-%d').date()
            ce = dt.strptime(date_to,   '%Y-%m-%d').date()
            if cs <= ce:
                delta = (ce - cs).days
                rows = (_daily_rows(txn_qs, inc_qs, cs, ce) if delta <= 31
                        else _monthly_rows(txn_qs, inc_qs, cs.year))
                return (rows, f"Custom Report {cs.strftime('%b %d')} – {ce.strftime('%b %d, %Y')}")
        except ValueError:
            pass

    # Default: monthly
    rows = _weekly_rows_for_month(txn_qs, inc_qs, b['month_start'], b['month_end'])
    return (rows, f"Monthly Report – {today.strftime('%B %Y')}")


# ─── views ───────────────────────────────────────────────────────────────────

@login_required
def reports_index(request):
    user  = request.user
    today = timezone.now().date()

    # Task 1 + 4: period defaults to 'monthly' — shows only that card unless changed
    period    = request.GET.get('period', 'monthly')
    branch_id = request.GET.get('branch', '')
    cat_param = request.GET.get('category', '')
    min_amt   = request.GET.get('min_amount', '')
    max_amt   = request.GET.get('max_amount', '')
    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')

    txn_qs, inc_qs = _base_qs(user)
    txn_qs, inc_qs = _apply_filters(txn_qs, inc_qs, branch_id, cat_param, min_amt, max_amt)

    b = _period_boundaries(today)

    # ── Performance fix (Task 3) ──────────────────────────────────────
    # Previously every render computed all 6 cards (~100+ aggregate queries)
    # only to discard 5 of them in the SHOW_MAP step below. We now derive
    # `show_keys` first and build only the cards that will actually render.
    # Quarterly / half-yearly / annual cards reuse one cached `monthly_rows`
    # so we never query the same month range twice.
    SHOW_MAP = {
        'daily':       ['daily'],
        'weekly':      ['weekly'],
        'monthly':     ['monthly'],
        'quarterly':   ['quarterly'],
        'half_yearly': ['half_yearly'],
        'annual':      ['annual'],
        'custom':      ['custom'],
        'all':         ['daily','weekly','monthly','quarterly','half_yearly','annual','custom'],
    }
    show_keys = SHOW_MAP.get(period, ['monthly'])

    # Lazy memo: monthly_rows is shared by quarterly / half / annual cards.
    _monthly_cache = {}
    def _get_monthly():
        if 'rows' not in _monthly_cache:
            _monthly_cache['rows'] = _monthly_rows(txn_qs, inc_qs, today.year)
        return _monthly_cache['rows']

    # Build a card only when its key is in show_keys.
    builders = {
        'daily': lambda: {
            'key':      'daily',
            'label':    'Daily Report',
            'sublabel': f"Today – {today.strftime('%A, %B %d, %Y')}",
            'detail':   [dict({'label': today.strftime('%A, %B %d, %Y')},
                              **_period(txn_qs, inc_qs, today, today))],
            **_period(txn_qs, inc_qs, today, today),
        },
        'weekly': lambda: {
            'key':      'weekly',
            'label':    'Weekly Report',
            'sublabel': 'This Week',
            'detail':   _daily_rows(txn_qs, inc_qs, b['week_start'], b['week_end']),
            **_period(txn_qs, inc_qs, b['week_start'], b['week_end']),
        },
        'monthly': lambda: {
            'key':      'monthly',
            'label':    'Monthly Report',
            'sublabel': today.strftime('%b %Y'),
            'detail':   _weekly_rows_for_month(txn_qs, inc_qs, b['month_start'], b['month_end']),
            **_period(txn_qs, inc_qs, b['month_start'], b['month_end']),
        },
        'quarterly': lambda: {
            'key':      'quarterly',
            'label':    'Quarterly Report',
            'sublabel': f"Q{b['q_num'] + 1} {today.year}",
            'detail':   _get_monthly()[b['q_num']*3:(b['q_num']+1)*3],
            **_period(txn_qs, inc_qs, b['q_start'], b['q_end']),
        },
        'half_yearly': lambda: {
            'key':      'half_yearly',
            'label':    'Half-Yearly Report',
            'sublabel': f"{'H1' if today.month <= 6 else 'H2'} {today.year}",
            'detail':   _get_monthly()[:6] if today.month <= 6 else _get_monthly()[6:],
            **_period(txn_qs, inc_qs, b['h_start'], b['h_end']),
        },
        'annual': lambda: {
            'key':      'annual',
            'label':    'Annual Report',
            'sublabel': f"Year {today.year}",
            'detail':   _get_monthly(),
            **_period(txn_qs, inc_qs, b['year_start'], b['year_end']),
        },
    }

    display_periods = [builders[k]() for k in show_keys if k in builders]

    # Custom date range — built only if explicitly requested via show_keys.
    if 'custom' in show_keys and date_from and date_to:
        try:
            cs = dt.strptime(date_from, '%Y-%m-%d').date()
            ce = dt.strptime(date_to,   '%Y-%m-%d').date()
            if cs <= ce:
                delta  = (ce - cs).days
                detail = (_daily_rows(txn_qs, inc_qs, cs, ce) if delta <= 31
                          else _monthly_rows(txn_qs, inc_qs, cs.year))
                display_periods.append({
                    'key':      'custom',
                    'label':    'Custom Report',
                    'sublabel': f"{cs.strftime('%b %d')} – {ce.strftime('%b %d, %Y')}",
                    'detail':   detail,
                    **_period(txn_qs, inc_qs, cs, ce),
                })
        except ValueError:
            pass

    if not display_periods:
        display_periods = [builders['monthly']()]

    admin_id   = get_admin_id(user)
    branches   = Branch.objects.filter(admin_id=admin_id)

    return render(request, 'reports/index.html', {
        'periods':      display_periods,
        'periods_json': json.dumps(display_periods),
        'branches':     branches,
        'all_categories': _all_categories(user),  # Task 3: global categories
        'filters': {
            'period':     period,
            'branch':     branch_id,
            'category':   cat_param,
            'min_amount': min_amt,
            'max_amount': max_amt,
            'date_from':  date_from,
            'date_to':    date_to,
        },
        'today': today,
    })


@login_required
def reports_download(request):
    user   = request.user
    fmt    = request.GET.get('format', 'xlsx')
    period = request.GET.get('period', 'monthly')
    today  = timezone.now().date()

    txn_qs, inc_qs = _base_qs(user)
    txn_qs, inc_qs = _apply_filters(
        txn_qs, inc_qs,
        request.GET.get('branch', ''),
        request.GET.get('category', ''),
        request.GET.get('min_amount', ''),
        request.GET.get('max_amount', ''),
    )

    rows, period_label = _get_download_rows(
        txn_qs, inc_qs, period, today,
        request.GET.get('date_from', ''),
        request.GET.get('date_to', ''),
    )
    fname = f"financial-report-{period}-{today.strftime('%Y%m%d')}"

    # Task 6: real XLSX
    if fmt == 'xlsx':
        content = _generate_xlsx(rows, period_label, today)
        resp = HttpResponse(content, content_type=(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))
        resp['Content-Disposition'] = f'attachment; filename="{fname}.xlsx"'
        return resp

    # Task 5: real PDF via reportlab
    if fmt == 'pdf':
        content = _generate_pdf(rows, period_label, today)
        resp = HttpResponse(content, content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="{fname}.pdf"'
        return resp

    return HttpResponse('Unsupported format', status=400)
