from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.http import JsonResponse
from django.contrib import messages
from .models import Project, Task, MachineLocation, WorkLog, WorkStatus
from .forms import ProjectForm, TaskForm
from accounts.views import get_admin_id
from employees.models import Employee
from branches.models import LocationSite
from datetime import date as date_type
import json

# Statuses that allow an employee's machine work to surface in reports.
# Matches the mobile rule in SPIM Lite (dashboard / machines / attendance
# screens): present + half_day are working; everything else hides machine
# display without deleting the underlying worklog row.
_WORKING_ATTENDANCE_STATUSES = ('present', 'half_day')


def _attendance_qualified_emps(emps, on_date):
    """
    Filter `emps` (iterable of Employee) to those whose AttendanceRecord
    for `on_date` is in {'present', 'half_day'}, OR who have no attendance
    record for that date at all (permissive default — matches mobile).

    Single batched query per call, so cost stays O(1) per worklog row.
    Returns a list in the original order.
    """
    from attendance.models import AttendanceRecord  # deferred: circular safety
    emp_list = list(emps)
    if not emp_list:
        return emp_list
    att_map = dict(
        AttendanceRecord.objects
        .filter(employee_id__in=[e.pk for e in emp_list], date=on_date)
        .values_list('employee_id', 'status')
    )
    return [
        e for e in emp_list
        if att_map.get(e.pk) is None
        or att_map[e.pk] in _WORKING_ATTENDANCE_STATUSES
    ]

# ─── Tenant-scoping helpers (Project / Task) ────────────────────────────────
# Project and Task both have an admin_id column on the DB (added by
# migration 0002). Admins see every project/task in their own org;
# non-admin users continue to see only the projects they personally own.
def _tenant_project_qs(user):
    if user.is_admin:
        return Project.objects.filter(admin_id=get_admin_id(user))
    return Project.objects.filter(owner=user, admin_id=get_admin_id(user))

def _tenant_task_qs(user):
    if user.is_admin:
        return Task.objects.filter(admin_id=get_admin_id(user))
    return Task.objects.filter(project__owner=user, admin_id=get_admin_id(user))


@login_required
def project_list(request):
    qs = _tenant_project_qs(request.user)
    return render(request, 'projects/list.html', {'projects': qs.select_related('client','owner')})

@login_required
def project_detail(request, pk):
    qs = _tenant_project_qs(request.user)
    p  = get_object_or_404(qs, pk=pk)
    return render(request, 'projects/detail.html', {
        'project': p, 'tasks': p.tasks.select_related('assigned_to'), 'task_form': TaskForm()
    })

@login_required
def add_project(request):
    form = ProjectForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        p = form.save(commit=False)
        p.owner    = request.user
        # Stamp the tenant so the row is isolated for every subsequent
        # admin_id-scoped query (list / detail / edit / delete / tasks).
        p.admin_id = get_admin_id(request.user)
        p.save()
        messages.success(request, "Project created.")
        return redirect('projects:detail', pk=p.pk)
    return render(request, 'projects/form.html', {'form': form, 'title': 'New Project'})

@login_required
def edit_project(request, pk):
    qs = _tenant_project_qs(request.user)
    p  = get_object_or_404(qs, pk=pk)
    form = ProjectForm(request.POST or None, instance=p)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Project updated.")
        return redirect('projects:detail', pk=p.pk)
    return render(request, 'projects/form.html', {'form': form, 'title': 'Edit Project', 'obj': p})

@login_required
def delete_project(request, pk):
    qs = _tenant_project_qs(request.user)
    p  = get_object_or_404(qs, pk=pk)
    if request.method == 'POST':
        p.delete()
        messages.success(request, "Deleted.")
    return redirect('projects:list')

@login_required
def add_task(request, project_pk):
    qs      = _tenant_project_qs(request.user)
    project = get_object_or_404(qs, pk=project_pk)
    form    = TaskForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        t = form.save(commit=False)
        t.project  = project
        # Inherit tenancy from the parent project so list / status / delete
        # queries scoped to admin_id pick this task up.
        t.admin_id = project.admin_id
        t.save()
        messages.success(request, "Task added.")
    return redirect('projects:detail', pk=project_pk)

@login_required
def update_task_status(request, pk):
    task_qs = _tenant_task_qs(request.user)
    task    = get_object_or_404(task_qs, pk=pk)
    ns   = request.POST.get('status')
    if ns in dict(Task.STATUS):
        task.status = ns
        task.save()
    return redirect('projects:detail', pk=task.project_id)

@login_required
def delete_task(request, pk):
    task_qs = _tenant_task_qs(request.user)
    task    = get_object_or_404(task_qs, pk=pk)
    pid  = task.project_id
    if request.method == 'POST':
        task.delete()
        messages.success(request, "Task removed.")
    return redirect('projects:detail', pk=pid)


# ── Work Log Views ──────────────────────────────────────────────────────────

@login_required
def work_log_dashboard(request):
    """
    Main entry point for /projects/ — Daily Work Log & Work Summary.
    AJAX JSON requests (format=json) return filtered data.
    Normal GET requests render the two-tab dashboard template.
    """
    admin_id = get_admin_id(request.user)

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.GET.get('format') == 'json':
        tab = request.GET.get('tab', 'log')
        if tab == 'summary':
            return _work_summary_json(request, admin_id)
        return _work_log_json(request, admin_id)

    machines     = list(MachineLocation.objects.filter(admin_id=admin_id).values('id', 'name'))
    employees    = list(Employee.objects.filter(admin_id=admin_id, status='active').values('id', 'employee_id', 'name'))
    statuses     = list(WorkStatus.objects.filter(admin_id=admin_id).values('id', 'name'))
    global_sites = list(LocationSite.objects.filter(admin_id=admin_id).values_list('name', flat=True))

    return render(request, 'projects/work_log.html', {
        'machines_json':     json.dumps(machines),
        'employees_json':    json.dumps(employees),
        'statuses_json':     json.dumps(statuses),
        'global_sites_json': json.dumps(global_sites),
    })


def _work_log_json(request, admin_id):
    """Return filtered WorkLog entries as JSON for the Daily Work Log tab."""
    qs = WorkLog.objects.filter(admin_id=admin_id).select_related('location').prefetch_related('employees')

    f_date     = request.GET.get('date', '').strip()
    f_location = request.GET.get('location', '').strip()
    f_site     = request.GET.get('site', '').strip()
    f_remarks  = request.GET.get('remarks', '').strip()

    if f_date:
        qs = qs.filter(date=f_date)
    if f_location:
        qs = qs.filter(location_id=f_location)
    if f_site:
        qs = qs.filter(site__icontains=f_site)
    if f_remarks:
        qs = qs.filter(remarks__icontains=f_remarks)

    # Daily Work Log renders exactly one row per MachineLocation. If the
    # DB still holds duplicate WorkLog rows for the same (admin_id, date,
    # location) — possible before migration 0007 runs — the client's
    # `logMap[e.location_id] = e` collapse would silently keep the LAST
    # iterated row (oldest, per Meta.ordering), which is typically the
    # empty stub from a legacy work_log_add / mobile race. Collapse here
    # instead so the JSON carries at most one canonical entry per
    # location: the FIRST hit under Meta.ordering `['-date','-created_at']`
    # — i.e. the newest, which is the row Work Summary displays. Rows
    # whose location was SET_NULL'd (location_id IS NULL) are omitted
    # because they cannot be attached to any machine row on this tab.
    entries_by_loc = {}
    for log in qs:
        loc_key = log.location_id
        if not loc_key:
            continue
        if loc_key in entries_by_loc:
            continue
        emps           = list(log.employees.all())
        emp_ids        = [e.pk for e in emps]
        emp_names      = [e.name for e in emps]
        qualified_emps = _attendance_qualified_emps(emps, log.date)
        entries_by_loc[loc_key] = {
            'id':            log.pk,
            'date':          log.date.strftime('%d-%m-%Y'),
            'date_iso':      log.date.isoformat(),
            'location_id':   loc_key,
            'location_name': log.location.name if log.location else '—',
            'site':          log.site or '',
            'work_details':  log.work_details or '',
            'tmp':           len(qualified_emps),
            'remarks':       log.remarks or '',
            'employee_ids':  emp_ids,
            'employee_names': emp_names,
            'locked':        bool(getattr(log, 'locked', False)),
        }
    entries = [{'sl': i, **e} for i, e in enumerate(entries_by_loc.values(), start=1)]

    locations = list(MachineLocation.objects.filter(admin_id=admin_id).values('id', 'name'))
    statuses  = list(WorkStatus.objects.filter(admin_id=admin_id).values('id', 'name'))
    employees = list(Employee.objects.filter(admin_id=admin_id, status='active').values('id', 'employee_id', 'name'))
    sites = list(
        WorkLog.objects.filter(admin_id=admin_id)
        .values_list('site', flat=True)
        .exclude(site='').distinct().order_by('site')
    )

    return JsonResponse({
        'entries':   entries,
        'locations': locations,
        'sites':     sites,
        'statuses':  statuses,
        'employees': employees,
    })


def _work_summary_json(request, admin_id):
    """Return filtered summary data as JSON for the Work Summary tab."""
    qs = WorkLog.objects.filter(admin_id=admin_id).select_related('location').prefetch_related('employees')

    f_date_from = request.GET.get('date_from', '').strip()
    f_date_to   = request.GET.get('date_to', '').strip()
    f_site      = request.GET.get('site', '').strip()
    f_location  = request.GET.get('location', '').strip()
    f_remarks   = request.GET.get('remarks', '').strip()

    if f_date_from:
        qs = qs.filter(date__gte=f_date_from)
    if f_date_to:
        qs = qs.filter(date__lte=f_date_to)
    if f_site:
        qs = qs.filter(site__icontains=f_site)
    if f_location:
        qs = qs.filter(location_id=f_location)
    if f_remarks:
        qs = qs.filter(remarks__icontains=f_remarks)

    rows = []
    total_tmp = 0
    for idx, log in enumerate(qs, start=1):
        # TMP = actual linked-employee count for this row, not the legacy
        # stored `tmp` integer (which mobile inserts default to 0).
        # Uses the prefetch_related cache — no extra query per row.
        # Then drop any employee whose attendance for this date is not
        # 'present' / 'half_day' so reports reflect the same business
        # rule the mobile UI uses for the Today's Machine card.
        emps = _attendance_qualified_emps(log.employees.all(), log.date)
        emp_count = len(emps)
        total_tmp += emp_count
        emp_names = ', '.join(e.name for e in emps) or '—'
        rows.append({
            'sl':            idx,
            'date':          log.date.strftime('%d-%m-%Y'),
            'location_name': log.location.name if log.location else '—',
            'site':          log.site or '—',
            'work_details':  log.work_details or '—',
            'tmp':           emp_count,
            'employees':     emp_names,
            'remarks':       log.remarks or '—',
        })

    locations = list(MachineLocation.objects.filter(admin_id=admin_id).values('id', 'name'))

    return JsonResponse({
        'rows':          rows,
        'total_tmp':     total_tmp,
        'total_entries': len(rows),
        'locations':     locations,
    })


@login_required
@require_POST
def work_log_add(request):
    """AJAX: legacy create endpoint — now routes through the same upsert
    as `work_log_upsert` so the (admin_id, date, location) uniqueness and
    the locked-on-save invariant hold for every write path."""
    try:
        data         = json.loads(request.body)
        admin_id     = get_admin_id(request.user)
        date_str     = (data.get('date') or '').strip()
        location_id  = data.get('location_id') or None
        site         = (data.get('site') or '').strip()
        work_details = (data.get('work_details') or '').strip()
        tmp          = int(data.get('tmp') or 0)
        company_name = (data.get('company_name') or '').strip()
        remarks      = (data.get('remarks') or '').strip()

        if not date_str:
            return JsonResponse({'success': False, 'error': 'Date is required.'})

        location = None
        if location_id:
            location = get_object_or_404(MachineLocation, pk=location_id, admin_id=admin_id)

        existing = WorkLog.objects.filter(
            admin_id=admin_id, date=date_str, location=location,
        ).only('id', 'locked').first()
        if existing and getattr(existing, 'locked', False):
            return JsonResponse({
                'success': False,
                'error':   'Entry is locked. Unlock to edit.',
                'locked':  True,
                'id':      existing.pk,
            })

        log, _created = WorkLog.objects.update_or_create(
            admin_id=admin_id,
            date=date_str,
            location=location,
            defaults={
                'site':         site,
                'work_details': work_details,
                'tmp':          tmp,
                'company_name': company_name,
                'remarks':      remarks,
                'created_by':   request.user,
                'locked':       True,
            },
        )
        return JsonResponse({'success': True, 'id': log.pk})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def work_log_edit(request, pk):
    """AJAX: update an existing WorkLog entry."""
    try:
        admin_id    = get_admin_id(request.user)
        log         = get_object_or_404(WorkLog, pk=pk, admin_id=admin_id)
        if getattr(log, 'locked', False):
            return JsonResponse({'success': False, 'error': 'Entry is locked. Unlock to edit.', 'locked': True})
        data        = json.loads(request.body)
        location_id = data.get('location_id') or None
        location    = None
        if location_id:
            location = get_object_or_404(MachineLocation, pk=location_id, admin_id=admin_id)

        log.date         = data.get('date', log.date)
        log.location     = location
        log.site         = (data.get('site') or '').strip()
        log.work_details = (data.get('work_details') or '').strip()
        log.tmp          = int(data.get('tmp') or 0)
        log.company_name = (data.get('company_name') or '').strip()
        log.remarks      = (data.get('remarks') or '').strip()
        log.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def work_log_delete(request, pk):
    """Delete a WorkLog entry via AJAX or form POST."""
    admin_id = get_admin_id(request.user)
    log = get_object_or_404(WorkLog, pk=pk, admin_id=admin_id)
    if getattr(log, 'locked', False):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'Entry is locked. Unlock to delete.', 'locked': True})
        return redirect('projects:list')
    log.delete()
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True})
    return redirect('projects:list')


@login_required
@require_POST
def work_log_upsert(request):
    """AJAX: create or update a WorkLog entry for a specific machine + date (attendance-style save)."""
    try:
        data        = json.loads(request.body)
        admin_id    = get_admin_id(request.user)
        date_str    = (data.get('date') or '').strip()
        location_id = data.get('location_id') or None

        if not date_str:
            return JsonResponse({'success': False, 'error': 'Date is required.'})

        try:
            entry_date = date_type.fromisoformat(date_str)
        except ValueError:
            return JsonResponse({'success': False, 'error': 'Invalid date format.'})

        if entry_date > date_type.today():
            return JsonResponse({'success': False, 'error': 'Future dates are not allowed.'})

        location = None
        if location_id:
            location = get_object_or_404(MachineLocation, pk=location_id, admin_id=admin_id)

        site         = (data.get('site') or '').strip()
        work_details = (data.get('work_details') or '').strip()
        tmp          = int(data.get('tmp') or 0)
        remarks      = (data.get('remarks') or '').strip()
        employee_ids = data.get('employee_ids') or []

        # Backend TMP enforcement
        if tmp > 0 and len(employee_ids) > tmp:
            return JsonResponse({
                'success': False,
                'error': f'Employee count ({len(employee_ids)}) exceeds TMP limit ({tmp}).',
            })

        # Machine + date lock: once an entry exists and is locked, block
        # further save/edit until the admin explicitly unlocks it.
        existing = WorkLog.objects.filter(
            admin_id=admin_id, date=date_str, location=location,
        ).only('id', 'locked').first()
        if existing and getattr(existing, 'locked', False):
            return JsonResponse({
                'success': False,
                'error': 'Entry is locked. Unlock to edit.',
                'locked': True,
                'id': existing.pk,
            })

        log, created = WorkLog.objects.update_or_create(
            admin_id=admin_id,
            date=date_str,
            location=location,
            defaults={
                'site':         site,
                'work_details': work_details,
                'tmp':          tmp,
                'remarks':      remarks,
                'created_by':   request.user,
                'locked':       True,
            },
        )

        if employee_ids:
            emps = Employee.objects.filter(pk__in=employee_ids, admin_id=admin_id)
            log.employees.set(emps)
        else:
            log.employees.clear()

        return JsonResponse({'success': True, 'id': log.pk, 'created': created})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def work_log_unlock(request, pk):
    """Clear the locked flag on a WorkLog row so the admin can edit/delete."""
    try:
        admin_id = get_admin_id(request.user)
        log = get_object_or_404(WorkLog, pk=pk, admin_id=admin_id)
        log.locked = False
        log.save(update_fields=['locked', 'updated_at'])
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def add_work_status(request):
    """AJAX: add a new WorkStatus option (e.g. Painting, Greasing)."""
    try:
        data     = json.loads(request.body)
        admin_id = get_admin_id(request.user)
        name     = (data.get('name') or '').strip()
        if not name:
            return JsonResponse({'success': False, 'error': 'Status name is required.'})
        ws, created = WorkStatus.objects.get_or_create(admin_id=admin_id, name=name)
        return JsonResponse({'success': True, 'id': ws.pk, 'name': ws.name, 'created': created})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def add_machine_location(request):
    """AJAX: add a new machine number to the MachineLocation list (module-local)."""
    try:
        data     = json.loads(request.body)
        admin_id = get_admin_id(request.user)
        name     = (data.get('name') or '').strip()
        if not name:
            return JsonResponse({'success': False, 'error': 'Machine number is required.'})
        loc, created = MachineLocation.objects.get_or_create(admin_id=admin_id, name=name)
        return JsonResponse({'success': True, 'id': loc.pk, 'name': loc.name, 'created': created})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def machine_monthly_report(request):
    """
    Return machine-wise monthly work-attendance summary as JSON.
    Used by the Work Summary tab to generate downloadable PDF / XLSX reports.

    Required params: year, month, location (MachineLocation pk)
    Optional params: site (text filter)
    """
    import calendar
    from dashboard.models import CompanySettings

    admin_id  = get_admin_id(request.user)
    year_str  = request.GET.get('year',  '').strip()
    month_str = request.GET.get('month', '').strip()
    loc_id    = request.GET.get('location', '').strip()
    site_flt  = request.GET.get('site', '').strip()

    if not (year_str and month_str and loc_id):
        return JsonResponse({'success': False, 'error': 'Year, month, and location are required.'})

    try:
        year  = int(year_str)
        month = int(month_str)
        if not (1 <= month <= 12) or year < 2000:
            raise ValueError
    except ValueError:
        return JsonResponse({'success': False, 'error': 'Invalid year or month.'})

    days_in_month = calendar.monthrange(year, month)[1]

    location = get_object_or_404(MachineLocation, pk=loc_id, admin_id=admin_id)

    qs = (
        WorkLog.objects
        .filter(admin_id=admin_id, location=location,
                date__year=year, date__month=month)
        .prefetch_related('employees')
        .order_by('date')
    )
    if site_flt:
        qs = qs.filter(site__icontains=site_flt)

    # day (1–31) → {emp_id_set, work_details, site}
    # TMP is derived from len(emp_ids) at output time below — the stored
    # `tmp` column is mobile-inserted as 0 and no longer reliable.
    day_map = {}
    # ordered dict: emp_id → emp_name  (insertion order = first appearance)
    emp_registry = {}

    for log in qs:
        d = log.date.day
        if d not in day_map:
            day_map[d] = {
                'emp_ids':     set(),
                'work_details': log.work_details or '',
                'site':        log.site or '',
            }
        # Same attendance-aware gate as _work_summary_json: only count
        # employees whose attendance for that date is present/half_day
        # (or absent — i.e. no record yet — permissive default).
        qualified = _attendance_qualified_emps(log.employees.all(), log.date)
        day_map[d]['emp_ids'].update(e.pk for e in qualified)
        day_map[d]['work_details']= log.work_details or day_map[d]['work_details']
        for e in qualified:
            if e.pk not in emp_registry:
                emp_registry[e.pk] = e.name

    # Site header value
    site_header = site_flt or (day_map[min(day_map)]['site'] if day_map else '')

    # Build employee rows: each row has 'cells' list of '' or '✓' for days 1–N
    emp_rows = []
    for sl, (eid, ename) in enumerate(emp_registry.items(), start=1):
        cells = [
            '✓' if (d in day_map and eid in day_map[d]['emp_ids']) else ''
            for d in range(1, days_in_month + 1)
        ]
        emp_rows.append({'sl': sl, 'emp_id': eid, 'name': ename, 'cells': cells})

    # TMP totals row (one value per day) — derived from the unique linked
    # employee count for that day so it stays in sync with employee_names.
    tmp_row = [
        len(day_map[d]['emp_ids']) if d in day_map else ''
        for d in range(1, days_in_month + 1)
    ]

    # Work-details row (one value per day, for reference)
    wd_row = [
        day_map[d]['work_details'] if d in day_map else ''
        for d in range(1, days_in_month + 1)
    ]

    company     = CompanySettings.get_settings(admin_id)
    month_names = [
        'January','February','March','April','May','June',
        'July','August','September','October','November','December',
    ]

    return JsonResponse({
        'success':       True,
        'company_name':  company.name if company else '',
        'location':      location.name,
        'site':          site_header,
        'year':          year,
        'month':         month,
        'month_name':    month_names[month - 1],
        'days_in_month': days_in_month,
        'dates':         list(range(1, days_in_month + 1)),
        'emp_rows':      emp_rows,
        'tmp_row':       tmp_row,
        'wd_row':        wd_row,
    })


# ─── Work Summary export (Task 5 / 2026-05-24) ───────────────────────────────
# Produces XLSX and PDF in the exact layout of the customer-supplied template:
#   Row 1: "Location  (machine nu)" | <machine#> | … | "Month" | <month>
#   Row 2: "Site "                  | <site>
#   Row 3: (blank)(blank) "Date "   repeated across day columns
#   Row 4: "Sl.no" | "Name/Emp Id"  | 1 .. N  (calendar-month days)
#   Row 5..: <sl> | <name>          | "Present" or "" per day
# ─────────────────────────────────────────────────────────────────────────────

def _build_work_summary_matrix(admin_id, year, month, location, site_flt):
    """Shared compute path for the export endpoints — mirrors the JSON view."""
    import calendar
    days_in_month = calendar.monthrange(year, month)[1]
    qs = (
        WorkLog.objects
        .filter(admin_id=admin_id, location=location,
                date__year=year, date__month=month)
        .prefetch_related('employees')
        .order_by('date')
    )
    if site_flt:
        qs = qs.filter(site__icontains=site_flt)

    day_map      = {}
    emp_registry = {}
    for log in qs:
        d = log.date.day
        if d not in day_map:
            day_map[d] = {'emp_ids': set(), 'site': log.site or ''}
        # Mirror the attendance gate used by the JSON summary endpoints so
        # XLSX / PDF exports show the same employee set as the screen view.
        qualified = _attendance_qualified_emps(log.employees.all(), log.date)
        day_map[d]['emp_ids'].update(e.pk for e in qualified)
        for e in qualified:
            if e.pk not in emp_registry:
                emp_registry[e.pk] = (e.employee_id or '', e.name)

    site_header = site_flt or (day_map[min(day_map)]['site'] if day_map else '')
    rows = []
    for sl, (eid, (emp_code, ename)) in enumerate(emp_registry.items(), start=1):
        cells = [
            'Present' if (d in day_map and eid in day_map[d]['emp_ids']) else ''
            for d in range(1, days_in_month + 1)
        ]
        # "Name/Emp Id" column: name + emp code if known
        label = f"{ename} ({emp_code})" if emp_code else ename
        rows.append({'sl': sl, 'name': label, 'cells': cells})

    return days_in_month, site_header, rows


@login_required
def machine_monthly_report_export(request):
    """
    GET /projects/report/machine-monthly/export/?year=&month=&location=&site=&format=xlsx|pdf
    Returns the work summary as a downloadable file using the customer template.
    """
    from django.http import HttpResponse
    from io import BytesIO

    admin_id  = get_admin_id(request.user)
    fmt       = (request.GET.get('format') or 'xlsx').lower()
    year_str  = request.GET.get('year',  '').strip()
    month_str = request.GET.get('month', '').strip()
    loc_id    = request.GET.get('location', '').strip()
    site_flt  = request.GET.get('site', '').strip()

    if not (year_str and month_str and loc_id):
        return HttpResponse('Year, month, and location are required.', status=400)
    try:
        year  = int(year_str)
        month = int(month_str)
        if not (1 <= month <= 12) or year < 2000:
            raise ValueError
    except ValueError:
        return HttpResponse('Invalid year or month.', status=400)

    location = get_object_or_404(MachineLocation, pk=loc_id, admin_id=admin_id)
    days_in_month, site_header, rows = _build_work_summary_matrix(
        admin_id, year, month, location, site_flt,
    )
    month_names = ['January','February','March','April','May','June',
                   'July','August','September','October','November','December']
    month_name = month_names[month - 1]
    fname_base = f"work-summary-{location.name}-{year}-{month:02d}"

    if fmt == 'xlsx':
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = 'Work Summary'

        # Row 1: Location (machine nu) | <machine> | ... | Month | <month_name>
        ws.cell(row=1, column=1, value='Location  (machine nu)').font = Font(bold=True)
        ws.cell(row=1, column=2, value=location.name)
        ws.cell(row=1, column=6, value='Month').font = Font(bold=True)
        ws.cell(row=1, column=7, value=month_name)
        # Row 2: Site
        ws.cell(row=2, column=1, value='Site ').font = Font(bold=True)
        ws.cell(row=2, column=2, value=site_header)
        # Row 3: Date header label across day columns
        for i in range(days_in_month):
            ws.cell(row=3, column=3 + i, value='Date ').font = Font(bold=True)
        # Row 4: Sl.no | Name/Emp Id | day numbers
        ws.cell(row=4, column=1, value='Sl.no').font = Font(bold=True)
        ws.cell(row=4, column=2, value='Name/Emp Id').font = Font(bold=True)
        for i in range(days_in_month):
            ws.cell(row=4, column=3 + i, value=i + 1).font = Font(bold=True)
        # Rows 5+: employee rows
        thin   = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for ri, r in enumerate(rows, start=5):
            ws.cell(row=ri, column=1, value=r['sl'])
            ws.cell(row=ri, column=2, value=r['name'])
            for ci, val in enumerate(r['cells']):
                ws.cell(row=ri, column=3 + ci, value=val)
        # Light border + sizing
        last_row = max(4, 4 + len(rows))
        last_col = 2 + days_in_month
        for rr in range(1, last_row + 1):
            for cc in range(1, last_col + 1):
                ws.cell(row=rr, column=cc).border = border
        ws.column_dimensions[get_column_letter(1)].width = 7
        ws.column_dimensions[get_column_letter(2)].width = 28
        for cc in range(3, last_col + 1):
            ws.column_dimensions[get_column_letter(cc)].width = 10

        buf = BytesIO()
        wb.save(buf)
        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="{fname_base}.xlsx"'
        return resp

    if fmt == 'pdf':
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A3, landscape
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
        buf = BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=landscape(A3),
            leftMargin=1*cm, rightMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm,
        )

        # Build table data mirroring the XLSX layout.
        header_top = ['Location  (machine nu)', location.name] + ['']*(days_in_month + 0 - 4) + ['Month', month_name]
        # Pad to total columns = 2 + days_in_month
        total_cols = 2 + days_in_month
        if len(header_top) < total_cols:
            header_top += [''] * (total_cols - len(header_top))
        else:
            header_top = header_top[:total_cols]

        site_row = ['Site ', site_header] + [''] * (total_cols - 2)
        date_label_row = ['', ''] + ['Date '] * days_in_month
        day_num_row    = ['Sl.no', 'Name/Emp Id'] + list(range(1, days_in_month + 1))
        data = [header_top, site_row, date_label_row, day_num_row]
        for r in rows:
            data.append([r['sl'], r['name']] + r['cells'])

        # Column widths: small for sl, wide for name, even for days
        col_widths = [1.2*cm, 5*cm] + [1.1*cm] * days_in_month
        t = Table(data, colWidths=col_widths, repeatRows=4)
        t.setStyle(TableStyle([
            ('FONTNAME',    (0,0), (-1,-1), 'Helvetica'),
            ('FONTSIZE',    (0,0), (-1,-1), 8),
            ('FONTNAME',    (0,0), (-1,3),  'Helvetica-Bold'),
            ('BACKGROUND',  (0,3), (-1,3),  colors.HexColor('#F1F5F9')),
            ('GRID',        (0,0), (-1,-1), 0.3, colors.HexColor('#CBD5E1')),
            ('ALIGN',       (2,3), (-1,-1), 'CENTER'),
            ('VALIGN',      (0,0), (-1,-1), 'MIDDLE'),
        ]))
        doc.build([t])
        resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="{fname_base}.pdf"'
        return resp

    return HttpResponse('Unsupported format. Use ?format=xlsx or pdf.', status=400)


# ═══════════════════════════════════════════════════════════════════════════
# [TEMP-DEBUG-WORKLOG-PAGE]  Admin-only diagnostic — remove after RCA.
# Route: /projects/debug/worklog/
# ═══════════════════════════════════════════════════════════════════════════
@login_required
def _debug_worklog(request):
    from django.http import HttpResponse
    from django.utils.html import escape as _esc
    from datetime import date as _date_type
    import json as _json

    if not getattr(request.user, 'is_admin', False):
        return HttpResponse('Forbidden', status=403)

    admin_id = get_admin_id(request.user)

    q_date = (request.GET.get('date') or '').strip()
    q_loc  = (request.GET.get('location_id') or '').strip()

    machines_all = list(
        MachineLocation.objects.all().values('id', 'admin_id', 'name').order_by('admin_id', 'name')
    )

    def _fmt(v):
        if v is None:
            return '<em style="color:#888">NULL</em>'
        return _esc(str(v))

    body = []
    body.append('<!doctype html><meta charset="utf-8"><title>WorkLog Debug</title>')
    body.append('<style>body{font-family:monospace;padding:1rem;background:#0f172a;color:#e2e8f0}'
                'table{border-collapse:collapse;margin:0.5rem 0;background:#1e293b}'
                'th,td{border:1px solid #334155;padding:.35rem .6rem;vertical-align:top;font-size:12px}'
                'th{background:#334155;text-align:left}'
                'h2{margin-top:1.6rem;color:#93c5fd;border-bottom:1px solid #334155;padding-bottom:.25rem}'
                'code,pre{background:#020617;color:#a5f3fc;padding:.15rem .3rem;border-radius:3px}'
                'pre{padding:.75rem;white-space:pre-wrap;word-break:break-all}'
                '.warn{color:#fca5a5;font-weight:bold}'
                '.ok{color:#86efac}'
                'form{background:#1e293b;padding:1rem;border-radius:6px;margin-bottom:1rem}'
                'input,select,button{padding:.4rem .6rem;background:#0f172a;color:#e2e8f0;border:1px solid #475569;border-radius:4px;font-family:inherit}'
                'button{background:#3b82f6;cursor:pointer;border-color:#3b82f6}'
                '</style>')
    body.append('<h1>WorkLog Debug — TEMPORARY</h1>')
    body.append(f'<p><strong>request.user</strong>: pk={request.user.pk}, email={_esc(getattr(request.user,"email","") or "")}, '
                f'role={_esc(getattr(request.user,"role","") or "")}, '
                f'user.admin_id (field)={_fmt(getattr(request.user,"admin_id", None))}<br>'
                f'<strong>get_admin_id(request.user)</strong> = <code>{_esc(admin_id)}</code></p>')

    # Form
    body.append('<form method="get">')
    body.append(f'<label>Date <input type="date" name="date" value="{_esc(q_date)}" required></label> &nbsp; ')
    body.append('<label>Machine <select name="location_id" required>')
    body.append('<option value="">— select —</option>')
    for m in machines_all:
        sel = ' selected' if str(m['id']) == q_loc else ''
        body.append(f'<option value="{m["id"]}"{sel}>#{m["id"]} — {_esc(m["name"])} '
                    f'(admin_id={_esc(m["admin_id"])})</option>')
    body.append('</select></label> &nbsp; <button type="submit">Diagnose</button>')
    body.append('</form>')

    if not (q_date and q_loc):
        return HttpResponse(''.join(body))

    # === A. current admin_id ===
    body.append('<h2>A. Current admin_id</h2>')
    body.append(f'<pre>{_esc(admin_id)}</pre>')

    # === B. MachineLocation ===
    try:
        loc_pk = int(q_loc)
    except (TypeError, ValueError):
        body.append('<p class="warn">Invalid location_id.</p>')
        return HttpResponse(''.join(body))

    m_obj = MachineLocation.objects.filter(pk=loc_pk).first()
    body.append('<h2>B. MachineLocation</h2>')
    if not m_obj:
        body.append(f'<p class="warn">MachineLocation pk={loc_pk} DOES NOT EXIST.</p>')
    else:
        body.append('<table><tr><th>id</th><th>name</th><th>admin_id</th>'
                    '<th>tenant match with request?</th></tr>')
        tenant_match = (m_obj.admin_id == admin_id)
        cls = 'ok' if tenant_match else 'warn'
        body.append(f'<tr><td>{m_obj.pk}</td><td>{_esc(m_obj.name)}</td>'
                    f'<td>{_esc(m_obj.admin_id)}</td>'
                    f'<td class="{cls}">{"YES" if tenant_match else "NO — cross-tenant"}</td></tr></table>')
        siblings = MachineLocation.objects.filter(name=m_obj.name).exclude(pk=m_obj.pk)
        if siblings.exists():
            body.append('<p class="warn">Other MachineLocation rows with the same name:</p><table>'
                        '<tr><th>id</th><th>admin_id</th><th>name</th></tr>')
            for s in siblings:
                body.append(f'<tr><td>{s.pk}</td><td>{_esc(s.admin_id)}</td><td>{_esc(s.name)}</td></tr>')
            body.append('</table>')

    # === C. Every WorkLog row for (admin_id, date, location_id=loc_pk) ===
    body.append('<h2>C. WorkLog rows for (current admin_id, chosen date, chosen location)</h2>')
    match_qs = (
        WorkLog.objects
        .filter(admin_id=admin_id, date=q_date, location_id=loc_pk)
        .prefetch_related('employees')
        .order_by('pk')
    )
    match_count = match_qs.count()
    body.append(f'<p>Row count = <code>{match_count}</code></p>')
    if match_count:
        body.append('<table><tr>'
                    '<th>pk</th><th>admin_id</th><th>location_id</th><th>date</th>'
                    '<th>site</th><th>work_details</th><th>tmp</th><th>remarks</th>'
                    '<th>locked</th><th>created_at</th><th>updated_at</th><th>employees</th>'
                    '</tr>')
        for w in match_qs:
            emps = list(w.employees.values_list('pk', 'employee_id', 'name'))
            body.append('<tr>'
                        f'<td>{w.pk}</td>'
                        f'<td>{_esc(w.admin_id)}</td>'
                        f'<td>{_fmt(w.location_id)}</td>'
                        f'<td>{_esc(str(w.date))}</td>'
                        f'<td>{_esc(w.site)}</td>'
                        f'<td>{_esc(w.work_details)}</td>'
                        f'<td>{w.tmp}</td>'
                        f'<td>{_esc(w.remarks)}</td>'
                        f'<td>{_fmt(getattr(w,"locked","MISSING"))}</td>'
                        f'<td>{_esc(str(w.created_at))}</td>'
                        f'<td>{_esc(str(w.updated_at))}</td>'
                        f'<td>{_esc(str(emps))}</td>'
                        '</tr>')
        body.append('</table>')

    # Also list every WorkLog for this date across all locations and admin_ids
    body.append('<h2>C-extra. All WorkLog rows on this date (any admin_id, any location)</h2>')
    all_on_date = (
        WorkLog.objects.filter(date=q_date).prefetch_related('employees').order_by('pk')
    )
    body.append(f'<p>Row count on {_esc(q_date)} = <code>{all_on_date.count()}</code></p>')
    if all_on_date.exists():
        body.append('<table><tr>'
                    '<th>pk</th><th>admin_id</th><th>location_id</th><th>site</th>'
                    '<th>work_details</th><th>locked</th><th>created_at</th></tr>')
        for w in all_on_date:
            row_tenant = 'ok' if w.admin_id == admin_id else 'warn'
            row_loc    = 'ok' if w.location_id == loc_pk else 'warn'
            body.append('<tr>'
                        f'<td>{w.pk}</td>'
                        f'<td class="{row_tenant}">{_esc(w.admin_id)}</td>'
                        f'<td class="{row_loc}">{_fmt(w.location_id)}</td>'
                        f'<td>{_esc(w.site)}</td>'
                        f'<td>{_esc(w.work_details)}</td>'
                        f'<td>{_fmt(getattr(w,"locked","MISSING"))}</td>'
                        f'<td>{_esc(str(w.created_at))}</td>'
                        '</tr>')
        body.append('</table>')

    # === D. Exact JSON that _work_log_json would return ===
    body.append('<h2>D. Exact JSON _work_log_json returns for this date</h2>')
    try:
        from django.test import RequestFactory as _RF
        _rf = _RF()
        _fake = _rf.get('/projects/', {'format': 'json', 'tab': 'log', 'date': q_date})
        _fake.user = request.user
        _fake.headers = {'X-Requested-With': 'XMLHttpRequest'}
        _resp = _work_log_json(_fake, admin_id)
        _payload = _json.loads(_resp.content.decode())
    except Exception as _e:
        _payload = {'ERROR': str(_e)}
    body.append(f'<pre>{_esc(_json.dumps(_payload, indent=2, default=str))}</pre>')

    entries_for_loc = [e for e in _payload.get('entries', []) if e.get('location_id') == loc_pk]
    body.append(f'<p>Entries in payload matching location_id=<code>{loc_pk}</code>: '
                f'<code>{len(entries_for_loc)}</code></p>')

    # === E. Duplicate detection ===
    body.append('<h2>E. Duplicate rows for (admin_id, date, location_id)</h2>')
    if match_count > 1:
        body.append(f'<p class="warn">DUPLICATES DETECTED — {match_count} rows share this key.</p>')
    else:
        body.append('<p class="ok">No duplicates for this key.</p>')

    # === F. Which row _work_log_json selects as canonical ===
    body.append('<h2>F. Canonical row picked by _work_log_json</h2>')
    if entries_for_loc:
        body.append(f'<pre>{_esc(_json.dumps(entries_for_loc[0], indent=2, default=str))}</pre>')
    else:
        body.append('<p class="warn">_work_log_json returned NO entry for this location_id.</p>')

    # === G. Locked state ===
    body.append('<h2>G. Locked state of the matching row</h2>')
    if match_qs.exists():
        for w in match_qs:
            body.append(f'<p>pk={w.pk} → locked=<code>{_fmt(getattr(w,"locked","MISSING"))}</code></p>')
    else:
        body.append('<p>No matching row.</p>')

    # === H. Orphan rows (location_id IS NULL) on this date ===
    body.append('<h2>H. Orphan WorkLogs on this date (location_id IS NULL)</h2>')
    orphans = WorkLog.objects.filter(date=q_date, location__isnull=True)
    body.append(f'<p>count = <code>{orphans.count()}</code></p>')
    if orphans.exists():
        body.append('<table><tr><th>pk</th><th>admin_id</th><th>site</th><th>work_details</th>'
                    '<th>remarks</th><th>tmp</th><th>locked</th><th>created_at</th></tr>')
        for w in orphans:
            body.append('<tr>'
                        f'<td>{w.pk}</td><td>{_esc(w.admin_id)}</td><td>{_esc(w.site)}</td>'
                        f'<td>{_esc(w.work_details)}</td><td>{_esc(w.remarks)}</td>'
                        f'<td>{w.tmp}</td><td>{_fmt(getattr(w,"locked","MISSING"))}</td>'
                        f'<td>{_esc(str(w.created_at))}</td></tr>')
        body.append('</table>')

    # === I. Queryset count from the exact filter _work_log_json uses ===
    body.append('<h2>I. Exact queryset count used by _work_log_json (admin_id + date filter only)</h2>')
    daily_qs = WorkLog.objects.filter(admin_id=admin_id, date=q_date)
    body.append(f'<p>count = <code>{daily_qs.count()}</code></p>')
    if daily_qs.exists():
        body.append('<table><tr><th>pk</th><th>location_id</th><th>site</th>'
                    '<th>work_details</th><th>locked</th><th>created_at</th></tr>')
        for w in daily_qs.order_by('-created_at'):
            body.append('<tr>'
                        f'<td>{w.pk}</td><td>{_fmt(w.location_id)}</td><td>{_esc(w.site)}</td>'
                        f'<td>{_esc(w.work_details)}</td>'
                        f'<td>{_fmt(getattr(w,"locked","MISSING"))}</td>'
                        f'<td>{_esc(str(w.created_at))}</td></tr>')
        body.append('</table>')

    # === J. What Work Summary would show for this date ===
    body.append('<h2>J. Rows Work Summary would display for date_from=date_to=this date</h2>')
    try:
        _fake2 = _rf.get('/projects/', {'format': 'json', 'tab': 'summary',
                                        'date_from': q_date, 'date_to': q_date})
        _fake2.user = request.user
        _fake2.headers = {'X-Requested-With': 'XMLHttpRequest'}
        _resp2 = _work_summary_json(_fake2, admin_id)
        _sum_payload = _json.loads(_resp2.content.decode())
    except Exception as _e:
        _sum_payload = {'ERROR': str(_e)}
    body.append(f'<pre>{_esc(_json.dumps(_sum_payload, indent=2, default=str))}</pre>')

    sum_rows_for_loc = [
        r for r in _sum_payload.get('rows', [])
        if m_obj and r.get('location_name') == m_obj.name
    ]
    body.append(f'<p>Summary rows matching this location by name: '
                f'<code>{len(sum_rows_for_loc)}</code></p>')

    return HttpResponse(''.join(body))

